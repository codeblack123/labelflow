import sys
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse
import pandas as pd
import fitz  # PyMuPDF
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io
import re
import base64
import os
import shutil
import glob
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List
from uuid import UUID
import urllib.request
import urllib.parse
from openpyxl import load_workbook
import zipfile
import math
import json


# --- FONT REGISTRATION ---
def register_fonts():
    try:
        # Standard paths for Windows fonts
        font_paths = {
            "Bahnschrift": "C:/Windows/Fonts/bahnschrift.ttf",
            # Add more fonts here if needed
        }
        
        for name, path in font_paths.items():
            if os.path.exists(path):
                pdfmetrics.registerFont(TTFont(name, path))
                print(f"[FONT] Registered: {name} from {path}")
            else:
                # Fallback to similar available fonts or just log it
                print(f"[FONT] Warning: {name} not found at {path}")
    except Exception as e:
        print(f"[FONT] Error registering fonts: {e}")

register_fonts()

app = FastAPI()

@app.get("/")
async def root():
    return {"status": "ok", "message": "Local backend is running"}

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- SKU MAPPING STORAGE (SUPABASE DIRECT) ---
import httpx
from pydantic import BaseModel

# Supabase Config
SUPABASE_URL = "https://lcexnrzqtyrixpuvifxg.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImxjZXhucnpxdHlyaXhwdXZpZnhnIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Njk3OTIxNTgsImV4cCI6MjA4NTM2ODE1OH0.HVtoklr7Y--yiYWLgfDA1M2qjR_xt7ihtDZoOR4IP5U"

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal"
}

class SkuMappingItem(BaseModel):
    id: str
    sku: str
    rak: str = ""

async def supabase_fetch(method: str, endpoint: str, data=None, params=None, headers=None):
    async with httpx.AsyncClient(timeout=60.0) as client:
        url = f"{SUPABASE_URL}/rest/v1/{endpoint}"
        req_headers = HEADERS.copy()
        
        # Merge custom headers if provided
        if headers:
            req_headers.update(headers)
        
        # If we expect a response (GET or explicit select), verify Prefer header
        if method == "GET":
            # GET always returns representation
            pass
        elif "select" in (params or {}):
            # If we want to select fields after POST/PATCH, we need representation
            req_headers["Prefer"] = "return=representation"
            
        try:
            response = await client.request(method, url, headers=req_headers, json=data, params=params)
            
            if response.status_code >= 400:
                error_detail = response.text
                try:
                    err_json = response.json()
                    error_detail = err_json.get('message', response.text)
                except:
                    pass
                print(f"[Supabase Error] {method} {endpoint} -> {response.status_code}: {error_detail}")
                raise HTTPException(status_code=response.status_code, detail=f"Database: {error_detail}")

            # GET returns list, POST/DELETE might differ based on Prefer header
            if (method == "GET" or req_headers.get("Prefer") == "return=representation") and response.text:
                 try:
                     return response.json()
                 except:
                     return response.text
            return response.text
        except httpx.RequestError as e:
            print(f"Supabase Request Error: {e}")
            raise HTTPException(status_code=502, detail="Gagal menghubungi Supabase")
        except HTTPException:
            raise
        except Exception as e:
            print(f"Internal Fetch Error: {e}")
            raise HTTPException(status_code=500, detail=str(e))

async def fetch_sku_mappings(limit: int = None, offset: int = 0, search: str = None, order_by: str = "custom_id", order_dir: str = "asc", is_multi_search: bool = False):
    # Helper to fetch SKU mappings (Internal Use)
    # If limit is None, fetches ALL data (handling 1000 row limit of Supabase)
    
    try:
        # Validasi order_by untuk keamanan (menghindari SQL injection via URL)
        safe_cols = ["custom_id", "sku", "rak"]
        if order_by not in safe_cols:
            order_by = "custom_id"
            
        def apply_search(p: dict):
            if not search: return
            if is_multi_search:
                import re
                terms = [t.strip() for t in re.split(r'[\n,]+', search) if t.strip()]
                if terms:
                    term_str = ",".join(terms)
                    p["or"] = f"(sku.in.({term_str}),custom_id.in.({term_str}))"
            else:
                p["or"] = f"(sku.ilike.%{search}%,custom_id.ilike.%{search}%)"
        
        if limit is None:
            # Fetch ALL
            all_data = []
            chunk_size = 1000
            current_offset = 0
            while True:
                params = {
                    "select": "custom_id,sku,rak",
                    "limit": chunk_size,
                    "offset": current_offset,
                    "order": f"{order_by}.{order_dir}"
                }
                
                
                apply_search(params)
                    
                chunk = await supabase_fetch("GET", "sku_mappings", params=params)
                if not chunk:
                    break
                all_data.extend(chunk)
                if len(chunk) < chunk_size:
                    break
                current_offset += chunk_size
            return [{"id": item["custom_id"], "sku": item["sku"], "rak": item.get("rak", "")} for item in all_data]
            
        else:
            # Fetch Paginated
            params = {
                "select": "custom_id,sku,rak",
                "limit": limit,
                "offset": offset,
                "order": f"{order_by}.{order_dir}"
            }
            
            apply_search(params)
            
            data = await supabase_fetch("GET", "sku_mappings", params=params)
            if not isinstance(data, list):
                return []
            return [{"id": item["custom_id"], "sku": item["sku"], "rak": item.get("rak", "")} for item in data]
            
    except Exception as e:
        print(f"Fetch SKU Error: {e}")
        import traceback
        traceback.print_exc()
        return []

async def count_sku_mappings(search: str = None):
    # Helper to count total
    endpoint = "sku_mappings?select=custom_id&limit=1" # Minimal fetch, rely on content-range header? 
    # Current supabase_fetch doesn't return headers.
    # Workaround: Use head=true logic or count=exact prefer header.
    # Our supabase_fetch implementation is simple.
    # Let's use a separate fetch with 'count=exact' Header if we can modify supabase_fetch, 
    # OR since we don't want to touch supabase_fetch too much, use a 'select count' if possible? No PostgREST doesn't support 'select=count'.
    
    # We will assume for now we can't easily get count without fetching all ID's or modifying supabase_fetch.
    # For performance, maybe just fetching IDs is light enough?
    # Or modify supabase_fetch to support returning count.
    
    # Let's try fetching just IDs with search to count?
    # If dataset is 50k, fetching 50k IDs is ~2MB. Acceptable? Maybe.
    # Better: Update supabase_fetch later. For now, let's use the 'Fetch All IDs' strategy for counting if search is active.
    # If search inactive, we might cache the count?
    
    # Actually, let's just use the `fetch_sku_mappings(limit=None)` on a valid search and count the length.
    # It's not optimal but it works without rewriting core utils.
    
    # WAIT! PostgREST supports HEAD request for count.
    # supabase_fetch logic: method="HEAD".
    # But `supabase_fetch` returns `response.json()` or `text`. Head has no body.
    # We need headers 'Content-Range'.
    
    # Let's stick to "Fetch All" for "Fetch All" endpoint.
    # For Paginated Endpoint, if we need count, we might have to be expensive for now or skip count.
    # User asked for pagination. Next/Prev is fine without "Page 1 of 100".
    # But "Page 1 of X" is better.
    
    # Let's implement a simple count by fetching all IDs (lightweight) if search is present?
    # Or just returning the page.
    pass

@app.get("/settings/sku-mappings")
async def get_sku_mappings():
    # Legacy Endpoint (returns ALL) - Enhanced to actually fetch ALL (>1000)
    return await fetch_sku_mappings(limit=None)

@app.get("/settings/sku-mappings-paginated")
async def get_sku_mappings_paginated(
    page: int = 1, 
    limit: int = 50, 
    search: str = "", 
    order_by: str = "custom_id", 
    order_dir: str = "asc",
    is_multi_search: str = "false"
):
    offset = (page - 1) * limit
    data = await fetch_sku_mappings(
        limit=limit, 
        offset=offset, 
        search=search if search else None,
        order_by=order_by,
        order_dir=order_dir,
        is_multi_search=(is_multi_search.lower() == 'true')
    )
    
    print(f"[DEBUG] Paginated SKU: Page {page}, Limit {limit}, Search {search}, Order {order_by} {order_dir} -> Found {len(data)} rows")
    return {
        "data": data,
        "page": page,
        "limit": limit
    }

@app.post("/settings/sku-mappings")
async def add_sku_mapping(item: SkuMappingItem):
    # Optimization: Check specific ID and SKU instead of fetching all 10k+ rows
    try:
        # 1. Check if ID exists
        check_id = await supabase_fetch("GET", f"sku_mappings?custom_id=eq.{urllib.parse.quote(item.id)}")
        if check_id and isinstance(check_id, list) and len(check_id) > 0:
            existing = check_id[0]
            if existing['sku'] == item.sku:
                 raise HTTPException(status_code=400, detail=f"Data ID '{item.id}' dengan SKU ini sudah ada")
            else:
                 raise HTTPException(status_code=400, detail=f"Konflik: ID '{item.id}' sudah digunakan oleh SKU '{existing['sku']}'")
        
        # 2. Check if SKU exists
        check_sku = await supabase_fetch("GET", f"sku_mappings?sku=eq.{urllib.parse.quote(item.sku)}")
        if check_sku and isinstance(check_sku, list) and len(check_sku) > 0:
             existing = check_sku[0]
             raise HTTPException(status_code=400, detail=f"Konflik: SKU '{item.sku}' sudah digunakan oleh ID '{existing['custom_id']}'")

        # 3. Insert new data
        await supabase_fetch("POST", "sku_mappings", data={
            "custom_id": item.id, 
            "sku": item.sku, 
            "rak": item.rak if item.rak else ""
        })
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Add Mapping Failed: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Terjadi kesalahan: {str(e)}")

@app.put("/settings/sku-mappings/{id}")
async def update_sku_mapping(id: str, item: SkuMappingItem):
    try:
        decoded_id = urllib.parse.unquote(id)
        
        # 1. Check if SKU exists and belongs to a DIFFERENT ID
        check_sku = await supabase_fetch("GET", f"sku_mappings?sku=eq.{urllib.parse.quote(item.sku)}")
        if check_sku and isinstance(check_sku, list) and len(check_sku) > 0:
            for existing in check_sku:
                if existing['custom_id'] != decoded_id:
                     raise HTTPException(status_code=400, detail=f"Konflik: SKU '{item.sku}' sudah digunakan oleh ID '{existing['custom_id']}'")
                     
        # 2. Check if the NEW custom_id already exists (if it's being changed)
        if decoded_id != item.id:
            check_id = await supabase_fetch("GET", f"sku_mappings?custom_id=eq.{urllib.parse.quote(item.id)}")
            if check_id and isinstance(check_id, list) and len(check_id) > 0:
                 raise HTTPException(status_code=400, detail=f"Konflik: ID Custom '{item.id}' sudah ada di database.")

        # 3. Update data
        await supabase_fetch("PATCH", f"sku_mappings?custom_id=eq.{urllib.parse.quote(decoded_id)}", data={
            "custom_id": item.id,
            "sku": item.sku,
            "rak": item.rak if item.rak else ""
        })
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Update Mapping Failed: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Terjadi kesalahan: {str(e)}")

@app.delete("/settings/sku-mappings/all")
async def delete_all_sku_mappings():
    try:
        print("[Admin] Nuclear Delete All: Deleting all rows directly...")
        # Cara paling efisien: DELETE dengan filter 'neq' yang selalu true
        # Ini menghapus SEMUA baris tanpa perlu fetch ID dulu (tidak ada limit 1000)
        await supabase_fetch("DELETE", "sku_mappings",
            params={"custom_id": "neq.________NEVER_MATCH________"},
            headers={"Prefer": "return=minimal"}
        )
        # Backup: jika filter params tidak bekerja, coba cara alternatif
        print("[Admin] Nuclear Delete completed.")
        return {"success": True}
    except Exception as e:
        print(f"[Admin] Delete All Error: {e}")
        # Fallback: endpoint style lama jika yang baru gagal
        try:
            await supabase_fetch("DELETE", "sku_mappings?custom_id=neq.________EMPTY_DUMMY________")
            return {"success": True}
        except Exception as e2:
            raise HTTPException(status_code=500, detail=f"Gagal mengosongkan database: {str(e2)}")

    except Exception as e:
        print(f"[Admin] Full Delete Error: {e}")
        raise HTTPException(status_code=500, detail=f"Gagal mengosongkan database: {str(e)}")

@app.delete("/settings/sku-mappings/{id}")
async def delete_sku_mapping(id: str):
    # Try deleting by custom_id first
    id = urllib.parse.unquote(id)
    await supabase_fetch("DELETE", f"sku_mappings?custom_id=eq.{urllib.parse.quote(id)}")
    return {"success": True}

class BulkDeleteSKUs(BaseModel):
    ids: list[str]

@app.post("/settings/sku-mappings/bulk-delete")
async def bulk_delete_sku_mappings(data: BulkDeleteSKUs):
    if not data.ids:
        return {"success": True, "count": 0}
    
    # Supabase allows 'in' filter: column=in.(val1,val2,...)
    # Using quote and comma separation
    ids_str = ",".join([f'"{val}"' for val in data.ids])
    endpoint = f"sku_mappings?custom_id=in.({urllib.parse.quote(ids_str)})"
    await supabase_fetch("DELETE", endpoint)
    return {"success": True, "count": len(data.ids)}

# --- ADMIN AUTH ---
class AdminPin(BaseModel):
    pin: str

@app.post("/toolkit/verify-pin")
async def verify_toolkit_pin(item: AdminPin):
    if item.pin in ["1995", "1088"]:
        return {"success": True}
    raise HTTPException(status_code=401, detail="PIN Salah")


@app.post("/toolkit/split-excel")
async def split_excel(
    file: UploadFile = File(...),
    batch_limit: int = Form(50),
    item_limit: int = Form(7000), 
    dynamic_batching: bool = Form(False),
    prioritize_satuan: bool = Form(False),
    bulky_multiplier: int = Form(1) # Default 1 (Normal)
):
    try:
        # Fetch Bulky SKUs
        bulky_skus = set()
        try:
             bulky_data = await supabase_fetch("GET", "sku_bulky?select=sku")
             bulky_skus = {str(b['sku']).strip().upper() for b in bulky_data}
             print(f"[SPLITTER] Loaded {len(bulky_skus)} Bulky SKUs.")
        except Exception as e:
             print(f"Error fetching bulky skus: {e}")

        # Check limit safety
        if batch_limit < 1: batch_limit = 50
        if item_limit < 10: item_limit = 7000 # Safety floor

        # Read Excel content
        content = await file.read()
        
        try:
            # READ AS OBJECT to preserve Python types (int/str) from engine (openpyxl).
            df = pd.read_excel(io.BytesIO(content), dtype=object)
            
            # Fix long numeric IDs (scientific notation)
            df = fix_excel_numeric_ids(df)
            
            # Safely convert everything else to string handling NaNs
            df = df.fillna("").astype(str)
            
            # Clean strings (remove .0 if appeared somehow, though astype(str) on int shouldn't have it)
            # Logic: If ID is "12345.0", make it "12345"
            # But with dtype=object, 12345 (int) becomes "12345".
            
        except Exception as e:
            raise HTTPException(status_code=400, detail="Format file tidak valid. Pastikan file Excel (.xlsx/.xls).")
            
        # Validate Columns
        cols_map = {c.lower(): c for c in df.columns}
        
        def find_col(keywords):
            for k in keywords:
                for c_lower in cols_map:
                    if k in c_lower:
                        return cols_map[c_lower]
            return None
            
        id_col = find_col(['id pesanan', 'order id', 'nomor pesanan'])
        
        # --- FIX MERGED CELLS (Critical for Ginee/Multi-row orders) ---
        # If ID column has merged cells, pandas reads them as NaN for the 2nd+ rows.
        # We must Forward Fill (ffill) to propagate the ID to all items in the order.
        if id_col:
             df[id_col] = df[id_col].replace('', float('nan')).ffill()

        sku_col = find_col(['msku', 'sku', 'product id'])
        qty_col = find_col(['jumlah', 'qty', 'quantity'])
        awb_col = find_col(['awb', 'no. tracking', 'tracking no', 'resi', 'no resi', 'tracking', 'no. resi'])
        
        missing = []
        if not id_col: missing.append("ID Pesanan")
        if not sku_col: missing.append("MSKU")
        if not qty_col: missing.append("Jumlah")
        
        if missing:
             raise HTTPException(status_code=400, detail=f"Kolom wajib tidak ditemukan: {', '.join(missing)}")
             
        # Filter DataFrame to keep ONLY required columns
        columns_to_keep = [id_col, sku_col, qty_col]
        rename_dict = {id_col: "ID Pesanan", sku_col: "MSKU", qty_col: "Jumlah"}
        
        if awb_col:
            # FIX: Only ffill AWB within the SAME order (same ID Pesanan), NOT across different orders.
            # Plain ffill() would incorrectly propagate one order's AWB into the next order's rows
            # when the next order has no AWB yet (e.g. AWB belum ada di Ginee).
            df[awb_col] = df[awb_col].replace('', float('nan'))
            if id_col:
                df[awb_col] = df.groupby(id_col, sort=False)[awb_col].transform(lambda x: x.ffill())
            else:
                df[awb_col] = df[awb_col].ffill()  # fallback: no ID col, use global ffill
            columns_to_keep.append(awb_col)
            rename_dict[awb_col] = "AWB/No. Tracking"
        else:
            df["AWB/No. Tracking"] = ""
            columns_to_keep.append("AWB/No. Tracking")
            rename_dict["AWB/No. Tracking"] = "AWB/No. Tracking"
            
        df = df[columns_to_keep]
        # Rename columns to standard generic format
        df = df.rename(columns=rename_dict)
        
        # Update references for the rest of processing
        id_col = "ID Pesanan"
        sku_col = "MSKU"
        qty_col = "Jumlah"
        awb_col = "AWB/No. Tracking"
              
        print(f"[SPLITTER] Standard Logic. Dynamic Batching: {dynamic_batching}. Prioritize Satuan: {prioritize_satuan}")
        
        # Ensure Order IDs are grouped together if they are scattered in the file
        # sort=False in groupby will handle this by Order of Appearance.
        
        # SMART BATCHING LOGIC
        # Group by ID Pesanan to keep orders together
        grouped_obj = df.groupby(id_col, sort=False)
        grouped_list = list(grouped_obj)

        # === BULKY SATUAN SEPARATION ===
        # When prioritize_satuan=True, we first extract all "bulky satuan" orders
        # and batch them separately based on total QTY (default max 10 pcs per batch).
        # Only satuan (1 unique MSKU) orders that contain a bulky SKU are affected.
        # All other satuan and pretelan use the existing logic below.

        bulky_satuan_orders = []   # (name, group) tuples for bulky satuan
        remaining_orders = []      # everything else (normal satuan + pretelan)

        if prioritize_satuan and bulky_skus:
            for name, group in grouped_list:
                group_skus_upper = [str(s).strip().upper() for s in group[sku_col] if pd.notna(s)]
                unique_mskus = list(set(group_skus_upper))
                is_satuan = len(unique_mskus) == 1
                is_bulky = any(s in bulky_skus for s in group_skus_upper)

                if is_satuan and is_bulky:
                    bulky_satuan_orders.append((name, group))
                else:
                    remaining_orders.append((name, group))
            print(f"[SPLITTER] Bulky Satuan: {len(bulky_satuan_orders)} orders separated for QTY-based batching.")
        else:
            remaining_orders = grouped_list

        # --- BULKY SATUAN: Batch by total QTY (max bulky_batch_qty per batch) ---
        # Sort by MSKU A-Z for zone efficiency
        bulky_satuan_orders.sort(key=lambda x: str(x[0]).upper())

        bulky_batches = []
        cur_bulky_dfs = []
        cur_bulky_qty = 0

        for name, group in bulky_satuan_orders:
            try:
                order_qty = int(float(group[qty_col].sum()))
            except Exception:
                order_qty = 1

            # Cut the batch if adding this order exceeds the QTY limit (but only if batch is not empty)
            if cur_bulky_qty + order_qty > bulky_multiplier and cur_bulky_dfs:
                bulky_batches.append(pd.concat(cur_bulky_dfs))
                cur_bulky_dfs = []
                cur_bulky_qty = 0

            cur_bulky_dfs.append(group)
            cur_bulky_qty += order_qty

        if cur_bulky_dfs:
            bulky_batches.append(pd.concat(cur_bulky_dfs))

        print(f"[SPLITTER] Bulky Satuan created {len(bulky_batches)} batches (max {bulky_multiplier} pcs/batch).")

        # --- REMAINING ORDERS: Sort and batch as before ---
        if prioritize_satuan:
            print("[SPLITTER] Sorting remaining groups: Satuan A-Z -> Pretelan original order")

            def get_sort_key(item_tuple):
                index, (name, group_df) = item_tuple
                mskus = [str(x).strip() for x in group_df[sku_col] if pd.notna(x)]
                unique_mskus = sorted(list(set(mskus)))
                is_pretelan = len(unique_mskus) > 1
                if not is_pretelan:
                    return (0, unique_mskus[0] if unique_mskus else "", 0)
                else:
                    return (1, "", index)

            remaining_with_index = list(enumerate(remaining_orders))
            sorted_items = sorted(remaining_with_index, key=get_sort_key)
            grouped = [item[1] for item in sorted_items]
        else:
            grouped = remaining_orders

        batches = []
        current_batch_dfs = []
        current_batch_load = 0
        current_batch_items = 0

        print(f"[SPLITTER] Processing remaining... Limit: {batch_limit} ({'Dynamic' if dynamic_batching else 'Static'})")

        for name, group in grouped:
            group_skus = [str(s).strip().upper() for s in group[sku_col] if pd.notna(s)]
            is_bulky_order = any(s in bulky_skus for s in group_skus)

            order_weight = 1
            if is_bulky_order:
                order_weight += max(0, bulky_multiplier - 1)

            group_items_count = len(group)
            is_order_pretelan = group_items_count > 1

            should_cut = False

            if dynamic_batching:
                target_limit = batch_limit
                if is_order_pretelan:
                    target_limit = 50
                soft_limit = int(target_limit * 0.8)
                if current_batch_load + order_weight > target_limit and current_batch_load > 0:
                    should_cut = True
                elif current_batch_load >= soft_limit and current_batch_items >= 200:
                    should_cut = True
            else:
                if current_batch_load + order_weight > batch_limit and current_batch_load > 0:
                    should_cut = True

            if should_cut:
                if current_batch_dfs:
                    batches.append(pd.concat(current_batch_dfs))
                    current_batch_dfs = []
                    current_batch_load = 0
                    current_batch_items = 0

            current_batch_dfs.append(group)
            current_batch_load += order_weight
            current_batch_items += group_items_count

        if current_batch_dfs:
            batches.append(pd.concat(current_batch_dfs))

        # Final batch list: Bulky Satuan first, then the rest
        batches = bulky_batches + batches
            
        print(f"[SPLITTER] Created {len(batches)} batches.")
        
        # Prepare in-memory Excel Single File (Multi-Sheet)
        output_buffer = io.BytesIO()
        
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
             # --- CREATE SUMMARY SHEET ---
             summary_data = []
             summary_data.append(["SUMMARY LAPORAN SPLIT BATCH", ""])
             summary_data.append(["Tanggal Proses", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
             summary_data.append(["Total Batch (Sheet)", len(batches)])
             # summary_data.append(["Total Pesanan", len(df)]) # df might be filtered
             summary_data.append(["", ""])
             summary_data.append(["DAFTAR BATCH", "JUMLAH BARIS"])
             
             for i, batch_df in enumerate(batches):
                 # Count unique orders for the sheet name
                 num_orders = len(batch_df[id_col].unique())
                 sheet_name = f"Batch {i+1} ({num_orders} Orders)"
                 summary_data.append([sheet_name, len(batch_df)])
                 
             print(f"[SPLITTER] WRITING SUMMARY SHEET...")
             summary_df = pd.DataFrame(summary_data)
             summary_df.to_excel(writer, sheet_name="SUMMARY", index=False, header=False)
             
             # --- WRITE BATCH SHEETS ---
             for i, batch_df in enumerate(batches):
                 num_orders = len(batch_df[id_col].unique())
                 sheet_name = f"Batch {i+1} ({num_orders} Orders)"
                 batch_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Return Single Excel File
        output_buffer.seek(0)
        return StreamingResponse(
            output_buffer, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers={"Content-Disposition": f"attachment; filename=splitted_batches_V2_{file.filename}"}
        )

    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Split Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/toolkit/split-excel-v3")
async def split_excel_v3(
    file: UploadFile = File(...),
    batch_limit: int = Form(50)
):
    try:
        # Check limit safety
        if batch_limit < 1: batch_limit = 50

        # Read Excel content
        content = await file.read()
        
        try:
             # Force string
            df = pd.read_excel(io.BytesIO(content), dtype=object)
            df = df.fillna("").astype(str)
        except Exception as e:
            raise HTTPException(status_code=400, detail="Format file tidak valid. Pastikan file Excel (.xlsx/.xls).")
            
        # Validate Columns
        cols_map = {c.lower(): c for c in df.columns}
        
        def find_col(keywords):
            for k in keywords:
                for c_lower in cols_map:
                    if k in c_lower:
                        return cols_map[c_lower]
            return None
            
        id_col = find_col(['id pesanan', 'order id', 'nomor pesanan'])
        sku_col = find_col(['msku', 'sku', 'product id'])
        qty_col = find_col(['jumlah', 'qty', 'quantity'])
        awb_col = find_col(['awb', 'no. tracking', 'tracking no', 'resi', 'no resi', 'tracking', 'no. resi'])
        
        missing = []
        if not id_col: missing.append("ID Pesanan")
        if not sku_col: missing.append("MSKU")
        if not qty_col: missing.append("Jumlah")
        
        if missing:
             raise HTTPException(status_code=400, detail=f"Kolom wajib tidak ditemukan: {', '.join(missing)}")
             
        # Filter DataFrame
        columns_to_keep = [id_col, sku_col, qty_col]
        rename_dict = {id_col: "ID Pesanan", sku_col: "MSKU", qty_col: "Jumlah"}
        
        if awb_col:
            # FIX: Only ffill AWB within the SAME order (same ID Pesanan), NOT across different orders.
            df[awb_col] = df[awb_col].replace('', float('nan'))
            if id_col:
                df[awb_col] = df.groupby(id_col, sort=False)[awb_col].transform(lambda x: x.ffill())
            else:
                df[awb_col] = df[awb_col].ffill()  # fallback: no ID col
            columns_to_keep.append(awb_col)
            rename_dict[awb_col] = "AWB/No. Tracking"
        else:
            df["AWB/No. Tracking"] = ""
            columns_to_keep.append("AWB/No. Tracking")
            rename_dict["AWB/No. Tracking"] = "AWB/No. Tracking"
            
        df = df[columns_to_keep]
        # Rename columns to standard generic format
        df = df.rename(columns=rename_dict)
        
        # Update references for the rest of processing
        id_col = "ID Pesanan"
        sku_col = "MSKU"
        qty_col = "Jumlah"
        awb_col = "AWB/No. Tracking"
        
        # Helper: Clean ID
        def clean_id(val):
            s = str(val).strip()
            if s.endswith(".0"): return s[:-2]
            return s
            
        df[id_col] = df[id_col].apply(clean_id)
        
        # Power Process:
        grouped = df.groupby(id_col, sort=False)
        order_units = []
        
        # Calculate Workload for ALL orders first to handle "Sisa" logic correctly
        # We need to preserve original order for Sheet1?
        # VBA: "dataRange.Sort Key1:=wsMaster.Range("H1"), Order1:=xlAscending" (Sorted by workload Ascending initially?)
        # But later "processRange.Sort Key1:=... Order1:=xlDescending"
        
        # Let's collect them all
        for name, group in grouped:
            workload = len(group)
            order_units.append({
                "id": name,
                "df": group,
                "workload": workload
            })
            
        # Logic V3 Adjusted: EXACT BATCHING (VBA Strict)
        total_orders = len(order_units)
        
        # 1. Calculate Number of Batches (Integer Division)
        # If Total < Batch Limit, num_batches = 0
        num_batches = total_orders // batch_limit
        
        # 2. Identify Processed vs Remaining
        if num_batches == 0 and total_orders > 0:
            num_batches = 1
            num_processed = total_orders
        else:
            num_processed = num_batches * batch_limit
            
        num_remaining = total_orders - num_processed
        
        # Split orders
        # VBA uses top N rows after some sort?
        # "If totalResi >= BATCH_SIZE Then ... resiYangDiproses = ..."
        # "processRange = wsMaster.Range("A1:H" & resiYangDiproses + 1)"
        # This implies it processes the FIRST N orders from the list.
        # AND the list was sorted Ascending by Workload earlier?
        # "Key1:=wsMaster.Range("H1"), Order1:=xlAscending" -> Lightest orders first?
        # Then it takes the first N (Lightest) orders to process?
        # AND THEN sorts them Descending for distribution?
        
        # Let's assume we engage the VBA logic:
        # 1. Sort ALL orders by Workload Ascending (Lightest first)
        order_units.sort(key=lambda x: x['workload']) # Ascending
        
        # 2. Slice Processed vs Remaining
        processed_units = order_units[:num_processed]
        remaining_units = order_units[num_processed:]
        
        # 3. Sort Processed Units Descending (For Greedy Distribution)
        processed_units.sort(key=lambda x: x['workload'], reverse=True)
        
        # Calculate Workload Stats
        total_items_processed = sum(u['workload'] for u in processed_units)
        target_workload = total_items_processed / num_batches if num_batches > 0 else 0
        
        print(f"[SPLITTER V3 Strict] Total: {total_orders}. Batches: {num_batches}. Processed: {num_processed}. Remaining: {num_remaining}")
        
        # 4. Distribute Processed Units
        batches = [[] for _ in range(num_batches)]
        batch_loads = [0] * num_batches
        
        if num_batches > 0:
            for unit in processed_units:
                min_load = min(batch_loads)
                target_batch_idx = batch_loads.index(min_load)
                
                batches[target_batch_idx].append(unit['df'])
                batch_loads[target_batch_idx] += unit['workload']
                
        # Compile Batch DataFrames
        final_batches = []
        for batch_list in batches:
            if batch_list:
                final_batches.append(pd.concat(batch_list))
            else:
                pass
                
        # Compile Remaining DataFrame
        if remaining_units:
            remaining_dfs = [u['df'] for u in remaining_units]
            remaining_df = pd.concat(remaining_dfs)
        else:
            remaining_df = pd.DataFrame(columns=[id_col, sku_col, qty_col, awb_col])
            
        # Generate Excel
        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
             # --- SUMMARY SHEET ---
             summary_data = []
             summary_data.append(["RINGKASAN PROSES (VBA Logic)", ""])
             summary_data.append(["Tanggal", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
             summary_data.append(["Total Resi Awal", total_orders])
             summary_data.append(["Resi Diproses", num_processed])
             summary_data.append(["Sisa Resi", num_remaining])
             summary_data.append(["Batch Dibuat", f"{num_batches} (@{batch_limit} resi)"])
             summary_data.append(["", ""])
             summary_data.append(["BEBAN KERJA PER BATCH", "JUMLAH SKU"])
             
             for i, items_count in enumerate(batch_loads):
                 batch_name = f"Batch {i+1}"
                 summary_data.append([batch_name, f"{items_count} SKU"])
                 
             # SUMMARY
             pd.DataFrame(summary_data).to_excel(writer, sheet_name="SUMMARY", index=False, header=False)
             
             # SISA DATA (Renamed & Deduped by ID Pesanan)
             if not remaining_df.empty:
                 # Strip whitespace from ID Pesanan
                 remaining_df[id_col] = remaining_df[id_col].astype(str).str.strip()
                 # BUG FIX: DO NOT drop_duplicates so we preserve all items in the 'Campur' order!

                 
             remaining_df.to_excel(writer, sheet_name="REMAINDER_SISA", index=False)
             
             # BATCH SHEETS
             for i, batch_df in enumerate(final_batches):
                 b_items = len(batch_df)
                 sheet_name = f"Batch {i+1} ({b_items} SKUs)"
                 batch_df.to_excel(writer, sheet_name=sheet_name, index=False)
             
        output_buffer.seek(0)
        return StreamingResponse(
            output_buffer, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers={"Content-Disposition": f"attachment; filename=splitted_batches_V3_{file.filename}"}
        )

    except HTTPException as he:
        raise he



@app.post("/toolkit/split-excel-v5")
async def split_excel_v5(
    file: UploadFile = File(...),
    batch_limit: int = Form(50)
):
    try:
        # Check limit safety
        if batch_limit < 1: batch_limit = 50

        # Read Excel content
        content = await file.read()
        
        try:
             # Force string
            df = pd.read_excel(io.BytesIO(content), dtype=object)
            df = df.fillna("").astype(str)
        except Exception as e:
            raise HTTPException(status_code=400, detail="Format file tidak valid. Pastikan file Excel (.xlsx/.xls).")
            
        # Validate Columns
        cols_map = {c.lower(): c for c in df.columns}
        
        def find_col(keywords):
            for k in keywords:
                for c_lower in cols_map:
                    if k in c_lower:
                        return cols_map[c_lower]
            return None
            
        id_col = find_col(['id pesanan', 'order id', 'nomor pesanan'])
        sku_col = find_col(['msku', 'sku', 'product id'])
        qty_col = find_col(['jumlah', 'qty', 'quantity'])
        awb_col = find_col(['awb', 'no. tracking', 'tracking no', 'resi', 'no resi', 'tracking', 'no. resi'])
        
        missing = []
        if not id_col: missing.append("ID Pesanan")
        if not sku_col: missing.append("MSKU")
        if not qty_col: missing.append("Jumlah")
        
        if missing:
             raise HTTPException(status_code=400, detail=f"Kolom wajib tidak ditemukan: {', '.join(missing)}")
             
        # Filter DataFrame
        columns_to_keep = [id_col, sku_col, qty_col]
        rename_dict = {id_col: "ID Pesanan", sku_col: "MSKU", qty_col: "Jumlah"}
        
        if awb_col:
            df[awb_col] = df[awb_col].replace('', float('nan'))
            if id_col:
                df[awb_col] = df.groupby(id_col, sort=False)[awb_col].transform(lambda x: x.ffill())
            else:
                df[awb_col] = df[awb_col].ffill() 
            columns_to_keep.append(awb_col)
            rename_dict[awb_col] = "AWB/No. Tracking"
        else:
            df["AWB/No. Tracking"] = ""
            columns_to_keep.append("AWB/No. Tracking")
            rename_dict["AWB/No. Tracking"] = "AWB/No. Tracking"
            
        df = df[columns_to_keep]
        df = df.rename(columns=rename_dict)
        
        id_col = "ID Pesanan"
        sku_col = "MSKU"
        qty_col = "Jumlah"
        awb_col = "AWB/No. Tracking"
        
        def clean_id(val):
            s = str(val).strip()
            if s.endswith(".0"): return s[:-2]
            return s
            
        df[id_col] = df[id_col].apply(clean_id)
        
        # --- LOGIKA V5 BARU (EXACT MATCH & PARTIAL MATCH) ---
        grouped = df.groupby(id_col, sort=False)
        all_order_units = []
        for name, group in grouped:
            # unique_skus tidak peduli qty, hanya peduli list MSKU unik yang diurutkan
            unique_skus = sorted(list(set(str(s).strip().upper() for s in group[sku_col] if pd.notna(s))))
            all_order_units.append({
                "id": name,
                "df": group,
                "workload": len(group),
                "unique_skus": unique_skus,
                "fingerprint": "|".join(unique_skus),
                "size": len(unique_skus)
            })

        # Pool initialization
        remaining_pool = all_order_units.copy()
        
        # Hitung frekuensi tiap fingerprint
        fingerprint_counts = {}
        for unit in remaining_pool:
            fp = unit['fingerprint']
            fingerprint_counts[fp] = fingerprint_counts.get(fp, 0) + 1
            
        # Pisahkan pesanan kembar (muncul >= 2 kali) dan unik (muncul cuma 1 kali)
        # Sesuai aturan: Kembar exact minimal 2 MSKU
        exact_units = []
        for unit in remaining_pool:
            if unit['size'] >= 2 and fingerprint_counts[unit['fingerprint']] >= 2:
                exact_units.append(unit)
                
        # Hapus exact units dari remaining_pool
        exact_ids = set(u['id'] for u in exact_units)
        remaining_pool = [u for u in remaining_pool if u['id'] not in exact_ids]
        
        final_sheets = []
        summary_data = [
            ["RINGKASAN PROSES V5 (EXACT & DYNAMIC PARTIAL MATCH)", ""],
            ["Tanggal", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Total Pesanan Keseluruhan", len(all_order_units)],
            ["", ""]
        ]
        
        def chunk_units(units, chunk_size):
            for i in range(0, len(units), chunk_size):
                yield units[i:i + chunk_size]

        # --- PHASE 1: BATCH PESANAN KEMBAR EXACT (100% SAMA) ---
        size_map_p1 = {}
        for unit in exact_units:
            sz = unit['size']
            if sz not in size_map_p1: size_map_p1[sz] = []
            size_map_p1[sz].append(unit)
            
        for sz in sorted(size_map_p1.keys()):
            orders_of_size = sorted(size_map_p1[sz], key=lambda x: x['fingerprint'])
            for idx, chunk in enumerate(chunk_units(orders_of_size, batch_limit)):
                df_chunk = pd.concat([u['df'] for u in chunk])
                sheet_name = f"Batch Kembar {sz} MSKU ({idx+1})"
                final_sheets.append((sheet_name, df_chunk))
            summary_data.append([f"Total Kembar Exact {sz} MSKU", f"{len(orders_of_size)} Resi"])

        summary_data.append(["", ""])

        # --- PHASES 2 - 5: BATCH KEMBAR PARSIAL (BEDA D = 1 sampai 4 MSKU) ---
        import itertools
        for d in [1, 2, 3, 4]:
            # Kandidat harus berukuran >= d + 2 agar minimal ada 2 MSKU yang sama/shared
            candidates = [u for u in remaining_pool if u['size'] >= d + 2]
            if not candidates:
                continue
                
            base_map = {}
            for unit in candidates:
                skus = unit['unique_skus']
                n = len(skus)
                k = n - d  # Jumlah MSKU yang sama
                for base_comb in itertools.combinations(skus, k):
                    b_key = tuple(sorted(base_comb))
                    if b_key not in base_map:
                        base_map[b_key] = []
                    base_map[b_key].append(unit)
                    
            # Urutkan kombinasi yang paling banyak resinya (paling populer)
            sorted_bases = sorted(base_map.items(), key=lambda x: len(x[1]), reverse=True)
            
            moved_ids = set()
            p_batches = []
            
            for b_key, units in sorted_bases:
                eligible_units = [u for u in units if u['id'] not in moved_ids]
                if len(eligible_units) >= 2:
                    p_batches.append((len(b_key), eligible_units))
                    for u in eligible_units:
                        moved_ids.add(u['id'])
                        
            # Hapus resi yang sudah digrup dari remaining_pool
            remaining_pool = [u for u in remaining_pool if u['id'] not in moved_ids]
            
            # Menyusun Sheet untuk Beda D
            p_counter = {}
            for base_sz, units_list in p_batches:
                if base_sz not in p_counter: p_counter[base_sz] = 1
                units_list.sort(key=lambda x: (x['size'], x['fingerprint']))
                
                for chunk in chunk_units(units_list, batch_limit):
                    df_chunk = pd.concat([u['df'] for u in chunk])
                    sheet_name = f"Batch Mirip {base_sz} MSKU (Beda {d}) - {p_counter[base_sz]}"
                    final_sheets.append((sheet_name, df_chunk))
                    p_counter[base_sz] += 1
                    
                summary_data.append([f"Grup Parsial Beda {d} (Sama {base_sz} MSKU)", f"{len(units_list)} Resi"])
                
            if p_batches:
                summary_data.append(["", ""])

        summary_data.append(["Total Pesanan Sisa (Bagi Rata)", f"{len(remaining_pool)} Resi"])

        # --- PHASE FINAL: BATCH PESANAN UNIK / SISA (BAGI RATA NORMAL) ---
        if remaining_pool:
            num_batches_p4 = len(remaining_pool) // batch_limit
            num_processed_p4 = num_batches_p4 * batch_limit
            
            remaining_pool.sort(key=lambda x: x['workload'])
            processed_p4 = remaining_pool[:num_processed_p4]
            sisa_p4 = remaining_pool[num_processed_p4:]
            
            processed_p4.sort(key=lambda x: x['workload'], reverse=True)
            
            p4_batches = [[] for _ in range(num_batches_p4)]
            p4_loads = [0] * num_batches_p4
            
            if num_batches_p4 > 0:
                for unit in processed_p4:
                    min_load = min(p4_loads)
                    target_idx = p4_loads.index(min_load)
                    p4_batches[target_idx].append(unit['df'])
                    p4_loads[target_idx] += unit['workload']
            
            for i, batch_list in enumerate(p4_batches):
                if batch_list:
                    df_b = pd.concat(batch_list)
                    final_sheets.append((f"Batch Bagi Rata {i+1} ({p4_loads[i]} SKU)", df_b))
                    summary_data.append([f"Batch Bagi Rata {i+1} Load", f"{p4_loads[i]} SKU"])
            
            if sisa_p4:
                remainder_df = pd.concat([u['df'] for u in sisa_p4])
                # Hapus duplikat untuk sheet sisa (opsional, ikuti versi asli)
                remainder_view = remainder_df.drop_duplicates(subset=[id_col])
                final_sheets.append(("Batch SISA_UNIK", remainder_view))
        
        # Write Excel
        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            pd.DataFrame(summary_data).to_excel(writer, sheet_name="SUMMARY", index=False, header=False)
            for s_name, s_df in final_sheets:
                s_df.to_excel(writer, sheet_name=s_name, index=False)
                
        output_buffer.seek(0)
        return StreamingResponse(
            output_buffer, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers={"Content-Disposition": f"attachment; filename=splitted_V5_Mega_{file.filename}"}
        )

    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Split V5 Mega Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Split V5 Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))





@app.post("/toolkit/preview-split-excel")
async def preview_split_excel(
    file: UploadFile = File(...)
):
    try:
        # Read Excel content
        content = await file.read()
        
        try:
            # Force string to preserve data integrity (leading zeros etc)
            df = pd.read_excel(io.BytesIO(content), dtype=str)
        except Exception as e:
            raise HTTPException(status_code=400, detail="Format file tidak valid. Pastikan file Excel (.xlsx/.xls).")
            
        # Validate Columns (Same logic as split_excel)
        cols_map = {c.lower(): c for c in df.columns}
        
        def find_col(keywords):
            for k in keywords:
                for c_lower in cols_map:
                    if k in c_lower:
                        return cols_map[c_lower]
            return None
            
        id_col = find_col(['id pesanan', 'order id', 'nomor pesanan'])
        sku_col = find_col(['msku', 'sku', 'product id'])
        qty_col = find_col(['jumlah', 'qty', 'quantity'])
        
        missing = []
        if not id_col: missing.append("ID Pesanan")
        if not sku_col: missing.append("MSKU")
        if not qty_col: missing.append("Jumlah")
        
        if missing:
             raise HTTPException(status_code=400, detail=f"Kolom wajib tidak ditemukan: {', '.join(missing)}")
             
        # Filter DataFrame
        df = df[[id_col, sku_col, qty_col]]
        
        # Group by ID Pesanan
        grouped = df.groupby(id_col, sort=False)
        
        preview_data = []
        
        for name, group in grouped:
            order_id = str(name)
            items = []
            
            records = group.to_dict('records')
            
            unique_mskus = set()
            total_qty = 0
            
            for row in records:
                msku = str(row[sku_col])
                qty_str = str(row[qty_col])
                try:
                    qty =  pd.to_numeric(qty_str, errors='coerce')
                    if pd.isna(qty): qty = 0
                except:
                    qty = 0
                
                items.append({
                    "msku": msku,
                    "qty": int(qty)
                })
                unique_mskus.add(msku)
                total_qty += int(qty)
                
            is_pretelan = len(unique_mskus) > 1
            
            preview_data.append({
                "id_pesanan": order_id,
                "items": items,
                "is_pretelan": is_pretelan,
                "msku_count": len(unique_mskus),
                "total_items": total_qty
            })
            
        return JSONResponse(content={"data": preview_data, "total_orders": len(preview_data)})

    except Exception as e:
        print(f"Error previewing excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Split Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/toolkit/extract-awb")
async def extract_awb_from_excel(file: UploadFile = File(...)):
    try:
        content = await file.read()
        # Read Excel, force string to preserve leading zeros
        df = pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False)
        
        # Cari kolom target (Case insensitive search)
        target_col = None
        possible_names = ["AWB/No. Tracking", "AWB", "No. Resi", "Tracking Number", "No. Tracking"]
        
        # 1. Exact match attempt
        for col in df.columns:
            if str(col).strip() in possible_names:
                target_col = col
                break
        
        # 2. Case insensitive partial match if not found
        if not target_col:
            for col in df.columns:
                col_upper = str(col).upper()
                if "AWB" in col_upper or "TRACKING" in col_upper or "RESI" in col_upper:
                    target_col = col
                    break
        
        if not target_col:
            raise HTTPException(status_code=400, detail="Kolom 'AWB/No. Tracking' tidak ditemukan dalam Excel.")
            
        # Extract data
        awbs = []
        for val in df[target_col]:
            val_str = str(val).strip()
            if val_str:
                awbs.append(val_str)
        
        # Remove duplicates while preserving order
        awbs = list(dict.fromkeys(awbs))
                
        return {"success": True, "count": len(awbs), "awbs": awbs, "column_name": target_col}
        
    except Exception as e:
        print(f"Extract AWB Error: {e}")
        raise HTTPException(status_code=500, detail=f"Gagal memproses file: {str(e)}")

@app.post("/toolkit/merge-pdfs")
async def merge_pdfs(files: List[UploadFile] = File(...)):
    try:
        if not files:
            raise HTTPException(status_code=400, detail="Tidak ada file yang diunggah")
            
        merged_pdf = fitz.open()
        
        for file in files:
            content = await file.read()
            # Try to open the uploaded file as a PDF
            try:
                pdf = fitz.open(stream=content, filetype="pdf")
                merged_pdf.insert_pdf(pdf)
                pdf.close()
            except Exception as e:
                print(f"[ERROR] Failed to process {file.filename}: {e}")
                raise HTTPException(status_code=400, detail=f"File {file.filename} bukan PDF yang valid atau rusak.")
                
        out_stream = io.BytesIO()
        merged_pdf.save(out_stream)
        merged_pdf.close()
        out_stream.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        return StreamingResponse(
            out_stream,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=Merged_Labels_{timestamp}.pdf",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Merge PDF Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Gagal menggabungkan PDF: {str(e)}")

@app.post("/toolkit/orderan-kilat")
async def process_orderan_kilat(file: UploadFile = File(...)):
    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False)
        
        if len(df) == 0:
            raise HTTPException(status_code=400, detail="File Excel kosong")
            
        col_id_pesanan = next((c for c in df.columns if 'ID PESANAN' in str(c).upper() or 'NO. PESANAN' in str(c).upper() or 'AWB' in str(c).upper()), None)
        col_msku = next((c for c in df.columns if 'MSKU' in str(c).upper()), None)
        
        if not col_id_pesanan or not col_msku:
            raise HTTPException(status_code=400, detail="Kolom 'ID Pesanan' atau 'MSKU' tidak ditemukan")
            
        # Ambil data SKU VIP dari Supabase
        vip_data = await supabase_fetch("GET", "sku_vip?select=sku")
        vip_skus = {e['sku'] for e in vip_data} if vip_data else set()
        
        if not vip_skus:
            raise HTTPException(status_code=404, detail="Data SKU VIP masih kosong di database")
            
        # Tangani merged cells dengan forward-fill pada ID Pesanan
        # Karena dtype=str dan keep_default_na=False, merged cells kosong mungkin berupa string kosong ''
        import numpy as np
        df_ffill = df.copy()
        df_ffill[col_id_pesanan] = df_ffill[col_id_pesanan].replace('', np.nan)
        df_ffill[col_id_pesanan] = df_ffill[col_id_pesanan].ffill()
        
        # Cari ID Pesanan mana saja yang memiliki MSKU VIP
        vip_order_ids = set()
        for idx, row in df_ffill.iterrows():
            msku = str(row[col_msku]).strip()
            order_id = str(row[col_id_pesanan]).strip()
            if msku in vip_skus:
                vip_order_ids.add(order_id)
                
        if not vip_order_ids:
            raise HTTPException(status_code=404, detail="Tidak ditemukan resi dengan SKU VIP di file ini")
            
        # Filter dataframe asli (dengan ID Pesanan yang sudah difill) untuk mempertahankan baris yang relevan
        df_filtered = df_ffill[df_ffill[col_id_pesanan].isin(vip_order_ids)]
        
        # Ekspor hasilnya
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df_filtered.to_excel(writer, index=False, sheet_name='Packing List')
            
            workbook  = writer.book
            worksheet = writer.sheets['Packing List']
            
            # Formats
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D9D9D9',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            cell_format_center = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'text_wrap': True
            })
            
            cell_format_left = workbook.add_format({
                'align': 'left',
                'valign': 'vcenter',
                'border': 1,
                'text_wrap': True
            })
            
            # Write headers
            for col_num, value in enumerate(df_filtered.columns.values):
                worksheet.write(0, col_num, value, header_format)
            
            # Set column widths
            worksheet.set_column('A:A', 15)  # ID / Rak
            worksheet.set_column('B:B', 40)  # SKU
            worksheet.set_column('C:C', 10)  # QTY
            worksheet.set_column('D:D', 30)  # NO. PESANAN
            
            # Write data with wrapping
            for row_num in range(len(df_filtered)):
                worksheet.write(row_num + 1, 0, df_filtered.iloc[row_num, 0], cell_format_center)
                worksheet.write(row_num + 1, 1, df_filtered.iloc[row_num, 1], cell_format_left)
                worksheet.write(row_num + 1, 2, df_filtered.iloc[row_num, 2], cell_format_center)
                worksheet.write(row_num + 1, 3, df_filtered.iloc[row_num, 3], cell_format_left)

        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Hasil_Orderan_Kilat.xlsx"}
        )
        
    except HTTPException as he:
        raise he
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/toolkit/verify-labels")
async def verify_labels(
    original_pdf: UploadFile = File(...),
    custom_pdf: UploadFile = File(...),
    excel_file: UploadFile = File(...)
):
    try:
        # 1. Parse Excel First (to build mappings)
        excel_content = await excel_file.read()
        df = pd.read_excel(io.BytesIO(excel_content), dtype=object)
        df = fix_excel_numeric_ids(df)
        
        col_mapping = {
            'AWB/No. Tracking': ['AWB/No. Tracking', 'AWB', 'No. Tracking', 'Tracking Number', 'Resi', 'No Resi'],
            'ID Pesanan': ['ID Pesanan', 'Order ID', 'No. Pesanan', 'Nomor Pesanan'],
            'MSKU': ['MSKU', 'SKU', 'Nama SKU', 'Product SKU', 'Master SKU'],
            'Jumlah': ['Jumlah', 'Qty', 'Quantity', 'QTY']
        }
        for target_col, alternatives in col_mapping.items():
            if target_col not in df.columns:
                for alt in alternatives:
                    if alt in df.columns:
                        df = df.rename(columns={alt: target_col})
                        break
                        
        required_cols = ['MSKU', 'Jumlah']
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            raise HTTPException(status_code=400, detail=f"File Excel salah. Kolom tidak ditemukan: {missing}")
            
        has_id_pesanan = 'ID Pesanan' in df.columns
        has_awb = 'AWB/No. Tracking' in df.columns
        if not has_id_pesanan and not has_awb:
            raise HTTPException(status_code=400, detail="Kolom 'ID Pesanan' atau 'AWB/No. Tracking' tidak ditemukan di Excel")
            
        if has_id_pesanan:
            df['ID Pesanan'] = df['ID Pesanan'].replace('', float('nan')).ffill()
        if has_awb:
            df['AWB/No. Tracking'] = df['AWB/No. Tracking'].replace('', float('nan'))
            if has_id_pesanan:
                df['AWB/No. Tracking'] = df.groupby('ID Pesanan', sort=False)['AWB/No. Tracking'].transform(lambda x: x.ffill())
            else:
                df['AWB/No. Tracking'] = df['AWB/No. Tracking'].ffill()
                
        excel_awbs = set()
        awb_to_id_mapping = {}
        id_to_awb_mapping = {}
        awb_to_items = {}
        
        for _, row in df.iterrows():
            id_pesanan = normalize_awb(row.get('ID Pesanan', '')) if has_id_pesanan else None
            awb = normalize_awb(row.get('AWB/No. Tracking', '')) if has_awb else None
            
            canonical_id = None
            if id_pesanan and id_pesanan not in ['NAN', 'NONE', 'NULL', '']:
                canonical_id = id_pesanan
            elif awb and awb not in ['NAN', 'NONE', 'NULL', '']:
                canonical_id = awb
                
            if not canonical_id: continue
            
            if id_pesanan and id_pesanan not in ['NAN', 'NONE', 'NULL', '']:
                excel_awbs.add(id_pesanan)
            if awb and awb not in ['NAN', 'NONE', 'NULL', '']:
                excel_awbs.add(awb)
                if id_pesanan and id_pesanan not in ['NAN', 'NONE', 'NULL', '']:
                    awb_to_id_mapping[awb] = id_pesanan
                    id_to_awb_mapping[id_pesanan] = awb
                    
            sku = str(row.get('MSKU', '')).strip()
            try:
                qty = int(float(row.get('Jumlah', 1)))
            except:
                qty = 1
                
            if sku and sku not in ['nan', 'NaN', 'NAN', '']:
                if canonical_id not in awb_to_items:
                    awb_to_items[canonical_id] = []
                awb_to_items[canonical_id].append({'sku': sku, 'qty': qty})

        # 2. Parse Original PDF
        orig_content = await original_pdf.read()
        orig_doc = fitz.open("pdf", orig_content)
        orig_canonical_ids = set()
        orig_canonical_to_text = {}
        for page_num in range(len(orig_doc)):
            page = orig_doc[page_num]
            text = page.get_text("text")
            candidates = extract_all_awb_candidates(text) + extract_order_ids(text)
            matched_awb = find_matching_awb(candidates, excel_awbs, page_num + 1)
            
            if matched_awb:
                canonical = awb_to_id_mapping.get(matched_awb, matched_awb)
                orig_canonical_ids.add(canonical)
                # Kumpulkan text PDF asli
                orig_canonical_to_text[canonical] = orig_canonical_to_text.get(canonical, "") + "\n" + text
            else:
                if candidates:
                    canonical = candidates[0]
                    orig_canonical_ids.add(canonical)
                    orig_canonical_to_text[canonical] = orig_canonical_to_text.get(canonical, "") + "\n" + text
                
        # 3. Parse Custom PDF
        custom_content = await custom_pdf.read()
        custom_doc = fitz.open("pdf", custom_content)
        custom_canonical_ids = set()
        custom_canonical_to_text = {}
        for page_num in range(len(custom_doc)):
            page = custom_doc[page_num]
            text = page.get_text("text")
            candidates = extract_all_awb_candidates(text) + extract_order_ids(text)
            matched_awb = find_matching_awb(candidates, excel_awbs, page_num + 1)
            
            if matched_awb:
                canonical = awb_to_id_mapping.get(matched_awb, matched_awb)
                custom_canonical_ids.add(canonical)
                # Kumpulkan text PDF custom
                custom_canonical_to_text[canonical] = custom_canonical_to_text.get(canonical, "") + "\n" + text
            else:
                if candidates:
                    canonical = candidates[0]
                    custom_canonical_ids.add(canonical)
                    custom_canonical_to_text[canonical] = custom_canonical_to_text.get(canonical, "") + "\n" + text

        missing_in_custom = list(orig_canonical_ids - custom_canonical_ids)
        
        mismatches = []
        # Check SKUs and Qty for matched AWBs
        matches = orig_canonical_ids.intersection(custom_canonical_ids)
        for canonical in matches:
            if canonical not in awb_to_items:
                # Should not happen for matched_awbs, but might happen for un-matched fallbacks
                mismatches.append({"awb": canonical, "reason": "ID Pesanan/Resi tidak ditemukan di file Excel"})
                continue
                
            # Basic validation: Check if custom PDF text contains the SKU (Fuzzy)
            import re
            custom_text = custom_canonical_to_text.get(canonical, "")
            norm_custom_text = re.sub(r'[\W_]+', '', custom_text.upper())
            
            orig_text = orig_canonical_to_text.get(canonical, "")
            norm_orig_text = re.sub(r'[\W_]+', '', orig_text.upper())
            
            items = awb_to_items[canonical]
            if not items:
                mismatches.append({"awb": canonical, "reason": "Tidak ada data produk (SKU) di file Excel untuk resi ini"})
                continue
                
            for item in items:
                sku = item['sku']
                norm_sku = re.sub(r'[\W_]+', '', str(sku).upper())
                target_sku = norm_sku[:12] # Ambil 12 karakter alfanumerik pertama
                
                if target_sku and target_sku not in norm_orig_text:
                    mismatches.append({"awb": canonical, "reason": f"SKU '{sku}' tidak terdeteksi di Label Asli (kemungkinan terpotong habis atau beda format)"})
                    break
                    
                if target_sku and target_sku not in norm_custom_text:
                    mismatches.append({"awb": canonical, "reason": f"SKU '{sku}' tidak terdeteksi di Label Custom buatan sistem"})
                    break
                
        return JSONResponse(content={
            "stats": {
                "original_total": len(orig_canonical_ids),
                "custom_total": len(custom_canonical_ids),
                "match_count": len(matches) - len(mismatches),
                "missing_count": len(missing_in_custom),
                "mismatch_count": len(mismatches)
            },
            "matches": list(matches),
            "missing_in_custom": missing_in_custom,
            "mismatches": mismatches
        })
    except Exception as e:
        import traceback
        print(f"Verify Labels Error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Gagal memverifikasi label: {str(e)}")

@app.post("/settings/import-sku")
async def import_sku_mappings(file: UploadFile = File(...)):
    try:
        content = await file.read()
        # Use dtype=str to preserve "0001", keep_default_na=False to avoid NaN for empty strings
        df = pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False)
        print(f"[IMPORT] Excel loaded. Rows: {len(df)}, Columns: {df.columns.tolist()}")
        
        # 0. Intelligent Column Detection
        cols_map = {str(c).lower().strip(): c for c in df.columns}
        
        def find_col(keywords):
            for k in keywords:
                for c_lower in cols_map:
                    if k in c_lower:
                        return cols_map[c_lower]
            return None
            
        id_col = find_col(['id custom', 'custom id', 'id pesanan', 'id'])
        sku_col = find_col(['kode sku', 'sku code', 'msku', 'sku'])
        rak_col = find_col(['lokasi rak', 'rak location', 'id rak', 'rak'])
        lorong_col = find_col(['lorong', 'zona lorong', 'zona'])
        
        # Fallback to column index if names don't match
        if not id_col or not sku_col:
            print("[IMPORT] Keywords not found, falling back to column index 0 and 1.")
            if len(df.columns) >= 2:
                id_col = df.columns[0]
                sku_col = df.columns[1]
                if len(df.columns) >= 3 and not rak_col:
                    rak_col = df.columns[2]
                if len(df.columns) >= 4 and not lorong_col:
                    lorong_col = df.columns[3]
            else:
                 raise HTTPException(status_code=400, detail="Format Excel minimal butuh 2 kolom (ID dan SKU).")
        
        print(f"[IMPORT] Using Columns -> LORONG: '{lorong_col}', RAK: '{rak_col}', ID: '{id_col}', SKU: '{sku_col}'")

        # De-duplikasi berdasarkan SKU (bukan ID!)
        # Alasan: file Excel pakai 1 ID untuk banyak SKU (1 ID = 1 rak = banyak produk)
        # Beberapa SKU bisa menempati rak/posisi yang sama → custom_id TIDAK perlu unik
        to_import = []
        seen_skus = set()
        duplicate_count = 0
        empty_count = 0
        
        for index, row in df.iterrows():
            raw_id = str(row[id_col]).strip()
            raw_sku = str(row[sku_col]).strip()
            raw_rak = str(row[rak_col]).strip() if rak_col and rak_col in row else ""
            raw_lorong = str(row[lorong_col]).strip() if lorong_col and lorong_col in row else ""
            
            # Skip baris kosong
            if not raw_id or raw_id in ('-', 'nan', 'None') or \
               not raw_sku or raw_sku in ('-', 'nan', 'None'):
                empty_count += 1
                continue
            
            # De-duplikasi berdasarkan SKU (setiap SKU hanya boleh ada 1 kali)
            if raw_sku in seen_skus:
                duplicate_count += 1
                continue
            
            # Gabungkan Lorong + RAK + ID → custom_id
            # Contoh: LORONG="12", RAK="EK", ID="06-22" → custom_id="12-EK-06-22"
            # custom_id TIDAK perlu unik — beberapa SKU bisa di lokasi yang sama
            parts = []
            if raw_lorong and raw_lorong not in ('-', 'nan', 'None'):
                parts.append(raw_lorong)
            if raw_rak and raw_rak not in ('-', 'nan', 'None'):
                parts.append(raw_rak)
            if raw_id:
                parts.append(raw_id)
                
            final_id = "-".join(parts) if parts else ""
                 
            to_import.append({
                "custom_id": final_id, 
                "sku": raw_sku, 
                "rak": raw_rak
            })
            seen_skus.add(raw_sku)
            
        print(f"[IMPORT] Summary -> Total rows: {len(df)}, Valid/Unik SKU: {len(to_import)}, Duplikat SKU: {duplicate_count}, Baris kosong: {empty_count}") 

        # 2. Hapus semua data lama (tanpa limit)
        try:
            print("[IMPORT] Clearing existing database...")
            await delete_all_sku_mappings()
            print("[IMPORT] Database cleared.")
        except Exception as e:
            print(f"[IMPORT] Warning: Clear failed but continuing insertion: {e}")
            
        # 3. Batch Upsert (tidak gagal jika ada konflik unique key)
        if to_import:
            chunk_size = 100
            total_batches = (len(to_import) + chunk_size - 1) // chunk_size
            inserted = 0
            for i in range(0, len(to_import), chunk_size):
                chunk = to_import[i:i + chunk_size]
                batch_num = i // chunk_size + 1
                try:
                    await supabase_fetch(
                        "POST", "sku_mappings",
                        data=chunk,
                        headers={
                            "Prefer": "resolution=merge-duplicates,return=minimal"
                        }
                    )
                    inserted += len(chunk)
                    print(f"[IMPORT] Batch {batch_num}/{total_batches} OK ({inserted}/{len(to_import)})")
                except Exception as batch_e:
                    print(f"[IMPORT] Batch {batch_num} FAILED: {batch_e}")
                    # Lanjut batch berikutnya, jangan stop
                
        return {"success": True, "count": len(to_import)}
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Import error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Gagal import: {str(e)}")

@app.get("/settings/export-sku")
async def export_sku_mappings():
    mappings = await get_sku_mappings() 
        
    df = pd.DataFrame(mappings)
    if not df.empty:
         df = df.rename(columns={'id': 'ID', 'sku': 'SKU', 'rak': 'RAK'})
    else:
         df = pd.DataFrame(columns=['ID', 'SKU', 'RAK'])
         
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='SKU Mappings')
        worksheet = writer.sheets['SKU Mappings']
        worksheet.set_column('A:A', 15)
        worksheet.set_column('B:B', 30)
        
    output.seek(0)
    filename = "SKU_Mappings_Export.xlsx"
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# --- SKU GROUPING / CATEGORIES ---

class SkuCategoryCreate(BaseModel):
    name: str

class SkuCategoryMemberCreate(BaseModel):
    sku: str

@app.get("/settings/categories")
async def get_sku_categories():
    try:
        # Get categories
        categories = await supabase_fetch("GET", "sku_categories?select=*&order=created_at.asc")
        
        # Get member counts for each category
        # Uses a separate query or join if possible, but simplest is separate count or fetch all members
        # For simplicity, we'll just fetch categories first. If we need counts, we can add a view in Supabase or fetch all members.
        # Let's fetch all members to count them in memory (efficient enough for small datasets)
        members = await supabase_fetch("GET", "sku_category_members?select=category_id")
        
        member_counts = {}
        for m in members:
            cid = m['category_id']
            member_counts[cid] = member_counts.get(cid, 0) + 1
            
        for cat in categories:
            cat['member_count'] = member_counts.get(cat['id'], 0)
            
        return categories
    except Exception as e:
        print(f"Get Categories Error: {e}")
        return []

@app.post("/settings/categories")
async def create_sku_category(item: SkuCategoryCreate):
    try:
        # Check duplicate name
        existing = await supabase_fetch("GET", f"sku_categories?name=eq.{urllib.parse.quote(item.name)}")
        if existing:
            raise HTTPException(status_code=400, detail="Nama kategori sudah ada.")
            
        res = await supabase_fetch("POST", "sku_categories", data={"name": item.name}, params={"select": "*"})
        return res[0] if res else {}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/settings/categories/{id}")
async def delete_sku_category(id: str):
    try:
        await supabase_fetch("DELETE", f"sku_categories?id=eq.{id}")
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/settings/categories/{id}/members")
async def get_category_members(id: str):
    try:
        # Join not supported directly via simple REST cleanly without definition, so we fetch members
        # And if we want to show custom_id, we might need to join locally or valid matches
        data = await supabase_fetch("GET", f"sku_category_members?category_id=eq.{id}&select=*")
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/settings/grouping-list")
async def get_all_grouping_members():
    # Fetch all members with category name
    # Syntax: select=id,sku,sku_categories(id,name)
    try:
        data = await supabase_fetch("GET", "sku_category_members?select=id,sku,category:sku_categories(id,name)")
        # Flatten structure
        result = []
        for item in data:
            cat = item.get('category') or {}
            result.append({
                "id": item['id'],
                "sku": item['sku'],
                "category_id": cat.get('id'),
                "category_name": cat.get('name')
            })
        return result
    except Exception as e:
        print(f"Error fetching grouping list: {e}")
        return []
    finally:
        print(f"[DEBUG] Grouping List: Returning {len(result) if 'result' in locals() else 0} rows")

class GroupingItem(BaseModel):
    category_name: str
    sku: str
    
@app.post("/settings/grouping-list")
async def add_grouping_row(item: GroupingItem):
    # 1. Get or Create Category
    cat_name = item.category_name.strip()
    sku = item.sku.strip()
    
    if not cat_name or not sku:
        raise HTTPException(status_code=400, detail="ID dan SKU harus diisi")
    
    existing_cat = await supabase_fetch("GET", f"sku_categories?name=eq.{urllib.parse.quote(cat_name)}")
    cat_id = None
            
    if existing_cat:
        cat_id = existing_cat[0]['id']
    else:
        new_cat = await supabase_fetch("POST", "sku_categories", data={"name": cat_name}, params={"select": "id"})
        if new_cat:
            cat_id = new_cat[0]['id']
            
    if not cat_id:
        raise HTTPException(status_code=500, detail="Gagal membuat kategori")

    # 2. Insert Member
    try:
        await supabase_fetch("POST", "sku_category_members", data={"category_id": cat_id, "sku": sku})
    except HTTPException as e:
        if e.status_code == 409: # Conflict
             raise HTTPException(status_code=400, detail="Data SKU ini sudah ada di grup tersebut")
        raise e
        
    return {"success": True}

@app.delete("/settings/grouping-list/{id}")
async def delete_grouping_member(id: str):
    await supabase_fetch("DELETE", f"sku_category_members?id=eq.{id}")
    return {"success": True}

@app.post("/settings/categories/{id}/members")
async def add_category_member(id: str, item: SkuCategoryMemberCreate):
    try:
        await supabase_fetch("POST", "sku_category_members", data={"category_id": id, "sku": item.sku})
        return {"success": True}
    except Exception as e:
        # Check duplicate
        raise HTTPException(status_code=400, detail="Gagal menambahkan member (mungkin duplikat)")

@app.delete("/settings/categories/{id}/members/{sku}")
async def delete_category_member(id: str, sku: str):
    sku = urllib.parse.unquote(sku)
    try:
        url = f"sku_category_members?category_id=eq.{id}&sku=eq.{urllib.parse.quote(sku)}"
        await supabase_fetch("DELETE", url)
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- PRIORITY BOTTOM ENDPOINTS ---

@app.get("/settings/priority-bottom")
async def get_priority_bottom():
    try:
        data = await supabase_fetch("GET", "sku_priority_bottom?select=id,sku")
        return data
    except Exception as e:
        print(f"Priority Get Error: {e}")
        return []
    finally:
        print(f"[DEBUG] Priority Bottom: Returning {len(data) if 'data' in locals() else 0} rows")

class PriorityItem(BaseModel):
    sku: str

@app.post("/settings/priority-bottom")
async def add_priority_bottom(item: PriorityItem):
    try:
        # Check if already exists? Unique constraint handles it.
        await supabase_fetch("POST", "sku_priority_bottom", data={"sku": item.sku})
        return {"success": True}
    except Exception as e:
        # Likely duplicate
        print(f"Priority Add Error: {e}")
        raise HTTPException(status_code=400, detail="SKU sudah ada di list atau error lain")

@app.delete("/settings/priority-bottom/{sku}")
async def delete_priority_bottom(sku: str):
    sku = urllib.parse.unquote(sku)
    try:
        url = f"sku_priority_bottom?sku=eq.{urllib.parse.quote(sku)}"
        await supabase_fetch("DELETE", url)
        return {"success": True}
    except Exception as e:
         raise HTTPException(status_code=500, detail=str(e))

@app.post("/settings/priority-bottom/bulk-delete")
async def bulk_delete_priority_bottom(data: BulkDeleteSKUs):
    if not data.ids:
        return {"success": True, "count": 0}
    # ids here are SKUs as strings
    ids_str = ",".join([f'"{val}"' for val in data.ids])
    endpoint = f"sku_priority_bottom?sku=in.({urllib.parse.quote(ids_str)})"
    await supabase_fetch("DELETE", endpoint)
    return {"success": True, "count": len(data.ids)}

@app.post("/settings/priority-bottom/export")
async def export_priority_bottom_selected(data: BulkDeleteSKUs):
    # Fetch data for selected SKUs
    if not data.ids:
         # Export all if none selected? User said "sesuai yang di select". 
         # But usually if they click export without selection, maybe export all? 
         # Let's stick to selected.
         mappings = await get_priority_bottom()
    else:
         ids_str = ",".join([f'"{val}"' for val in data.ids])
         endpoint = f"sku_priority_bottom?sku=in.({urllib.parse.quote(ids_str)})&select=id,sku"
         mappings = await supabase_fetch("GET", endpoint)

    df = pd.DataFrame(mappings)
    if not df.empty:
         df = df.rename(columns={'id': 'ID', 'sku': 'SKU'})
    else:
         df = pd.DataFrame(columns=['ID', 'SKU'])
         
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Priority Bottom')
        
    output.seek(0)
    filename = f"Priority_Bottom_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.post("/settings/import-priority-bottom")
async def import_priority_bottom(file: UploadFile = File(...)):
    try:
        content = await file.read()
        
        # Read Excel, expect column "SKU" (or first column)
        df = pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False)
        
        # Determine SKU column
        col_sku = next((c for c in df.columns if 'SKU' in str(c).upper()), None)
        if not col_sku:
             # Fallback to first column
             col_sku = df.columns[0]
             
        skus = set()
        for val in df[col_sku]:
            s = str(val).strip()
            if s: skus.add(s)
            
        print(f"[DEBUG] Priority Import: Found {len(skus)} SKUs")
        
        # Bulk Insert (Individual or Chunked)
        # Check existing first to avoid error spam?
        # Or just try insert and ignore errors.
        
        # Fetch existing
        existing_data = await supabase_fetch("GET", "sku_priority_bottom?select=sku")
        existing_skus = {e['sku'] for e in existing_data}
        
        new_skus = [s for s in skus if s not in existing_skus]
        
        if new_skus:
            # Chunk insert 100
            chunk_size = 100
            total_added = 0
            for i in range(0, len(new_skus), chunk_size):
                chunk = [{"sku": s} for s in new_skus[i:i + chunk_size]]
                await supabase_fetch("POST", "sku_priority_bottom", data=chunk)
                total_added += len(chunk)
            
            return {"success": True, "count": total_added}
        
        return {"success": True, "count": 0}

    except Exception as e:
        print(f"Import Priority Error: {e}")
        raise HTTPException(status_code=500, detail=f"Gagal import: {str(e)}")


# --- BARANG KHUSUS (SUPABASE) ---

@app.get("/settings/barang-khusus")
async def get_barang_khusus_list():
    try:
        data = await supabase_fetch("GET", "sku_barang_khusus?select=id,sku")
        return data
    except Exception as e:
        print(f"Barang Khusus Get Error: {e}")
        return []

@app.post("/settings/barang-khusus")
async def add_barang_khusus(item: PriorityItem):
    try:
        await supabase_fetch("POST", "sku_barang_khusus", data={"sku": item.sku})
        return {"success": True}
    except Exception as e:
        print(f"Barang Khusus Add Error: {e}")
        raise HTTPException(status_code=400, detail="SKU sudah ada di list atau error lain")

@app.delete("/settings/barang-khusus/{sku}")
async def delete_barang_khusus(sku: str):
    sku = urllib.parse.unquote(sku)
    try:
        url = f"sku_barang_khusus?sku=eq.{urllib.parse.quote(sku)}"
        await supabase_fetch("DELETE", url)
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/settings/barang-khusus/bulk-delete")
async def bulk_delete_barang_khusus(payload: BulkDeleteSKUs):
    if not payload.ids:
        return {"success": True, "count": 0}
    ids_str = ",".join([f'"{val}"' for val in payload.ids])
    endpoint = f"sku_barang_khusus?sku=in.({urllib.parse.quote(ids_str)})"
    await supabase_fetch("DELETE", endpoint)
    return {"success": True, "count": len(payload.ids)}

@app.post("/settings/barang-khusus/export")
async def export_barang_khusus(payload: BulkDeleteSKUs):
    if not payload.ids:
        mappings = await get_barang_khusus_list()
    else:
        ids_str = ",".join([f'"{val}"' for val in payload.ids])
        endpoint = f"sku_barang_khusus?sku=in.({urllib.parse.quote(ids_str)})&select=id,sku"
        mappings = await supabase_fetch("GET", endpoint)
        
    df = pd.DataFrame(mappings)
    if not df.empty:
         df = df.rename(columns={'id': 'ID', 'sku': 'SKU'})
    else:
         df = pd.DataFrame(columns=['ID', 'SKU'])
         
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Barang Khusus')
        
    output.seek(0)
    filename = f"Barang_Khusus_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.post("/settings/import-barang-khusus")
async def import_barang_khusus(file: UploadFile = File(...)):
    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False)
        col_sku = next((c for c in df.columns if 'SKU' in str(c).upper()), None)
        if not col_sku:
             col_sku = df.columns[0]
             
        skus = set()
        for val in df[col_sku]:
            s = str(val).strip()
            if s: skus.add(s)
            
        existing_data = await supabase_fetch("GET", "sku_barang_khusus?select=sku")
        existing_skus = {e['sku'] for e in existing_data}
        
        new_skus = [s for s in skus if s not in existing_skus]
        
        if new_skus:
            chunk_size = 100
            total_added = 0
            for i in range(0, len(new_skus), chunk_size):
                chunk = [{"sku": s} for s in new_skus[i:i + chunk_size]]
                await supabase_fetch("POST", "sku_barang_khusus", data=chunk)
                total_added += len(chunk)
            
            return {"success": True, "count": total_added}
            
        return {"success": True, "count": 0}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gagal import: {str(e)}")

# --- SKU VIP (>10K) ---

@app.get("/settings/sku-vip")
async def get_sku_vip_list():
    try:
        data = await supabase_fetch("GET", "sku_vip?select=id,sku")
        return data
    except Exception:
        return []

@app.post("/settings/sku-vip")
async def add_sku_vip(item: PriorityItem):
    try:
        await supabase_fetch("POST", "sku_vip", data={"sku": item.sku})
        return {"success": True}
    except Exception:
        raise HTTPException(status_code=400, detail="SKU sudah ada di list atau error lain")

@app.delete("/settings/sku-vip/{sku}")
async def delete_sku_vip(sku: str):
    sku = urllib.parse.unquote(sku)
    try:
        url = f"sku_vip?sku=eq.{urllib.parse.quote(sku)}"
        await supabase_fetch("DELETE", url)
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/settings/sku-vip/bulk-delete")
async def bulk_delete_sku_vip(payload: BulkDeleteSKUs):
    if not payload.ids:
        return {"success": True, "count": 0}
    ids_str = ",".join([f'"{val}"' for val in payload.ids])
    endpoint = f"sku_vip?sku=in.({urllib.parse.quote(ids_str)})"
    await supabase_fetch("DELETE", endpoint)
    return {"success": True, "count": len(payload.ids)}

@app.post("/settings/sku-vip/export")
async def export_sku_vip(payload: BulkDeleteSKUs):
    if not payload.ids:
        mappings = await get_sku_vip_list()
    else:
        ids_str = ",".join([f'"{val}"' for val in payload.ids])
        mappings = await supabase_fetch("GET", f"sku_vip?sku=in.({urllib.parse.quote(ids_str)})&select=id,sku")
    
    if not mappings:
        raise HTTPException(status_code=404, detail="Tidak ada data untuk diexport")
        
    df = pd.DataFrame(mappings)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='SKU VIP')
    
    output.seek(0)
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=SKU_VIP_Export.xlsx"}
    )

@app.post("/settings/import-sku-vip")
async def import_sku_vip(file: UploadFile = File(...)):
    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False)
        col_sku = next((c for c in df.columns if 'SKU' in str(c).upper()), None)
        if not col_sku:
             col_sku = df.columns[0]
             
        skus = set()
        for val in df[col_sku]:
            s = str(val).strip()
            if s: skus.add(s)
            
        existing_data = await supabase_fetch("GET", "sku_vip?select=sku")
        existing_skus = {e['sku'] for e in existing_data} if existing_data else set()
        
        new_skus = [s for s in skus if s not in existing_skus]
        
        if new_skus:
            chunk_size = 100
            total_added = 0
            for i in range(0, len(new_skus), chunk_size):
                chunk = [{"sku": s} for s in new_skus[i:i + chunk_size]]
                await supabase_fetch("POST", "sku_vip", data=chunk)
                total_added += len(chunk)
            
            return {"success": True, "count": total_added}
            
        return {"success": True, "count": 0}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gagal import: {str(e)}")


# --- FORMATTING RULES MODELS ---
class FormattingColorItem(BaseModel):
    keyword: str
    color_code: str

class FormattingStyleItem(BaseModel):
    keyword: str
    font_size: int = 11
    is_bold: bool = False

# --- FORMATTING ENDPOINTS ---

@app.get("/settings/formatting/colors")
async def get_formatting_colors():
    try:
        data = await supabase_fetch("GET", "sku_formatting_colors?select=*&order=created_at.desc")
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        print(f"[DEBUG] Table Colors: Returning {len(data) if 'data' in locals() else 0} rows")

@app.post("/settings/formatting/colors")
async def add_formatting_color(item: FormattingColorItem):
    try:
        # Check duplicate keyword
        existing = await supabase_fetch("GET", f"sku_formatting_colors?keyword=eq.{item.keyword}")
        if existing:
            raise HTTPException(status_code=400, detail="Keyword warna sudah ada")
        
        await supabase_fetch("POST", "sku_formatting_colors", data=item.dict(), params={"select": "id"})
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/settings/formatting/colors/{id}")
async def delete_formatting_color(id: UUID):
    try:
        await supabase_fetch("DELETE", f"sku_formatting_colors?id=eq.{id}")
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/settings/formatting/styles")
async def get_formatting_styles():
    try:
        data = await supabase_fetch("GET", "sku_formatting_styles?select=*&order=created_at.desc")
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        print(f"[DEBUG] Table Styles: Returning {len(data) if 'data' in locals() else 0} rows")

@app.post("/settings/formatting/styles")
async def add_formatting_style(item: FormattingStyleItem):
    try:
         # Check duplicate
        existing = await supabase_fetch("GET", f"sku_formatting_styles?keyword=eq.{item.keyword}")
        if existing:
             raise HTTPException(status_code=400, detail="Keyword style sudah ada")

        await supabase_fetch("POST", "sku_formatting_styles", data=item.dict(), params={"select": "id"})
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/settings/formatting/styles/{id}")
async def delete_formatting_style(id: UUID):
    try:
        await supabase_fetch("DELETE", f"sku_formatting_styles?id=eq.{id}")
        return {"success": True}
    except Exception as e:
         raise HTTPException(status_code=500, detail=str(e))

# --- COLUMN SETTINGS ---
class ColumnSettingItem(BaseModel):
    column_name: str
    column_width: float = 20
    font_size: int = 16
    font_name: str = "Rockwell"
    is_bold: bool = False
    text_align: str = "center"

class ColumnSettingUpdate(BaseModel):
    column_width: float = None
    font_size: int = None
    font_name: str = None
    is_bold: bool = None
    text_align: str = None

@app.get("/settings/formatting/columns")
async def get_column_settings():
    try:
        data = await supabase_fetch("GET", "sku_column_settings?select=*&order=column_name.asc")
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        print(f"[DEBUG] Table Columns: Returning {len(data) if 'data' in locals() else 0} rows")

@app.post("/settings/formatting/columns")
async def add_column_setting(item: ColumnSettingItem):
    try:
        # Check duplicate column name
        existing = await supabase_fetch("GET", f"sku_column_settings?column_name=eq.{item.column_name}")
        if existing:
            raise HTTPException(status_code=400, detail="Kolom sudah ada")
        
        await supabase_fetch("POST", "sku_column_settings", data=item.dict(), params={"select": "id"})
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- BULKY SKU ENDPOINTS ---

@app.get("/settings/bulky-skus")
async def get_bulky_skus():
    try:
        data = await supabase_fetch("GET", "sku_bulky?select=id,sku&order=created_at.desc")
        return data
    except Exception as e:
        print(f"Bulky SKU Get Error: {e}")
        return []
    finally:
        print(f"[DEBUG] Bulky SKUs: Returning {len(data) if 'data' in locals() else 0} rows")

class BulkySkuItem(BaseModel):
    sku: str

@app.post("/settings/bulky-skus")
async def add_bulky_sku(item: BulkySkuItem):
    try:
        await supabase_fetch("POST", "sku_bulky", data={"sku": item.sku})
        return {"success": True}
    except Exception as e:
        print(f"Bulky SKU Add Error: {e}")
        raise HTTPException(status_code=400, detail="SKU sudah ada di list bulky atau error lain")

@app.delete("/settings/bulky-skus/{sku}")
async def delete_bulky_sku(sku: str):
    sku = urllib.parse.unquote(sku)
    try:
        url = f"sku_bulky?sku=eq.{urllib.parse.quote(sku)}"
        await supabase_fetch("DELETE", url)
        return {"success": True}
    except Exception as e:
         raise HTTPException(status_code=500, detail=str(e))

@app.post("/settings/import-bulky-skus")
async def import_bulky_skus(file: UploadFile = File(...)):
    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False)
        col_sku = next((c for c in df.columns if 'SKU' in str(c).upper()), df.columns[0])
             
        skus = {str(val).strip() for val in df[col_sku] if str(val).strip()}
        
        existing_data = await supabase_fetch("GET", "sku_bulky?select=sku")
        existing_skus = {e['sku'] for e in existing_data}
        
        new_skus = [s for s in skus if s not in existing_skus]
        if new_skus:
            chunk_size = 100
            for i in range(0, len(new_skus), chunk_size):
                chunk = [{"sku": s} for s in new_skus[i:i + chunk_size]]
                await supabase_fetch("POST", "sku_bulky", data=chunk)
            return {"success": True, "count": len(new_skus)}
        return {"success": True, "count": 0}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gagal import bulky SKUs: {str(e)}")

# --- LABEL TABLE LAYOUT CONFIG ---

LABEL_CONFIG_KEY = "label_table_config"

LABEL_CONFIG_DEFAULTS = {
    # Format Standar (2 kolom: MSKU + Qty)
    "std_col_msku": 220,        # lebar kolom MSKU (pt)
    "std_col_qty": 50,          # lebar kolom Qty (pt)
    "std_font_msku": 8,         # ukuran font kolom MSKU (pt)
    "std_font_qty": 13.5,       # ukuran font kolom Qty (pt)
    "std_font_size": 8,         # legacy / fallback (pt)
    "std_row_height": 18,       # tinggi baris minimum (pt)
    # Format Rak & ID (3 kolom: Rak | MSKU | Qty)
    "ext_col_rak": 80,          # lebar kolom Rak & ID (pt)
    "ext_col_msku": 150,        # lebar kolom MSKU (pt)
    "ext_col_qty": 50,          # lebar kolom Qty (pt)
    "ext_font_rak": 10,          # ukuran font kolom Rak & ID (pt)
    "ext_font_msku": 8,         # ukuran font kolom MSKU (pt)
    "ext_font_qty": 13.5,       # ukuran font kolom Qty (pt)
    "ext_font_size": 8,         # legacy / fallback (pt)
    "ext_row_height": 25,       # tinggi baris minimum (pt)
    # Umum
    "border_thickness": 0.5,    # ketebalan garis border (pt)
    "header_bg": "#ffffff",     # warna background header (hex)
    "header_color": "#000000",  # warna teks header (hex)
    
    # Font Family (Baru)
    "std_font_msku_family": "Helvetica",
    "std_font_qty_family": "Bahnschrift",
    "ext_font_rak_family": "Bahnschrift",
    "ext_font_msku_family": "Helvetica",
    "ext_font_qty_family": "Bahnschrift",
    "std_font_msku_bold": False,
    "std_font_qty_bold": True,
    "ext_font_rak_bold": False,
    "ext_font_msku_bold": False,
    "ext_font_qty_bold": True,
}

LABEL_CONFIG_FILE = "label_config.json"

async def get_label_config() -> dict:
    """Ambil label config dari Supabase (app_settings), fallback ke defaults jika gagal."""
    import json
    try:
        res = await supabase_fetch("GET", "app_settings?key=eq.label_table_config")
        if res and isinstance(res, list) and len(res) > 0:
            val = res[0].get('value', '{}')
            cfg = json.loads(val) if isinstance(val, str) else val
            return {**LABEL_CONFIG_DEFAULTS, **cfg}
    except Exception as e:
        print(f"[LabelConfig] Supabase fetch error: {e}")
        
    # Fallback ke file lokal jika offline
    try:
        if os.path.exists(LABEL_CONFIG_FILE):
            with open(LABEL_CONFIG_FILE, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
            return {**LABEL_CONFIG_DEFAULTS, **cfg}
    except Exception:
        pass
        
    return LABEL_CONFIG_DEFAULTS

@app.get("/settings/label-table-config")
async def get_label_table_config():
    cfg = await get_label_config()
    return cfg

class LabelTableConfigItem(BaseModel):
    std_col_msku: float = 220
    std_col_qty: float = 50
    std_font_msku: float = 8
    std_font_qty: float = 8
    std_font_size: float = 8
    std_row_height: float = 18
    ext_col_rak: float = 70
    ext_col_msku: float = 150
    ext_col_qty: float = 50
    ext_font_rak: float = 8
    ext_font_msku: float = 8
    ext_font_qty: float = 8
    ext_font_size: float = 8
    ext_row_height: float = 18
    border_thickness: float = 0.5
    header_bg: str = "#000000"
    header_color: str = "#ffffff"
    std_font_msku_family: str = "Helvetica"
    std_font_qty_family: str = "Helvetica"
    ext_font_rak_family: str = "Helvetica"
    ext_font_msku_family: str = "Helvetica"
    ext_font_qty_family: str = "Helvetica"
    std_font_msku_bold: bool = False
    std_font_qty_bold: bool = False
    ext_font_rak_bold: bool = False
    ext_font_msku_bold: bool = False
    ext_font_qty_bold: bool = False

@app.post("/settings/label-table-config")
async def save_label_table_config(item: LabelTableConfigItem):
    import json
    try:
        cfg_dict = item.dict()
        cfg_str = json.dumps(cfg_dict)
        
        # Save to Supabase
        try:
            existing = await supabase_fetch("GET", "app_settings?key=eq.label_table_config")
            if existing and len(existing) > 0:
                await supabase_fetch("PATCH", "app_settings?key=eq.label_table_config", data={"value": cfg_str})
            else:
                await supabase_fetch("POST", "app_settings", data={"key": "label_table_config", "value": cfg_str})
        except Exception as e:
            print(f"[LabelConfig] Supabase save error: {e}")
            
        # Also save locally as fallback
        with open(LABEL_CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(cfg_dict, f, indent=2)
            
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- MENU SETTINGS ---
from typing import List

class MenuSettingsItem(BaseModel):
    hidden_menus: List[str]
    menu_order: List[str]

@app.get("/settings/menu")
async def get_menu_settings():
    cache_file = "menu_settings_cache.json"
    try:
        data = await supabase_fetch("GET", "menu_settings?select=hidden_menus,menu_order&limit=1")
        if data and isinstance(data, list) and len(data) > 0:
            result = data[0]
            try:
                with open(cache_file, 'w', encoding='utf-8') as f: json.dump(result, f)
            except: pass
            return result
    except Exception as e:
        print(f"Menu Settings Get Error: {e}")
    # Fallback to local cache
    if os.path.exists(cache_file):
        try:
            with open(cache_file, 'r', encoding='utf-8') as f: return json.load(f)
        except: pass
    return {"hidden_menus": [], "menu_order": []}

@app.post("/settings/menu")
async def save_menu_settings(item: MenuSettingsItem):
    try:
        # Check if row exists
        existing = await supabase_fetch("GET", "menu_settings?select=id&limit=1")
        if existing and isinstance(existing, list) and len(existing) > 0:
            row_id = existing[0]['id']
            await supabase_fetch("PATCH", f"menu_settings?id=eq.{row_id}", data={
                "hidden_menus": item.hidden_menus,
                "menu_order": item.menu_order,
                "updated_at": datetime.now().isoformat()
            })
        else:
            await supabase_fetch("POST", "menu_settings", data={
                "hidden_menus": item.hidden_menus,
                "menu_order": item.menu_order,
                "updated_at": datetime.now().isoformat()
            })
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- TOOLKIT ORDER ---

@app.get("/settings/toolkit-order")
async def get_toolkit_order():
    cache_file = "toolkit_order_cache.json"
    try:
        res = await supabase_fetch("GET", "app_settings?key=eq.toolkit_order")
        if res and isinstance(res, list) and len(res) > 0:
            val = res[0].get('value', '[]')
            result = json.loads(val) if isinstance(val, str) else val
            try:
                with open(cache_file, 'w', encoding='utf-8') as f: json.dump(result, f)
            except: pass
            return result
    except Exception as e:
        print(f"Toolkit Order Get Error: {e}")
    # Fallback to local cache
    if os.path.exists(cache_file):
        try:
            with open(cache_file, 'r', encoding='utf-8') as f: return json.load(f)
        except: pass
    return []

class ToolkitOrderUpdate(BaseModel):
    order: List[str]

@app.post("/settings/toolkit-order")
async def update_toolkit_order(item: ToolkitOrderUpdate):
    try:
        order_str = json.dumps(item.order)
        existing = await supabase_fetch("GET", "app_settings?key=eq.toolkit_order")
        if existing and len(existing) > 0:
            await supabase_fetch("PATCH", "app_settings?key=eq.toolkit_order", data={"value": order_str})
        else:
            await supabase_fetch("POST", "app_settings", data={"key": "toolkit_order", "value": order_str})
        return {"success": True}
    except Exception as e:
        print(f"Toolkit Order Post Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# --- TOOLKIT FEATURE LOCKS ---


@app.get("/settings/toolkit-features")
async def get_toolkit_features():
    cache_file = "toolkit_features_cache.json"
    try:
        data = await supabase_fetch("GET", "toolkit_feature_locks?select=feature_key,is_locked")
        try:
            with open(cache_file, 'w', encoding='utf-8') as f: json.dump(data, f)
        except: pass
        return data
    except Exception as e:
        print(f"Toolkit Features Get Error: {e}")
    # Fallback to local cache
    if os.path.exists(cache_file):
        try:
            with open(cache_file, 'r', encoding='utf-8') as f: return json.load(f)
        except: pass
    return []

class ToolkitFeatureItem(BaseModel):
    feature_key: str
    is_locked: bool

@app.post("/settings/toolkit-features")
async def upsert_toolkit_feature(item: ToolkitFeatureItem):
    try:
        existing = await supabase_fetch("GET", f"toolkit_feature_locks?feature_key=eq.{urllib.parse.quote(item.feature_key)}")
        if existing:
            await supabase_fetch("PATCH", f"toolkit_feature_locks?feature_key=eq.{urllib.parse.quote(item.feature_key)}", data={"is_locked": item.is_locked})
        else:
            await supabase_fetch("POST", "toolkit_feature_locks", data={"feature_key": item.feature_key, "is_locked": item.is_locked})
        return {"success": True}
    except Exception as e:
        print(f"Toolkit Feature Upsert Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/settings/formatting/columns/{id}")
async def update_column_setting(id: UUID, item: ColumnSettingUpdate):
    try:
        update_data = {k: v for k, v in item.dict().items() if v is not None}
        if not update_data:
            raise HTTPException(status_code=400, detail="Tidak ada data untuk diupdate")
        
        await supabase_fetch("PATCH", f"sku_column_settings?id=eq.{id}", data=update_data)
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/settings/formatting/columns/{id}")
async def delete_column_setting(id: UUID):
    try:
        await supabase_fetch("DELETE", f"sku_column_settings?id=eq.{id}")
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class LabelPriorityItem(BaseModel):
    format_type: str
    keyword: str

@app.get("/settings/label-priority-bottom")
async def get_label_priority_bottom(format_type: str = None):
    try:
        url = "label_bottom_priorities"
        if format_type:
            url += f"?format_type=eq.{format_type}"
        data = await supabase_fetch("GET", url)
        return data
    except Exception as e:
        print(f"Error GET label priority: {e}")
        return []

@app.post("/settings/label-priority-bottom")
async def add_label_priority_bottom(item: LabelPriorityItem):
    import urllib.parse
    try:
        data = {"format_type": item.format_type, "keyword": item.keyword}
        existing = await supabase_fetch("GET", f"label_bottom_priorities?format_type=eq.{item.format_type}&keyword=eq.{urllib.parse.quote(item.keyword)}")
        if not existing:
            await supabase_fetch("POST", "label_bottom_priorities", data=data)
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/settings/label-priority-bottom/{id}")
async def delete_label_priority_bottom(id: str):
    try:
        await supabase_fetch("DELETE", f"label_bottom_priorities?id=eq.{id}")
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/settings/import-label-priority-bottom")
async def import_label_priority_bottom(
    file: UploadFile = File(...),
    format_type: str = Form(...)
):
    import io
    import pandas as pd
    import urllib.parse
    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content), dtype=str)
        if df.empty or len(df.columns) == 0:
            raise ValueError("File Excel kosong atau tidak memiliki kolom")
        
        col_name = df.columns[0]
        keywords = df[col_name].dropna().astype(str).tolist()
        
        # Ambil data lama agar tidak duplikat
        existing_res = await supabase_fetch("GET", f"label_bottom_priorities?format_type=eq.{format_type}")
        existing_kws = {item['keyword'].strip().upper() for item in existing_res} if existing_res else set()
        
        new_items = []
        for kw in keywords:
            clean_kw = kw.strip().upper()
            if clean_kw and clean_kw not in existing_kws:
                new_items.append({"format_type": format_type, "keyword": clean_kw})
                existing_kws.add(clean_kw) # cegah duplikat dari excel itu sendiri
                
        if new_items:
            chunk_size = 500
            for i in range(0, len(new_items), chunk_size):
                await supabase_fetch("POST", "label_bottom_priorities", data=new_items[i:i+chunk_size])
                
        return {"success": True, "count": len(new_items)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/settings/label-priority-bottom/bulk-delete")
async def bulk_delete_label_priority_bottom(data: BulkDeleteSKUs):
    if not data.ids:
        return {"success": True, "count": 0}
    # ids here are internal IDs (UUIDs as strings)
    ids_str = ",".join([f'{val}' for val in data.ids]) # UUIDs don't need extra quotes in in.() usually, but depends on supabase_fetch
    endpoint = f"label_bottom_priorities?id=in.({ids_str})"
    await supabase_fetch("DELETE", endpoint)
    return {"success": True, "count": len(data.ids)}

@app.post("/settings/label-priority-bottom/export")
async def export_label_priority_bottom_selected(data: BulkDeleteSKUs):
    if not data.ids:
         data_res = await supabase_fetch("GET", "label_bottom_priorities")
    else:
         ids_str = ",".join([f'{val}' for val in data.ids])
         endpoint = f"label_bottom_priorities?id=in.({ids_str})"
         data_res = await supabase_fetch("GET", endpoint)

    df = pd.DataFrame(data_res)
    if not df.empty:
         df = df.rename(columns={'id': 'ID', 'format_type': 'Format Type', 'keyword': 'Keyword'})
    else:
         df = pd.DataFrame(columns=['ID', 'Format Type', 'Keyword'])
         
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Label Priority Bottom')
        
    output.seek(0)
    filename = f"Label_Priority_Bottom_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )




@app.post("/settings/formatting/reset")
async def reset_formatting_rules():
    try:
        colors = await supabase_fetch("GET", "sku_formatting_colors?select=id")
        for c in colors:
            await supabase_fetch("DELETE", f"sku_formatting_colors?id=eq.{c['id']}")
            
        styles = await supabase_fetch("GET", "sku_formatting_styles?select=id")
        for s in styles:
            await supabase_fetch("DELETE", f"sku_formatting_styles?id=eq.{s['id']}")
            
        cols = await supabase_fetch("GET", "sku_column_settings?select=id")
        for c in cols:
            await supabase_fetch("DELETE", f"sku_column_settings?id=eq.{c['id']}")
            
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/settings/import-grouping-data")
async def import_grouping_data(file: UploadFile = File(...)):
    try:
        content = await file.read()
        print(f"[DEBUG] Grouping Import size: {len(content)} bytes")
        
        # Read Excel, expect columns like "ID" (Category) and "SKU BOX" (Member)
        # Use dtype=str to avoid float conversion of IDs
        df = pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False)
        
        # Normalize columns
        # User defined format: A=ID, B=SKU BOX
        # Let's map indexes 0 and 1 if names don't match exactly
        if len(df.columns) >= 2:
            # Rename first two columns to internal names
            df.columns.values[0] = 'category_name'
            df.columns.values[1] = 'sku'
        else:
            raise HTTPException(status_code=400, detail="Format Excel salah. Minimal 2 kolom (ID, SKU BOX).")
            
        # Parse Data
        # Map: CategoryName -> Set of SKUs
        category_map = {} 
        
        for index, row in df.iterrows():
            cat_name = str(row['category_name']).strip()
            sku = str(row['sku']).strip()
            
            if not cat_name or not sku: continue
            
            if cat_name not in category_map:
                category_map[cat_name] = set()
            category_map[cat_name].add(sku)
            
        print(f"[DEBUG] Found {len(category_map)} categories to process.")
        
        # Process Categories and Insert Members
        total_pushed = 0
        
        for cat_name, skus in category_map.items():
            # 1. Get or Create Category
            # Check if exists
            existing_cat = await supabase_fetch("GET", f"sku_categories?name=eq.{urllib.parse.quote(cat_name)}")
            cat_id = None
            
            if existing_cat:
                cat_id = existing_cat[0]['id']
            else:
                # Create new
                new_cat = await supabase_fetch("POST", "sku_categories", data={"name": cat_name}, params={"select": "id"})
                if new_cat:
                    cat_id = new_cat[0]['id']
            
            if not cat_id:
                print(f"[ERROR] Failed to get/create category: {cat_name}")
                continue
                
            # 2. Bulk Insert Members
            # We need to check existing members to avoid duplicates error (or just try insert)
            # Fetch existing members for this category to filter
            existing_members = await supabase_fetch("GET", f"sku_category_members?category_id=eq.{cat_id}&select=sku")
            existing_member_skus = {m['sku'] for m in existing_members}
            
            to_insert = []
            for sku in skus:
                if sku not in existing_member_skus:
                    to_insert.append({"category_id": cat_id, "sku": sku})
            
            if to_insert:
                # Chunk insert if needed, but for now direct
                # Supabase bulk insert limit is usually high enough for typical usage
                # If very large, might need chunking. Let's do 100 chunk.
                chunk_size = 100
                for i in range(0, len(to_insert), chunk_size):
                    chunk = to_insert[i:i + chunk_size]
                    await supabase_fetch("POST", "sku_category_members", data=chunk)
                    total_pushed += len(chunk)
                    
        return {"success": True, "count": total_pushed}

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Gagal import: {str(e)}")

# Add bulk add endpoint for convenience
@app.post("/settings/categories/{id}/members/bulk")
async def bulk_add_category_members(id: str, skus: List[str]):
    try:
        data = [{"category_id": id, "sku": sku} for sku in skus]
        # Supabase insert ignores duplicates? No, it errors.
        # We need to filter existing?
        # Or just try insert and ignore error? Supabase REST doesn't support ON CONFLICT IGNORE easily without stored proc.
        # Strict parsing: Fetch existing first.
        existing = await supabase_fetch("GET", f"sku_category_members?category_id=eq.{id}&select=sku")
        existing_set = {r['sku'] for r in existing}
        
        new_items = [d for d in data if d['sku'] not in existing_set]
        
        if new_items:
             await supabase_fetch("POST", "sku_category_members", data=new_items)
             
        return {"success": True, "added": len(new_items)}
    except Exception as e:
         raise HTTPException(status_code=500, detail=str(e))


class PinRequest(BaseModel):
    pin: str

@app.post("/admin/verify-pin")
async def verify_pin(request: PinRequest):
    # Hardcoded safe PIN in backend
    if request.pin in ["1995", "1088"]:
        return {"success": True}
    else:
        raise HTTPException(status_code=401, detail="PIN Salah")


# ========== BACKUP CONFIGURATION ==========
BACKUP_FOLDER = Path(os.path.expanduser("~")) / "OneDrive" / "Dokumen" / "SCRIPT" / "backup_data_localhost"
BACKUP_RETENTION_DAYS = 7  # File akan dihapus setelah 7 hari

def ensure_backup_folder():
    """Buat folder backup jika belum ada"""
    if not BACKUP_FOLDER.exists():
        BACKUP_FOLDER.mkdir(parents=True, exist_ok=True)
        print(f"[BACKUP] Created backup folder: {BACKUP_FOLDER}")
    return BACKUP_FOLDER

def cleanup_old_backups():
    """Hapus file backup yang lebih dari 7 hari"""
    try:
        ensure_backup_folder()
        now = datetime.now()
        deleted_count = 0
        
        # Cari semua subfolder dalam backup folder
        for item in BACKUP_FOLDER.iterdir():
            if item.is_dir():
                # Cek usia folder dari nama (format: YYYYMMDD_HHMMSS_...)
                try:
                    folder_name = item.name
                    # Extract date from folder name (first 8 chars = YYYYMMDD)
                    date_str = folder_name[:8]
                    folder_date = datetime.strptime(date_str, "%Y%m%d")
                    
                    # Jika lebih dari 7 hari, hapus
                    if (now - folder_date).days >= BACKUP_RETENTION_DAYS:
                        shutil.rmtree(item)
                        deleted_count += 1
                        print(f"[BACKUP] Deleted old backup: {item.name}")
                except (ValueError, IndexError):
                    # Folder name tidak sesuai format, skip
                    continue
        
        if deleted_count > 0:
            print(f"[BACKUP] Cleanup completed: {deleted_count} old backups deleted")
    except Exception as e:
        print(f"[BACKUP] Cleanup error: {e}")

def save_backup(excel_filename: str, pdf_filenames: list, excel_data: bytes, pdf_data: bytes, result_pdf: bytes, filtered_excel_data: bytes = None):
    """Simpan backup file Excel, PDF asli, dan PDF hasil proses"""
    try:
        ensure_backup_folder()
        
        # Buat subfolder dengan timestamp + nama PDF pertama
        # FIX: Sertakan nama PDF pertama di nama folder agar setiap batch unik,
        # mencegah packing list salah batch saat pakai Upload Massal dengan Excel yang sama.
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_excel_name = re.sub(r'[<>:"/\\|?*]', '_', excel_filename.replace('.xlsx', ''))
        
        # Ambil nama PDF pertama (tanpa ekstensi) sebagai identitas batch
        if pdf_filenames:
            safe_pdf_stem = re.sub(r'[<>:"/\\|?*]', '_', pdf_filenames[0].replace('.pdf', '').replace('.PDF', ''))
        else:
            safe_pdf_stem = 'nopdf'
        
        backup_subfolder = BACKUP_FOLDER / f"{timestamp}_{safe_excel_name}__{safe_pdf_stem}"
        backup_subfolder.mkdir(parents=True, exist_ok=True)
        
        # Simpan file Excel asli
        excel_path = backup_subfolder / excel_filename
        with open(excel_path, 'wb') as f:
            f.write(excel_data)
        print(f"[BACKUP] Saved Excel: {excel_path}")
        
        # Simpan PDF asli (jika ada data)
        if pdf_data:
            # Simpan sebagai file gabungan (fallback)
            pdf_original_path = backup_subfolder / "original_labels.pdf"
            with open(pdf_original_path, 'wb') as f:
                f.write(pdf_data)
            print(f"[BACKUP] Saved original PDF (combined): {pdf_original_path}")
            
            # FIX: Simpan juga salinan PDF dengan nama asli setiap PDF
            # Ini memungkinkan content-aware filter bekerja via nama file
            for pdf_fn in pdf_filenames:
                safe_fn = re.sub(r'[<>:"/\\|?*]', '_', pdf_fn)
                if safe_fn and safe_fn != 'original_labels.pdf':
                    named_pdf_path = backup_subfolder / safe_fn
                    if not named_pdf_path.exists():  # Jangan overwrite jika sudah ada
                        with open(named_pdf_path, 'wb') as f:
                            f.write(pdf_data)
                        print(f"[BACKUP] Saved named PDF copy: {named_pdf_path}")
        
        # Simpan PDF hasil proses
        result_path = backup_subfolder / f"result_{timestamp}.pdf"
        with open(result_path, 'wb') as f:
            f.write(result_pdf)
        print(f"[BACKUP] Saved result PDF: {result_path}")
        
        # Simpan Filtered Excel (jika ada)
        if filtered_excel_data:
            filtered_path = backup_subfolder / f"filtered_{excel_filename}"
            with open(filtered_path, 'wb') as f:
                f.write(filtered_excel_data)
            print(f"[BACKUP] Saved Filtered Excel: {filtered_path}")
        
        # Simpan info file PDF names
        info_path = backup_subfolder / "info.txt"
        with open(info_path, 'w', encoding='utf-8') as f:
            f.write(f"Timestamp: {datetime.now().isoformat()}\n")
            f.write(f"Excel: {excel_filename}\n")
            f.write(f"PDF Files: {', '.join(pdf_filenames)}\n")
            f.write(f"PDF Stem (Batch ID): {safe_pdf_stem}\n")
        
        print(f"[BACKUP] Backup completed: {backup_subfolder}")
        return str(backup_subfolder)
    except Exception as e:
        print(f"[BACKUP] Error saving backup: {e}")
        return None

def find_backup_folder(date_str: str, excel_filename: str, required_pdf_name: str = None) -> Optional[Path]:
    """
    Cari folder backup yang cocok berdasarkan timestamp history dan nama file excel.
    Fiksasi Timezone: Convert History (UTC) ke Local Time.
    Toleransi waktu: +- 2 menit (dikurangi dari 5 menit agar tidak salah batch).
    
    Strategi pencarian (prioritas turun):
    1. Folder name mengandung PDF stem (PALING AKURAT - folder baru menyertakan nama PDF)
    2. File PDF dengan nama yang cocok di dalam folder (konten check)
    3. Best time match (fallback untuk folder lama sebelum fix ini)
    
    Args:
        date_str: Timestamp dari history (UTC ISO format)
        excel_filename: Nama file excel
        required_pdf_name: (Optional) Nama file PDF / label batch ini.
                           Digunakan untuk membedakan batch yang diproses berdekatan dengan Excel sama.
    """
    try:
        print(f"[DEBUG] Received date_str: '{date_str}'")
        
        # Masalah URL Decoding: "+" sering berubah jadi " " (space)
        # Fix: Kembalikan space menjadi + jika ada indikasi offset format
        if ' ' in date_str:
            date_str = date_str.replace(' ', '+')
            print(f"[DEBUG] Fixed date_str: '{date_str}'")

        # Parse timestamp dari history (ISO format) -> biasanya UTC
        # Handle 'Z' suffix manual jaga-jaga
        history_time_utc = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        
        # Convert ke Local Time (Naive) agar match dengan folder name (yang dibuat pakai datetime.now())
        # astimezone(None) menggunakan local system timezone
        history_time_local = history_time_utc.astimezone(None).replace(tzinfo=None)
        
        safe_excel_name = re.sub(r'[<>:"/\\|?*]', '_', excel_filename.replace('.xlsx', ''))
        
        # FIX: Buat safe_pdf_stem untuk pencocokan nama folder (strategi 1)
        safe_pdf_stem = None
        if required_pdf_name:
            safe_pdf_stem = re.sub(r'[<>:"/\\|?*]', '_', required_pdf_name.replace('.pdf', '').replace('.PDF', '')).lower()
        
        candidates = []
        ensure_backup_folder()
        
        print(f"[DEBUG] Finding backup for: {excel_filename}")
        print(f"[DEBUG] History Time (Local Converted): {history_time_local}")
        if required_pdf_name:
            print(f"[DEBUG] Required PDF name: {required_pdf_name} | PDF Stem: {safe_pdf_stem}")
        
        for item in BACKUP_FOLDER.iterdir():
            if item.is_dir() and safe_excel_name in item.name:
                # Coba parse timestamp folder (YYYYMMDD_HHMMSS)
                try:
                    folder_ts_str = item.name[:15] # Ambil bagian YYYYMMDD_HHMMSS
                    folder_time = datetime.strptime(folder_ts_str, "%Y%m%d_%H%M%S")
                    
                    # Hitung selisih detik
                    diff = abs((history_time_local - folder_time).total_seconds())
                    
                    # FIX: Kurangi toleransi dari 5 menit menjadi 2 menit
                    # Ini mencegah batch yang diproses dalam jarak berdekatan saling silang
                    if diff < 120: # Toleransi 2 menit (sebelumnya 5 menit)
                        candidates.append((diff, item))
                    elif diff < 300:
                        # Terima dengan toleransi lebar HANYA jika folder name mengandung PDF stem
                        # (Untuk kompatibilitas folder lama sebelum fix ini)
                        if safe_pdf_stem and safe_pdf_stem in item.name.lower():
                            candidates.append((diff, item))
                except ValueError:
                    continue
        
        if candidates:
            # Sort by time difference (ascending)
            candidates.sort(key=lambda x: x[0])
            
            # --- STRATEGI 1: PENCOCOKAN NAMA FOLDER (PRIORITAS TERTINGGI) ---
            # Folder baru diberi nama {timestamp}_{excel}__{pdf_stem}
            # Jika folder name mengandung PDF stem yang dicari, ini adalah match paling akurat
            if safe_pdf_stem:
                print(f"[DEBUG] Strategy 1: Checking folder names for PDF stem '{safe_pdf_stem}'")
                for diff, folder_path in candidates:
                    folder_lower = folder_path.name.lower()
                    # Cek apakah nama folder mengandung '__' (separator baru) dan PDF stem
                    if '__' in folder_lower and safe_pdf_stem in folder_lower:
                        print(f"[DEBUG] Strategy 1 MATCH (folder name): {folder_path.name} (Diff: {diff:.2f}s)")
                        return folder_path
            
            # --- STRATEGI 2: KONTEN CHECK (FILE PDF DI DALAM FOLDER) ---
            # Untuk folder lama (sebelum fix) yang belum menyertakan nama PDF di folder name
            if required_pdf_name:
                target_pdf_stem = required_pdf_name.lower().replace('.pdf', '').replace('.PDF', '')
                
                print(f"[DEBUG] Strategy 2: Checking folder contents for PDF file '{target_pdf_stem}'")
                
                for diff, folder_path in candidates:
                    found_pdf = False
                    for existing_pdf in folder_path.glob("*.pdf"):
                        existing_stem = existing_pdf.name.lower().replace('.pdf', '')
                        # Pencocokan eksak atau fuzzy (folder lama mungkin punya nama sedikit berbeda)
                        if existing_stem == target_pdf_stem:
                            found_pdf = True
                            break
                        # Cek juga via safe name (karakter khusus sudah diganti underscore)
                        safe_target = re.sub(r'[<>:"/\\|?*]', '_', target_pdf_stem)
                        if existing_stem == safe_target:
                            found_pdf = True
                            break
                    
                    if found_pdf:
                        print(f"[DEBUG] Strategy 2 MATCH (file content): {folder_path.name} (Diff: {diff:.2f}s)")
                        return folder_path
                
                print(f"[DEBUG] WARNING: PDF '{required_pdf_name}' tidak ditemukan di folder manapun.")
                print(f"[DEBUG] Strategy 3: Falling back to best time match.")
            
            # --- STRATEGI 3: BEST TIME MATCH (FALLBACK) ---
            best_match = candidates[0][1]
            print(f"[DEBUG] Strategy 3 MATCH (best time): {best_match.name}")
            return best_match
            
        print("[DEBUG] No matching backup folder found")
        return None
    except Exception as e:
        print(f"[BACKUP] Find Error: {e}")
        return None

@app.get("/download-backup")
async def download_backup(type: str, date: str, excel: str):
    """
    Download file backup.
    type: 'excel', 'pdf-original', 'pdf-result'
    date: created_at from history
    excel: excel_filename from history
    """
    folder = find_backup_folder(date, excel)
    if not folder:
        raise HTTPException(status_code=404, detail="Backup tidak ditemukan (mungkin sudah dihapus > 7 hari)")
    
    file_path = None
    filename = ""
    
    if type == 'excel':
        # FIX: Cari file .xlsx yang EXACT match dengan nama file history dulu
        # Jika tidak ada, baru fallback ke logic lama
        exact_path = folder / excel
        
        if exact_path.exists():
            file_path = exact_path
            filename = excel
            print(f"[DOWNLOAD] Found exact Excel match: {file_path}")
        else:
            # Fallback: Cari file .xlsx apapun di dalam folder
            # TAPI hindari file "filtered_" jika memungkinkan, kecuali jika itu satu-satunya
            files = list(folder.glob("*.xlsx"))
            
            # Filter out temporary files or filtered files if possible
            candidates = [f for f in files if not f.name.startswith("filtered_") and not f.name.startswith("~$")]
            
            if candidates:
                file_path = candidates[0]
                filename = file_path.name
                print(f"[DOWNLOAD] Found fallback Excel (non-filtered): {file_path}")
            elif files:
                file_path = files[0] # Last resort
                filename = file_path.name
                print(f"[DOWNLOAD] Found fallback Excel (any): {file_path}")
                
    elif type == 'pdf-original':
        file_path = folder / "original_labels.pdf"
        filename = "original_labels.pdf"
    elif type == 'pdf-result':
        # Cari file result_*.pdf
        files = list(folder.glob("result_*.pdf"))
        if files:
            file_path = files[0]
            filename = files[0].name
            
    if file_path and file_path.exists():
        return FileResponse(
            path=file_path, 
            filename=filename, 
            media_type='application/octet-stream'
        )
    
    raise HTTPException(status_code=404, detail=f"File {type} tidak ditemukan di backup")


@app.get("/generate-packing-list")
async def generate_packing_list(date: str, excel: str, pdf_name: str = None):
    """
    Generate Packing List Excel from original backup Excel.
    Columns: A: ID (Empty), B: SKU (MSKU), C: QTY (Sum), D: NO. PESANAN (Join).
    """
    folder = find_backup_folder(date, excel, required_pdf_name=pdf_name)
    if not folder:
        raise HTTPException(status_code=404, detail="Backup tidak ditemukan (mungkin sudah dihapus > 7 hari)")
    
    files = list(folder.glob("*.xlsx"))
    if not files:
        raise HTTPException(status_code=404, detail="Excel file tidak ditemukan di backup")
    
    # Prioritaskan file 'filtered_' jika ada
    filtered_files = list(folder.glob(f"filtered_*.xlsx"))
    if filtered_files:
        excel_path = filtered_files[0]
        print(f"[PACKING LIST] Using Filtered Excel source: {excel_path.name}")
    else:
        # Fallback cari file original yang sesuai nama (approx)
        safe_name = re.sub(r'[<>:"/\\|?*]', '_', excel.replace('.xlsx', ''))
        # Cari yang BUKAN filtered_
        orig_files = [f for f in files if not f.name.startswith('filtered_')]
        if orig_files:
             excel_path = orig_files[0] # Ambil yang pertama saja jika spesifik susah
             print(f"[PACKING LIST] Using Original Excel source: {excel_path.name}")
        else:
             excel_path = files[0] # Last resort
             print(f"[PACKING LIST] Using any Excel found: {excel_path.name}")
    
    try:
        # Read Excel - first pass to get column names
        df_temp = pd.read_excel(excel_path, nrows=0)
        df_temp.columns = df_temp.columns.astype(str).str.strip()
        
        # Identify ID column to force it as string (preserve large numbers like TikTok order IDs)
        col_id_temp = next((c for c in df_temp.columns if 'ID Pesanan' in c or 'Nomor Pesanan' in c), None)
        
        # Read Excel with ID column as string to preserve precision
        dtype_dict = {}
        if col_id_temp:
            dtype_dict[col_id_temp] = str
        
        df = pd.read_excel(excel_path, dtype=dtype_dict)
        
        # Normalize columns logic (copy from process_labels)
        df.columns = df.columns.astype(str).str.strip()
        
        # Identify columns
        col_msku = next((c for c in df.columns if 'MSKU' in c or 'No. Referensi SKU' in c), None)
        col_qty = next((c for c in df.columns if 'Jumlah' in c or 'QTY' in c or 'Qty' in c), None)
        col_id = next((c for c in df.columns if 'ID Pesanan' in c or 'Nomor Pesanan' in c), None)
        col_notes = next((c for c in df.columns if 'Catatan Pembeli' in c), None)
        
        if not (col_msku and col_qty and col_id):
             raise HTTPException(status_code=400, detail="Kolom MSKU, Jumlah, atau ID Pesanan tidak ditemukan di Excel")
        
        # Use openpyxl to read ID Pesanan values directly from cells (preserves exact values)
        try:
            wb = load_workbook(excel_path, data_only=False)
            ws = wb.active
            
            # Find ID Pesanan column index
            id_col_idx = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value and str(cell.value).strip() in [col_id, 'ID Pesanan', 'Nomor Pesanan', 'Order ID']:
                    id_col_idx = idx
                    break
            
            if id_col_idx:
                id_values = []
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=id_col_idx)
                    if cell.value is not None:
                        val = cell.value
                        if isinstance(val, (int, float)):
                            id_values.append(f'{int(val):d}')
                        else:
                            id_values.append(str(val).strip())
                    else:
                        id_values.append('')
                
                if len(id_values) == len(df):
                    df[col_id] = id_values
                    print(f"[PACKING LIST] ID Pesanan read via openpyxl: {id_values[:5]}")
            
            wb.close()
        except Exception as e:
            print(f"[PACKING LIST] openpyxl reading failed: {e}, using pandas values")
            # Fallback: clean up pandas-read values
            df[col_id] = df[col_id].astype(str).str.strip()
            df[col_id] = df[col_id].apply(lambda x: x[:-2] if x.endswith('.0') else x)
        
        # Forward-fill empty Order IDs to handle merged cells from Ginee Excel
        df[col_id] = df[col_id].replace(['', 'nan', 'None'], pd.NA).ffill()
        df[col_id] = df[col_id].fillna('')
        
        print(f"[PACKING LIST] Sample Order IDs from Excel: {df[col_id].head(5).tolist()}")
        
        # ============ PDF-BASED ORDER ID EXTRACTION ============
        # Extract correct order IDs from PDF backup and match with AWB
        pdf_files = list(folder.glob("*.pdf"))
        if pdf_files and 'AWB/No. Tracking' in df.columns:
            print(f"[PACKING LIST] Extracting order IDs from {len(pdf_files)} PDF backup files...")
            awb_to_pdf_order = {}  # AWB -> Correct Order ID from PDF
            
            for pdf_path in pdf_files:
                try:
                    pdf_doc = fitz.open(pdf_path)
                    for page_num in range(len(pdf_doc)):
                        page = pdf_doc[page_num]
                        text = page.get_text("text")
                        
                        # Extract AWB candidates from this page
                        awb_candidates = extract_all_awb_candidates(text)
                        
                        # Extract order ID candidates from this page
                        order_candidates = extract_order_ids(text)
                        
                        # Also look for TikTok-style numeric order IDs (15-18 digits)
                        tiktok_pattern = r'\b(\d{15,18})\b'
                        tiktok_matches = re.findall(tiktok_pattern, text)
                        for m in tiktok_matches:
                            if m not in order_candidates:
                                order_candidates.append(m)
                        
                        # Map each AWB found to the order IDs on the same page
                        for awb in awb_candidates:
                            awb_norm = normalize_awb(awb)
                            if order_candidates and awb_norm not in awb_to_pdf_order:
                                # Use the first/longest order ID found
                                best_order = max(order_candidates, key=len) if order_candidates else None
                                if best_order:
                                    awb_to_pdf_order[awb_norm] = best_order
                    
                    pdf_doc.close()
                except Exception as e:
                    print(f"[PACKING LIST] Error reading PDF {pdf_path.name}: {e}")
            
            print(f"[PACKING LIST] Found {len(awb_to_pdf_order)} AWB->OrderID mappings from PDF")
            
            # Now update DataFrame with correct order IDs from PDF
            col_awb = 'AWB/No. Tracking'
            if col_awb in df.columns and len(awb_to_pdf_order) > 0:
                def get_correct_order_id(row):
                    awb_val = row.get(col_awb, '')
                    if pd.notna(awb_val):
                        awb_norm = normalize_awb(awb_val)
                        if awb_norm in awb_to_pdf_order:
                            return awb_to_pdf_order[awb_norm]
                    # Fallback to original value
                    return row.get(col_id, '')
                
                original_ids = df[col_id].head(5).tolist()
                df[col_id] = df.apply(get_correct_order_id, axis=1)
                print(f"[PACKING LIST] Order IDs updated from PDF!")
                print(f"[PACKING LIST] Before: {original_ids}")
                print(f"[PACKING LIST] After: {df[col_id].head(5).tolist()}")
        
        print(f"[PACKING LIST] Final Sample Order IDs: {df[col_id].head(5).tolist()}")
              
        # Fetch Priority Bottom List
        priority_skus = set()
        try:
            p_data = await supabase_fetch("GET", "sku_priority_bottom?select=sku")
            priority_skus = {p['sku'].strip().upper() for p in p_data}
            print(f"[PACKING LIST] Loaded {len(priority_skus)} priority bottom SKUs")
        except Exception as e:
            print(f"[PACKING LIST] Failed to load priority list: {e}")

        # Process Grouping
        # Group by MSKU
        results = []
        # Handle nan in MSKU? groupby handles excludes nan by default.
        for msku, group in df.groupby(col_msku):
            total_qty = group[col_qty].sum()
            
        # Sub-group by Order ID to aggregate Qty per Order
            order_groups = group.groupby(col_id)[col_qty].sum()
            
            # Format: ID(Qty) sorted by ID
            detail_strs = []
            for oid, qty in sorted(order_groups.items()):
                 detail_strs.append(f"{oid}({int(qty)})")
            
            joined_details = '\n'.join(detail_strs)
            
            # Get Catatan Pembeli (Buyer Notes)
            # Collect unique notes for this SKU
            notes_list = []
            if col_notes:
                for note in group[col_notes]:
                    if pd.notna(note) and str(note).strip():
                        notes_list.append(str(note).strip())
            
            # Format: SKU\nCatatan Pembeli:notes (or - if empty)
            if notes_list:
                unique_notes = list(set(notes_list))
                notes_str = '; '.join(unique_notes)  # Join multiple notes with semicolon
                sku_with_notes = f"{msku}\nCatatan Pembeli:{notes_str}"
            else:
                sku_with_notes = f"{msku}\nCatatan Pembeli:-"
            
            results.append({
                'SKU': sku_with_notes,
                'RAW_SKU': msku, # Keep raw for lookup
                'QTY': total_qty,
                'NO. PESANAN': joined_details
            })
            
        # SORTING LOGIC:
        # Step 1: Sort A-Z by SKU (RAW_SKU column) - using natural sort (remove special chars)
        # Step 2: Apply priority bottom as the LAST step (items in priority_skus go to bottom)
        
        # Helper function for natural sort (matches Excel A-Z behavior)
        def natural_sort_key(item):
            raw_sku = str(item['RAW_SKU']).strip().upper()
            # Remove non-alphanumeric characters for natural sorting (matches Excel behavior)
            # This makes "PULPENC" sort correctly between "PULPEN-B" and "PULPEN-G"
            clean_sku = re.sub(r'[^A-Z0-9]', '', raw_sku)
            return clean_sku
        
        # First: Sort alphabetically by SKU (A-Z) using natural sort
        results.sort(key=natural_sort_key)
        print(f"[PACKING LIST] Sorted A-Z by SKU (natural). First 3: {[r['RAW_SKU'] for r in results[:3]]}")
        
        # Then: Separate into normal and bottom items, preserving A-Z order within each group
        normal_items = []
        bottom_items = []
        for item in results:
            raw_sku = str(item['RAW_SKU']).strip().upper()
            if raw_sku in priority_skus:
                bottom_items.append(item)
            else:
                normal_items.append(item)
        
        # Final result: normal items (A-Z) + bottom items (A-Z)
        results = normal_items + bottom_items
        print(f"[PACKING LIST] Priority applied. Normal: {len(normal_items)}, Bottom: {len(bottom_items)}")
        
        output_df = pd.DataFrame(results)
        
        # --- LOOKUP CUSTOM ID FROM SUPABASE ---
        # --- LOOKUP CUSTOM ID FROM SUPABASE ---
        print("[PACKING LIST] Fetching SKU Mappings for ID Lookup...")
        
        def normalize_sku_key(s):
             if not s: return ""
             return str(s).strip().upper()

        sku_map = {}
        try:
             # Reuse existing fetch logic (it's async, but we are in async def)
             mappings = await get_sku_mappings()
             # Create lookup: Normalized SKU -> CustomID
             # Mappings in DB is {id, sku} where 'id' is actually the custom_id due to get_sku_mappings transformation
             
             for m in mappings:
                 norm_key = normalize_sku_key(m['sku'])
                 if norm_key: # Skip empty keys
                     sku_map[norm_key] = str(m['id']) # 'id' here is custom_id e.g. "0005"
             
             print(f"[PACKING LIST] Loaded {len(sku_map)} mappings (Normalized). Sample: {list(sku_map.items())[:3]}")
        except Exception as e:
             print(f"[PACKING LIST] Failed to load mappings: {e}")
             sku_map = {}

        # Apply Lookup
        def get_custom_id(row):
            # Try exact match on raw SKU
            raw_sku = str(row['RAW_SKU'])
            norm_sku = normalize_sku_key(raw_sku)
            
            if norm_sku in sku_map:
                return sku_map[norm_sku]
            
            # Debug failures for specific SKU mentioned by user
            if 'BAG-DCB' in norm_sku:
                print(f"[DEBUG LOOKUP FAIL] SKU: '{raw_sku}' -> Norm: '{norm_sku}' caused fail. Map has '{norm_sku}'? {norm_sku in sku_map}")
            
            return "-"

        output_df['ID'] = output_df.apply(get_custom_id, axis=1)
        
        # Ensure correct column order
        output_df = output_df[['ID', 'SKU', 'QTY', 'NO. PESANAN']]
        
        # Fetch Formatting Rules
        format_colors = []
        format_styles = []
        try:
             format_colors = await supabase_fetch("GET", "sku_formatting_colors?select=keyword,color_code")
             format_styles = await supabase_fetch("GET", "sku_formatting_styles?select=keyword,font_size,is_bold")
             print(f"[PACKING LIST] Loaded {len(format_colors)} color rules, {len(format_styles)} style rules.")
        except Exception as e:
             print(f"[PACKING LIST] Failed to load formatting rules: {e}")

        # Fetch Column Settings
        column_settings = {}
        try:
            cs_data = await supabase_fetch("GET", "sku_column_settings?select=*")
            for cs in cs_data:
                column_settings[cs['column_name']] = cs
            print(f"[PACKING LIST] Loaded {len(column_settings)} column settings.")
        except Exception as e:
            print(f"[PACKING LIST] Failed to load column settings: {e}")

        # Default column settings if not in DB
        default_columns = {
            'ID': {'column_width': 9.64, 'font_size': 16, 'font_name': 'Rockwell', 'is_bold': False, 'text_align': 'center'},
            'SKU': {'column_width': 40.55, 'font_size': 16, 'font_name': 'Rockwell', 'is_bold': False, 'text_align': 'left'},
            'QTY': {'column_width': 15.82, 'font_size': 20, 'font_name': 'Rockwell', 'is_bold': True, 'text_align': 'center'},
            'NO. PESANAN': {'column_width': 29.91, 'font_size': 12, 'font_name': 'Rockwell', 'is_bold': False, 'text_align': 'center'}
        }

        # Merge with DB settings
        for col_name, defaults in default_columns.items():
            if col_name not in column_settings:
                column_settings[col_name] = defaults

        # Determine output stream
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            workbook = writer.book
            worksheet = workbook.add_worksheet('Sheet1')
            
            # --- DEFINE FORMATS ---
            # Header Format (Fixed)
            header_format = workbook.add_format({
                'font_name': 'Times New Roman', 'font_size': 20, 'bold': True,
                'align': 'center', 'valign': 'vcenter', 'bg_color': '#666666',
                'font_color': 'white', 'border': 1, 'text_wrap': True
            })
            
            # Dynamic Column Formats
            col_formats = {}
            col_order = ['ID', 'SKU', 'QTY', 'NO. PESANAN']
            for col_name in col_order:
                cs = column_settings.get(col_name, default_columns.get(col_name, {}))
                fmt = workbook.add_format({
                    'font_name': cs.get('font_name', 'Rockwell'),
                    'font_size': cs.get('font_size', 16),
                    'bold': cs.get('is_bold', False),
                    'align': cs.get('text_align', 'center'),
                    'valign': 'vcenter',
                    'border': 1,
                    'text_wrap': True
                })
                col_formats[col_name] = fmt

            # Assign specific format references for easier code use
            fmt_id = col_formats.get('ID')
            fmt_sku_base = col_formats.get('SKU')
            fmt_qty = col_formats.get('QTY')
            fmt_order = col_formats.get('NO. PESANAN')

            # Dynamic Color Formats Cache
            # Map: color_code -> Format Object (Base SKU + Color)
            color_formats = {} 
            for rule in format_colors:
                cc = rule['color_code']
                if cc not in color_formats:
                    cs = column_settings.get('SKU', default_columns.get('SKU', {}))
                    f = workbook.add_format({
                        'font_name': cs.get('font_name', 'Rockwell'),
                        'font_size': cs.get('font_size', 16),
                        'align': cs.get('text_align', 'left'),
                        'valign': 'vcenter',
                        'border': 1,
                        'text_wrap': True
                    })
                    f.set_font_color(cc)
                    color_formats[cc] = f

            # Dynamic Style Formats Cache
            # Map: (size, bold) -> Format Object (Just Font Metrics)
            style_formats = {}
            for rule in format_styles:
                key = (rule['font_size'], rule['is_bold'])
                if key not in style_formats:
                    # For rich string segments, we just define font props used in write_rich_string
                    # These segment formats do NOT need alignment/border usually, just font.
                    cs = column_settings.get('SKU', default_columns.get('SKU', {}))
                    f = workbook.add_format({'font_name': cs.get('font_name', 'Rockwell')})
                    if rule['font_size']: f.set_font_size(rule['font_size'])
                    if rule['is_bold']: f.set_bold()
                    style_formats[key] = f

            # --- WRITE HEADER ---
            headers = ['ID', 'SKU', 'QTY', 'NO. PESANAN']
            for col_num, header in enumerate(headers):
                worksheet.write(0, col_num, header, header_format)
            
            # Set Column Widths from Settings
            col_letters = ['A', 'B', 'C', 'D']
            for i, col_name in enumerate(col_order):
                cs = column_settings.get(col_name, default_columns.get(col_name, {}))
                width = cs.get('column_width', 20)
                worksheet.set_column(f'{col_letters[i]}:{col_letters[i]}', width)

            # --- WRITE DATA ---
            # Output DF: ID, SKU, QTY, NO. PESANAN
            # Iterating manually to apply rich text
            
            current_row = 1
            for index, row in output_df.iterrows():
                # 1. ID
                worksheet.write(current_row, 0, row['ID'], fmt_id)
                
                # 2. SKU (Rich Text Logic)
                sku_text = str(row['SKU'])
                
                # A. Determine Base Cell Color (from Color Rules)
                # Check formatting_colors
                # Find first matching rule? Or specific priority?
                # User: "jika ada sku ... mengandung kata BLUE maka ... teksnya warna merah"
                base_format = fmt_sku_base
                matched_color = None
                
                for rule in format_colors:
                    if rule['keyword'].upper() in sku_text.upper():
                        cc = rule['color_code']
                        if cc in color_formats:
                            base_format = color_formats[cc]
                            matched_color = cc
                        break # First match wins for cell color
                
                # B. Determine Rich Text Segments (from Style Rules)
                # This is complex if multiple keywords exist.
                # Simplification: Find ALL style keywords present.
                # Flatten the string into [ (text, format), (text, format), ... ]
                
                segments = []
                
                # Identify styled tokens
                # We need a list of (start_index, end_index, format_obj)
                style_matches = []
                for rule in format_styles:
                    kw = rule['keyword']
                    if not kw: continue
                    # Find all occurrences
                    start = 0
                    while True:
                        idx = sku_text.upper().find(kw.upper(), start)
                        if idx == -1: break
                        
                        # Store match
                        k_fmt = style_formats.get((rule['font_size'], rule['is_bold']))
                        style_matches.append({
                            'start': idx,
                            'end': idx + len(kw),
                            'fmt': k_fmt,
                            'priority': len(kw) # Longer keywords priority?
                        })
                        start = idx + 1
                
                # If no style matches, simple write
                if not style_matches:
                    worksheet.write(current_row, 1, sku_text, base_format)
                else:
                    # Sort matches by start position
                    style_matches.sort(key=lambda x: x['start'])
                    
                    # Handle overlaps? Simplest: First come first serve or verify no overlap.
                    # We will construct segments linearly.
                    
                    final_segments = []
                    last_pos = 0
                    
                    # Deduplicate/Resolve overlaps linearly
                    for match in style_matches:
                        if match['start'] < last_pos:
                            continue # Skip overlapping
                        
                        # Text before match - MUST include format for proper font!
                        if match['start'] > last_pos:
                            final_segments.append(base_format)  # Apply base format to non-styled text
                            final_segments.append(sku_text[last_pos:match['start']])
                        
                        # The Match - Add Format Object then Text
                        final_segments.append(match['fmt'])
                        final_segments.append(sku_text[match['start']:match['end']])
                        
                        last_pos = match['end']
                    
                    # Remaining text after last match - MUST include format!
                    if last_pos < len(sku_text):
                        final_segments.append(base_format)  # Apply base format to remaining text
                        final_segments.append(sku_text[last_pos:])
                    
                    # Write Rich String
                    # Args: row, col, *segments, cell_format (cell_format applies to cell properties like border)
                    # Each text segment should be preceded by its format
                    
                    try:
                        worksheet.write_rich_string(current_row, 1, *final_segments, base_format)
                    except Exception as e:
                        print(f"Rich Text Error: {e}")
                        worksheet.write(current_row, 1, sku_text, base_format)

                # 3. QTY
                worksheet.write(current_row, 2, row['QTY'], fmt_qty)
                
                # 4. NO PESANAN
                worksheet.write(current_row, 3, row['NO. PESANAN'], fmt_order)
                
                # Set Row Height (auto or fixed?)
                # user didn't ask, but multiline needs height. 
                # Excel usually auto-fits height if text wrap is on, but xlsxwriter sometimes needs explicit or separate step.
                # We'll leave default for now (xlsxwriter auto height is implied effectively if valid).
                # Actually, set row height to allow multi-line visibility if needed?
                # Using text_wrap=True usually handles it in Excel viewer.
                
                current_row += 1
                
            # Set Row Height for Header (Auto or fixed? Size 20 font needs space)
            worksheet.set_row(0, 30)
            
            # PAGE SETUP (from VBA specification + user top margin)
            # Margins: left/right 0.25", top 1.8" (pas untuk QR+teks header, jarak kecil ke tabel), bottom 0.75"
            worksheet.set_margins(left=0.25, right=0.25, top=1.8, bottom=0.75)
            
            # Print Title Rows: Row 1 repeats on each page
            worksheet.repeat_rows(0, 0)
            
            # Print Area: A1:D{lastRow}
            last_row = len(output_df) + 1  # +1 because header is row 1
            worksheet.print_area(f'A1:D{last_row}')

            # --- SHEET 2: UNIQUE ORDER IDs ---
            # Extract unique order IDs from NO. PESANAN column (column D)
            # Each cell may have multiple lines like: "582160320679544823(2)\n582160347278706463(1)..."
            # We split by newlines, extract ID before "(", and deduplicate
            
            unique_order_ids = set()
            for val in output_df['NO. PESANAN']:
                if pd.isna(val): continue
                lines = str(val).split('\n')
                for line in lines:
                    line = line.strip()
                    if not line: continue
                    # Remove quantity suffix like "(2)" or "(1)"
                    # Pattern: ID(qty) -> extract just ID
                    if '(' in line:
                        order_id = line.split('(')[0].strip()
                    else:
                        order_id = line.strip()
                    if order_id:
                        unique_order_ids.add(order_id)
            
            # Create Sheet2 and write unique IDs
            sheet2 = workbook.add_worksheet('Sheet2')
            sheet2.write(0, 0, 'ID Pesanan')
            # Write starting from A2 (row index 1)
            row_idx = 1
            for order_id in sorted(unique_order_ids):
                sheet2.write(row_idx, 0, order_id)
                row_idx += 1
            
            print(f"[PACKING LIST] Sheet2 created with {len(unique_order_ids)} unique order IDs.")

            # --- SHEET 3: UNIQUE AWBS ---
            # Cari kolom AWB di DataFrame asli
            col_awb = next((c for c in df.columns if 'AWB' in str(c).upper() or 'TRACKING' in str(c).upper() or 'RESI' in str(c).upper()), None)
            if col_awb:
                unique_awbs = set()
                for val in df[col_awb]:
                    if pd.notna(val):
                        awb = str(val).strip()
                        if awb:
                            unique_awbs.add(awb)
                
                sheet3 = workbook.add_worksheet('Sheet3')
                sheet3.write(0, 0, 'AWB/No. Tracking')
                row_idx = 1
                for awb in sorted(unique_awbs):
                    sheet3.write(row_idx, 0, awb)
                    row_idx += 1
                print(f"[PACKING LIST] Sheet3 created with {len(unique_awbs)} unique AWBs.")
            else:
                print(f"[PACKING LIST] Sheet3 skipped, AWB/No. Tracking column not found.")

            # 1. Generate QR Image locally (no internet needed)
            qr_file_path = None
            try:
                import qrcode
                # Use pdf_name for QR data (without extension)
                qr_data = (pdf_name if pdf_name else excel).replace('.pdf', '').replace('.xlsx', '').replace('.xls', '')
                print(f"[PACKING LIST] Generating QR locally for: {qr_data}")

                qr = qrcode.QRCode(
                    version=None,
                    error_correction=qrcode.constants.ERROR_CORRECT_M,
                    box_size=3,
                    border=1,
                )
                qr.add_data(qr_data)
                qr.make(fit=True)

                img = qr.make_image(fill_color="black", back_color="white")

                # Save to temp file
                temp_filename = f"temp_qr_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.png"
                img.save(temp_filename)
                qr_file_path = temp_filename
                print(f"[PACKING LIST] QR saved to: {temp_filename}")

            except Exception as e:
                print(f"[PACKING LIST] QR Generation Error: {e}")

            # 2. Set Header: &[Picture]        nama file PDF (tanpa extension)
            # Use pdf_name if provided, otherwise fallback to excel
            header_name = pdf_name if pdf_name else excel
            header_display = header_name.replace('.pdf', '').replace('.xlsx', '').replace('.xls', '')
            if qr_file_path:
                # QR di baris pertama, teks nama di baris kedua (center, di bawah QR)
                header_str = f'&C&G\n&"Courier New,Bold"&16{header_display}'
                worksheet.set_header(header_str, {'image_center': qr_file_path, 'header_margin': 0.5})
            else:
                # Fallback text only if QR fails
                header_str = f'&C&"Courier New,Bold"&16[NO QR] {header_display}'
                worksheet.set_header(header_str, {'header_margin': 0.5})
            
            # 3. Set Footer: Page &[Page] of &[Pages] (Arial Black, Size 28)
            footer_str = '&C&"Arial Black,Regular"&28Page &P of &N'
            worksheet.set_footer(footer_str)
            
        output.seek(0)
        
        # Cleanup temp QR file
        if qr_file_path and os.path.exists(qr_file_path):
            try:
                os.remove(qr_file_path)
            except:
                pass
        
        # Use pdf_name if provided, otherwise fallback to excel name
        base_name = pdf_name if pdf_name else excel.replace('.xlsx', '').replace('.xls', '')
        filename = f"{base_name}.xlsx"

        # --- BACKUP TO LOCAL FOLDER ---
        try:
            # SAVE TO PROCESS SPECIFIC FOLDER (folder variable from find_backup_folder)
            if folder and folder.exists():
                local_backup_path = folder / filename
                
                # Save BytesIO content to file
                output.seek(0)
                with open(local_backup_path, 'wb') as f:
                    f.write(output.read())
                
                print(f"[PACKING LIST] Saved backup to: {local_backup_path}")
            else:
                 print(f"[PACKING LIST] Backup folder not found, skipping local save.")

        except Exception as e:
            print(f"[PACKING LIST] Failed to save local backup: {e}")
        # -----------------------------
        
        output.seek(0) # Reset pointer for response
        
        return StreamingResponse(
            output, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# Jalankan cleanup saat startup
cleanup_old_backups()



# Register Unicode fonts
# Try common font paths
FONT_PATHS = [
    # Windows
    "C:/Windows/Fonts/seguiemj.ttf",  # Segoe UI Emoji
    "C:/Windows/Fonts/arial.ttf",
    "C:/Windows/Fonts/arialuni.ttf",  # Arial Unicode MS
    # Linux
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
]

UNICODE_FONT = None
for font_path in FONT_PATHS:
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont('UnicodeFont', font_path))
            UNICODE_FONT = 'UnicodeFont'
            print(f"[FONT] Registered Unicode font: {font_path}")
            break
        except Exception as e:
            print(f"[FONT] Failed to register {font_path}: {e}")

if not UNICODE_FONT:
    print("[FONT] WARNING: No Unicode font found, falling back to Helvetica")
    UNICODE_FONT = 'Helvetica'


def clean_text_for_pdf(text):
    """
    Bersihkan teks untuk PDF:
    - Ganti newline/carriage return dengan spasi
    - Hapus karakter control yang tidak bisa diprint
    - Pertahankan karakter Unicode seperti ♡
    """
    if not text:
        return ""
    
    # Replace line breaks with space
    text = text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
    
    # Remove control characters (except normal space)
    # Keep only printable characters, Unicode letters, numbers, punctuation
    cleaned = []
    for char in text:
        # Keep if: printable ASCII, or Unicode char (not control)
        if char == ' ' or (ord(char) >= 32 and ord(char) != 127):
            cleaned.append(char)
        else:
            cleaned.append(' ')  # Replace control chars with space
    
    # Remove multiple spaces
    result = ' '.join(''.join(cleaned).split())
    return result


def fix_excel_numeric_ids(df):
    """
    Mencegah long numeric IDs (seperti ID Pesanan) berubah jadi scientific notation (5.8E+17).
    Mengubah kolom yang mengandung keyword 'ID', 'Pesanan', 'AWB', 'Tracking', 'Resi' menjadi string murni.
    """
    for col in df.columns:
        col_str = str(col).upper()
        if any(k in col_str for k in ['ID', 'PESANAN', 'AWB', 'TRACKING', 'RESI']):
            # Convert scientific/float to full string integer
            df[col] = df[col].apply(lambda x: '{:.0f}'.format(x) if isinstance(x, (int, float)) and pd.notna(x) else x)
            # Force as string and clean up 'nan' strings
            df[col] = df[col].astype(str).replace(['nan', 'NaN', 'None', 'NAN', 'null'], None)
    return df


def normalize_awb(awb):
    """Normalize AWB: hapus spasi, dash, # prefix, dan uppercase."""
    if not awb or pd.isna(awb):
        return ""
    # Strip # prefix (used to preserve large numbers as text in Excel)
    result = str(awb).strip()
    if result.startswith('#'):
        result = result[1:]
    return re.sub(r'[\s\-\.]+', '', result).upper().strip()


def extract_all_awb_candidates(text):
    """Ekstrak semua kandidat AWB dari teks halaman PDF."""
    candidates = []
    
    # Pattern yang lebih lengkap untuk berbagai ekspedisi
    patterns = [
        # Shopee Express
        r'\bSPXID\d{10,18}\b',            # SPXID + 10-18 digit
        r'\bSPX[A-Z]{2}\d{10,}\b',       # SPXID variant (SPXPH, SPXMY, dll)
        
        # J&T Express / Cargo
        r'\bJT\d{10,18}\b',              # JT + 10-18 digit
        r'\bJX\d{10,18}\b',              # JX (Cargo)
        r'\bJD\d{10,18}\b',              # JD variant
        r'\b5\d{11}\b',                   # J&T Cargo (TikTok): 12 digit starting with 5
        
        # SiCepat
        r'\b00\d{10,15}\b',              # Dimulai dengan 00
        r'\b000\d{9,14}\b',              # Dimulai dengan 000
        
        # Ninja Xpress / Ninja Van (EXTENDED)
        r'\bNVID\d{8,15}\b',             # NVID prefix
        r'\bNV\d{10,18}\b',              # NV prefix
        r'\bNINJA\d{8,15}\b',            # NINJA prefix
        r'\bNJVTT\d{8,18}\b',            # NJVTT prefix (Ninja Van Thailand/Transit)
        r'\bNJV[A-Z]{0,3}\d{8,18}\b',    # NJV + 0-3 huruf + digit (catch all Ninja variants)
        
        # Grab Express
        r'\bGB\d{10,18}\b',              # GB prefix
        r'\bGRAB\d{8,15}\b',             # GRAB prefix
        
        # AnterAja / Tokopedia / GoTo Logistics
        r'\bTKP\d{8,18}\b',              # TKP prefix
        r'\bANT\d{10,18}\b',             # ANT prefix
        r'\bGTL\d{8,15}\b',              # GTL prefix (GoTo Logistics)
        
        # Lazada / LEX
        r'\bID\d{10,20}\b',              # ID prefix (LEX ID)
        r'\bLEX\d{8,18}\b',              # LEX prefix
        r'\bLAZ\d{8,18}\b',              # LAZ prefix
        r'\bLXAD[\-]?\d{8,18}\b',        # Lazada LXAD (LXAD-5107623190)
        
        # SAP Express
        r'\bSAP\d{10,18}\b',             # SAP prefix
        
        # Pos Indonesia
        r'\bPOS\d{10,18}\b',             # POS prefix
        
        # Blibli
        r'\bBLI[A-Z0-9]{8,20}\b',        # BLIJC, BLI, dll
        
        # JNE
        r'\bJNE\d{10,18}\b',             # JNE prefix
        r'\bJNAP[\-]?\d{8,18}\b',         # JNAP (JNE via Lazada, contoh: JNAP-0251117042)
        r'\b\d{10}[A-Z]{2,4}\d{4,6}\b',  # Format JNE lama (10digit + huruf + digit)
        r'\bCM\d{8,18}\b',               # CM prefix (User requested)
        r'\b11\d{10,18}\b',              # 11 prefix (User requested)
        
        # Kerry Express
        r'\bKERRY\d{8,15}\b',            # KERRY prefix
        r'\bKE\d{10,18}\b',              # KE prefix
        
        # Lion Parcel
        r'\bLION\d{8,15}\b',             # LION prefix
        r'\bLP\d{10,18}\b',              # LP prefix
        
        # Wahana
        r'\bWHN\d{8,15}\b',              # WHN prefix
        
        # RPX
        r'\bRPX\d{8,15}\b',              # RPX prefix
        
        # Tiki
        r'\bTIKI\d{8,15}\b',             # TIKI prefix
        
        # ========== GENERIC AUTO-DETECT PATTERNS ==========
        # Patterns ini akan menangkap format ekspedisi BARU yang belum dikenal
        # Urutkan dari yang paling spesifik ke yang paling umum
        
        r'\b[A-Z]{2,5}\d{8,20}\b',       # 2-5 huruf + 8-20 digit (PALING UMUM - tangkap semua)
        r'\b[A-Z]{1,2}\d{12,20}\b',      # 1-2 huruf + 12-20 digit
        # FIX: Regex ini sebelumnya menangkap kata biasa seperti "HIGHLIGHTER" karena allow [A-Z0-9] tanpa wajib digit.
        # Sekarang WAJIB ada angka minimal 1 digit.
        # Format: 3-6 huruf prefix + (sequence berisi minimal 1 angka)
        # FIX: Regex ini sebelumnya menangkap kata biasa seperti "HIGHLIGHTER" karena allow [A-Z0-9] tanpa wajib digit.
        # Sekarang WAJIB ada angka minimal 1 digit.
        # Format: 3-6 huruf prefix + (sequence berisi minimal 1 angka)
        r'\b[A-Z]{3,6}(?=[A-Z0-9]*\d)[A-Z0-9]{8,15}\b',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        candidates.extend([normalize_awb(m) for m in matches])
    
    # Unique while preserving order
    seen = set()
    unique = []
    for c in candidates:
        if c and c not in seen and len(c) >= 8:  # Min 8 karakter
            seen.add(c)
            unique.append(c)
    
    return unique


def extract_order_ids(text):
    """
    Ekstrak Order ID / ID Pesanan dari teks PDF.
    Menggunakan pola spesifik agar tidak mengambil data sampah/barcode lain.
    """
    import re
    if not text: return []
    
    order_ids = []
    
    # Pattern spesifik (diambil dari Script Lama)
    patterns = [
        # TikTok Order ID: "TT Order ID :" prefix + 15-20 digit
        r'TT\s*Order\s*ID\s*[:\uff1a]\s*(\d{15,20})',
        
        # Lazada Order ID: "Nomor Order :" prefix + 16 digit
        r'Nomor\s*Order\s*[:\-]?\s*(\d{15,18})',
        
        # Lazada Order ID murni numerik: 16 digit mulai 268x atau 277x (format ID Lazada Indonesia)
        r'\b(268\d{13})\b',
        r'\b(277\d{13})\b',
        
        # Shopee Order ID: 6 digit (YYMMDD) + 6-12 alfanumerik (contoh: 230909xxxx)
        r'\b(\d{6}[A-Z0-9]{6,12})\b',
        
        # Pattern eksplisit setelah kata kunci
        r'(?:Order\s*ID|No\.?\s*Pesanan|ID\s*Pesanan)\s*[:\-]?\s*([A-Z0-9]{12,20})',
        
        # Tokopedia/Shopee long order ID (Format gabungan huruf angka spesifik)
        r'\b([A-Z]{2,6}\d{10,18})\b',
        
        # Pure numeric TikTok order ID (15-20 digits starting with 57/58)
        # Kita batasi agar tidak mengambil barcode produk (biasanya 13 digit ean)
        r'\b(5[7-9]\d{13,18})\b',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for m in matches:
            # Bersihkan hasil
            clean_m = str(m).strip().upper().replace(' ', '').replace('-', '')
            if len(clean_m) >= 10:
                order_ids.append(clean_m)
    
    return list(dict.fromkeys(order_ids))  # dedup tapi PERTAHANKAN urutan


def find_matching_awb(candidates, known_awbs, page_num=None):
    """
    Cari AWB yang match dengan Excel.
    Returns: matched AWB dari known_awbs, atau None jika tidak ditemukan
    """
    # Normalize semua known AWBs untuk comparison
    known_normalized = {normalize_awb(awb): awb for awb in known_awbs}
    
    for candidate in candidates:
        norm_candidate = normalize_awb(candidate)
        
        # 1. Exact match (normalized) - PRIORITAS TERTINGGI
        if norm_candidate in known_normalized:
            print(f"  [OK] MATCH (exact): {candidate} -> {known_normalized[norm_candidate]}")
            return known_normalized[norm_candidate]
        
        # 2. Partial match - KETAT: hanya jika perbedaan panjang <= 4 karakter
        # Ini untuk menghindari false positive dari substring yang terlalu pendek
        for norm_known, original_known in known_normalized.items():
            # Cek jika candidate mengandung known (atau sebaliknya)
            # TAPI hanya jika perbedaan panjang tidak terlalu besar
            if norm_candidate in norm_known:
                len_diff = len(norm_known) - len(norm_candidate)
                if 0 < len_diff <= 4:  # Candidate adalah substring dari known, max 4 char lebih pendek
                    print(f"  [OK] MATCH (partial-in): {candidate} -> {original_known}")
                    return original_known
            elif norm_known in norm_candidate:
                len_diff = len(norm_candidate) - len(norm_known)
                if 0 < len_diff <= 4:  # Known adalah substring dari candidate, max 4 char lebih pendek
                    print(f"  ✓ MATCH (partial-in): {candidate} -> {original_known}")
                    return original_known
        
        # 3. Suffix match - cek jika 12+ karakter terakhir sama (lebih ketat dari 10)
        for norm_known, original_known in known_normalized.items():
            if len(norm_candidate) >= 12 and len(norm_known) >= 12:
                if norm_candidate[-12:] == norm_known[-12:]:
                    print(f"  [OK] MATCH (suffix-12): {candidate} -> {original_known}")
                    return original_known
    
    # Log jika tidak ditemukan
    if page_num is not None:
        print(f"  [NO] NO MATCH for page {page_num}. Top candidates: {candidates[:5]}")
    
    return None


def find_clear_start_position(page):
    """
    Cari posisi Y untuk memulai area putih (clear area).
    Logika:
    1. Cari teks "TANPA VIDEO UNBOXING, KOMPLIEN TIDAK DITERIMA"
    2. Jika ada "Catatan Pembeli:" di bawahnya, start dari bawah Catatan Pembeli
    3. Jika tidak ada Catatan Pembeli, start dari bawah teks TANPA VIDEO...
    4. Fallback 1: Cari Header Tabel Asli (Nama Produk / SKU / Qty) -> Start dari ATAS header ini (timpa)
    5. Fallback 2: ke y=300 jika tidak ditemukan
    """
    text_instances = page.get_text("dict")
    blocks = text_instances.get("blocks", [])
    
    lipat_y = None
    unboxing_y = None
    catatan_y = None
    table_header_y = None
    
    # Cari posisi teks
    for block in blocks:
        if block.get("type") == 0:  # Text block
            for line in block.get("lines", []):
                line_text = ""
                # Bbox line: [x0, y0, x1, y1] -> Top=y0, Bottom=y1
                line_top = 9999
                line_bottom = 0
                
                for span in line.get("spans", []):
                    line_text += span.get("text", "")
                    bbox = span.get("bbox", [0, 0, 0, 0])
                    line_top = min(line_top, bbox[1])
                    line_bottom = max(line_bottom, bbox[3])
                
                line_text_upper = line_text.upper()
                
                # Cek teks LIPAT, SEMBUNYIKAN (Blibli)
                if "LIPAT, SEMBUNYIKAN" in line_text_upper or "PISAHKAN BAGIAN INI" in line_text_upper:
                    lipat_y = line_bottom
                
                # Cek teks TANPA VIDEO UNBOXING
                if "TANPA VIDEO UNBOXING" in line_text_upper:
                    unboxing_y = line_bottom
                
                # Cek teks Catatan Pembeli
                if "CATATAN PEMBELI" in line_text_upper and (unboxing_y is not None or lipat_y is not None):
                    catatan_y = line_bottom

                # Header Tabel
                is_header_match = ("NAMA PRODUK" in line_text_upper or "PRODUCT NAME" in line_text_upper or 
                                   ("#" in line_text and "NAMA" in line_text_upper) or
                                   ("SKU" in line_text_upper and "QTY" in line_text_upper))
                
                if is_header_match:
                     if table_header_y is None or line_top < table_header_y:
                         table_header_y = line_top
    
    # Tentukan posisi start clear area
    if lipat_y is not None:
        clear_start = lipat_y + 3
    elif catatan_y is not None:
        clear_start = catatan_y + 5
    elif unboxing_y is not None:
        clear_start = unboxing_y + 5
    elif table_header_y is not None:
        clear_start = table_header_y - 2
    else:
        clear_start = 300
    
    return clear_start


def calculate_available_rows(page_height, clear_start_y, row_height=18, margin_bottom=25):
    """
    Hitung berapa baris yang muat di area yang tersedia.
    margin_bottom=25 → ruang untuk No:xxx dan indikator halaman.
    """
    available_height = page_height - clear_start_y - margin_bottom + 5
    max_rows = int(available_height / row_height) - 1
    return max(0, max_rows)


def is_junk_page(page):
    """
    Deteksi apakah halaman adalah 'sampah' yang harus dihapus.
    """
    text = page.get_text()
    text_upper = text.upper()
    text_lower = text.lower()
    
    # ===== DETEKSI TABEL LAMA (dari PDF marketplace asli) =====
    has_old_table = any([
        re.search(r'#\s+SKU\s+Qty', text, re.IGNORECASE),
        re.search(r'Product\s+Name', text, re.IGNORECASE),  # Header tabel Tokopedia
        re.search(r'Seller\s+SKU', text, re.IGNORECASE),    # Header tabel marketplace
        re.search(r'No\.?\s+SKU\s+Qty', text, re.IGNORECASE),
        re.search(r'Nama\s+Produk', text, re.IGNORECASE),   # Header tabel Lazada
        re.search(r'Produk\s+dalam\s+paket\s+ini', text, re.IGNORECASE), # Header tabel Blibli
        re.search(r'No\.\s+item\s+pesanan', text, re.IGNORECASE), # Kolom Blibli
    ])
    
    # ===== DETEKSI TABEL BARU (dari kita) =====
    has_new_table = re.search(r'MSKU\s+Qty', text, re.IGNORECASE) is not None
    
    # ===== DETEKSI PATTERN KHUSUS =====
    has_in_transit = re.search(r'In\s+transit\s+by\s*:', text, re.IGNORECASE) is not None
    has_qty_total = re.search(r'Qty\s*Total\s*:', text, re.IGNORECASE) is not None
    has_marketplace_footer = any([
        re.search(r'tokopedia.*shop', text, re.IGNORECASE),  # "tokopedia | Shop"
        re.search(r'shopee.*express', text, re.IGNORECASE),  # Shopee footer
        'order id:' in text_lower or 'package id:' in text_lower,  # ID tanpa label
    ])
    
    # ===== HALAMAN LANJUTAN KITA - JANGAN HAPUS =====
    is_our_continuation = 'LANJUTAN AWB:' in text_upper
    if is_our_continuation:
        return False
    
    # ===== DETEKSI INFO PENGIRIMAN ESENSIAL =====
    has_awb_barcode = any([
        re.search(r'SPXID\d{10,}', text_upper),   # Shopee Express
        re.search(r'JT\d{10,}', text_upper),       # J&T
        re.search(r'JX\d{10,}', text_upper),       # J&T Cargo
        re.search(r'5\d{11}', text),               # J&T Cargo TikTok
        re.search(r'00\d{10,}', text_upper),       # SiCepat
        re.search(r'NVID\d{8,}', text_upper),      # Ninja
        re.search(r'NJVTT\d{8,}', text_upper),     # Ninja Van
        re.search(r'NJV[A-Z]{0,3}\d{8,}', text_upper),  # All Ninja Van
        re.search(r'TKP\d{8,}', text_upper),       # Tokopedia
        re.search(r'ANT\d{10,}', text_upper),      # AnterAja
        re.search(r'GTL\d{8,}', text_upper),       # GoTo Logistics
        re.search(r'JNE\d{10,}', text_upper),      # JNE
        re.search(r'LION\d{8,}', text_upper),      # Lion Parcel
        re.search(r'LP\d{10,}', text_upper),       # Lion Parcel
        re.search(r'KERRY\d{8,}', text_upper),     # Kerry Express
        re.search(r'WHN\d{8,}', text_upper),       # Wahana
        re.search(r'RPX\d{8,}', text_upper),       # RPX
        re.search(r'SAP\d{10,}', text_upper),      # SAP Express
        re.search(r'LEX\d{8,}', text_upper),       # Lazada LEX
        re.search(r'LXAD[\-]?\d{8,}', text_upper), # Lazada LXAD
        re.search(r'JNAP[\-]?\d{8,}', text_upper), # Lazada JNAP
        re.search(r'BLI[A-Z0-9]{8,20}', text_upper), # Blibli
        re.search(r'TT\s*Order\s*ID', text, re.IGNORECASE),
        re.search(r'[A-Z]{2,5}\d{8,}', text_upper),
    ])
    
    # Alamat penerima yang LENGKAP
    has_recipient_address = any([
        ('penerima' in text_lower and re.search(r'(kecamatan|kelurahan|kab\.|kota|kabupaten)', text_lower)),
        ('pengirim' in text_lower and re.search(r'(kecamatan|kelurahan|kab\.|kota|kabupaten)', text_lower)),
        re.search(r'(penerima|pengirim).*(\+62|08\d{8,})', text_lower),
        re.search(r'\b\d{5}\b', text) and ('penerima' in text_lower or 'pengirim' in text_lower),
    ])
    
    # Marker label resi valid
    has_label_marker = any([
        'tanpa video unboxing' in text_lower,
        'harap rekam unboxing' in text_lower,
        'wajib video unboxing' in text_lower,
        'tidak terima komplain' in text_lower,
    ])
    
    # ===== LOGIKA PENGHAPUSAN =====
    if has_qty_total: return True
    if has_old_table and has_in_transit and not has_awb_barcode and not has_recipient_address: return True
    if has_old_table and has_marketplace_footer and not has_awb_barcode and not has_recipient_address and not has_label_marker: return True
    if has_old_table and not has_new_table and not has_awb_barcode and not has_recipient_address and not has_label_marker: return True
    if has_new_table and not has_old_table:
        if not (is_our_continuation or has_awb_barcode or has_recipient_address or has_label_marker): return True
    
    return False


# Maksimal karakter MSKU yang muat di 1 baris pada kolom extended (lebar 140pt, font 8)
MSKU_MAX_CHARS_EXTENDED = 22   # ≈ 140pt / ~6.3pt per char
MSKU_SAFETY_CHARS = 20          # margin agar tidak terlalu mepet border

def split_msku_at_dash(msku: str, max_chars: int = MSKU_SAFETY_CHARS):
    """
    Selalu pecah di DASH PERTAMA jika string lebih panjang dari kolom.
    """
    if len(msku) <= max_chars:
        return [msku]
    # Cari dash pertama
    first_dash = msku.find('-')
    if first_dash == -1:
        return [msku[:max_chars], msku[max_chars:]]
    line1 = msku[:first_dash + 1]
    line2 = msku[first_dash + 1:]
    return [line1, line2]

def format_msku_for_wrapping(msku: str, font_name: str, font_size: float, max_width: float) -> str:
    import re
    import html
    from reportlab.pdfbase.pdfmetrics import stringWidth
    
    tokens = re.split(r'([-/])', msku)
    lines = []
    current_line = ""
    
    for token in tokens:
        if not token: continue
        if token in ('-', '/'):
            current_line += token
        else:
            if current_line and stringWidth(current_line + token, font_name, font_size) > max_width:
                lines.append(current_line)
                current_line = token
            else:
                current_line += token
                
    if current_line:
        lines.append(current_line)
        
    return "<br/>".join(html.escape(line) for line in lines)


def get_pdf_font_name(family: str, is_bold: bool) -> str:
    """Helper untuk mendapatkan nama font ReportLab (standard/bold)."""
    if family == "Helvetica":
        return "Helvetica-Bold" if is_bold else "Helvetica"
    if family == "Courier":
        return "Courier-Bold" if is_bold else "Courier"
    if family == "Times-Roman":
        return "Times-Bold" if is_bold else "Times-Roman"
    if family == "Bahnschrift":
        return "Bahnschrift"
    return family

def shrink_cfg_for_page1(label_cfg, scale_f):
    """Skala seragam khusus Halaman 1 agar item pertama pasti muat tanpa menabrak teks."""
    if not label_cfg or scale_f >= 1.0:
        return label_cfg
        
    cfg = label_cfg.copy()
    
    # Scale fonts
    for k in ['ext_font_rak', 'ext_font_msku', 'ext_font_qty', 'ext_font_size', 'std_font_msku', 'std_font_qty', 'std_font_size']:
        val = float(cfg.get(k, 8.0))
        cfg[k] = val * scale_f
            
    # Scale row heights
    for k in ['ext_row_height', 'std_row_height']:
        val = float(cfg.get(k, 18.0))
        cfg[k] = val * scale_f
        
    return cfg

def scale_cfg_for_wide_label(label_cfg, W_pts, is_extended=True):
    """Skala otomatis ukuran tabel (font, kolom, tinggi) jika PDF sumber lebar."""
    if not W_pts or W_pts <= 350:
        return label_cfg
        
    cfg = label_cfg.copy() if label_cfg else {}
    if is_extended:
        base_w = float(cfg.get('ext_col_rak', 70)) + float(cfg.get('ext_col_msku', 150)) + float(cfg.get('ext_col_qty', 50))
    else:
        base_w = float(cfg.get('std_col_msku', 220)) + float(cfg.get('std_col_qty', 50))
        
    target_w = W_pts - 14 # margin 7 kanan kiri
    if target_w > base_w + 10:
        scale_f = target_w / base_w
        font_scale = min(scale_f, 1.6) 
        
        # Scale fonts
        for k in ['ext_font_rak', 'ext_font_msku', 'ext_font_qty', 'ext_font_size', 'std_font_msku', 'std_font_qty', 'std_font_size']:
            val = float(cfg.get(k, 8.0))
            cfg[k] = val * font_scale
                
        # Scale row heights
        for k in ['ext_row_height', 'std_row_height']:
            val = float(cfg.get(k, 18.0))
            cfg[k] = val * font_scale
                
        # Scale widths
        for k, default_w in [('ext_col_rak', 70.0), ('ext_col_msku', 150.0), ('ext_col_qty', 50.0), ('std_col_msku', 220.0), ('std_col_qty', 50.0)]:
            val = float(cfg.get(k, default_w))
            cfg[k] = val * scale_f
                
    return cfg

def create_table(data, row_heights=None, span_cmds=None, label_cfg=None):
    """Buat tabel ReportLab untuk MSKU dan Qty."""
    cfg = label_cfg or {}
    num_cols = len(data[0]) if data else 2
    if num_cols == 3:
        col_widths = [
            float(cfg.get('ext_col_rak',  70)),
            float(cfg.get('ext_col_msku', 150)),
            float(cfg.get('ext_col_qty',  50)),
        ]
        font_size = float(cfg.get('ext_font_size', 8))
    else:
        col_widths = [
            float(cfg.get('std_col_msku', 220)),
            float(cfg.get('std_col_qty',  50)),
        ]
        font_size = float(cfg.get('std_font_size', 8))

    border = float(cfg.get('border_thickness', 0.5))
    hdr_bg_hex   = str(cfg.get('header_bg',    '#000000')).lstrip('#')
    hdr_txt_hex  = str(cfg.get('header_color', '#ffffff')).lstrip('#')
    try:
        hdr_bg  = colors.HexColor(f'#{hdr_bg_hex}')
        hdr_txt = colors.HexColor(f'#{hdr_txt_hex}')
    except Exception:
        hdr_bg  = colors.black
        hdr_txt = colors.white

    if row_heights is None:
        row_heights = [max(18, font_size * 2.5)] * len(data)

    t = Table(data, colWidths=col_widths, rowHeights=row_heights)
    
    style = [
        ('GRID',        (0, 0), (-1, -1), border, colors.black),
        ('FONTSIZE',    (0, 0), (-1, -1), font_size),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME',    (0, 0), (-1,  0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1,  0), max(9.0, font_size * 1.1)), 
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',(0, 0), (-1, -1), 4),
        ('ALIGN',       (-1, 0), (-1, -1), 'CENTER'),
        ('ALIGN',       (0, 0),  (0,  -1), 'CENTER'),
        ('TEXTCOLOR',   (0, 0), (-1,  0), hdr_txt),
        ('TEXTCOLOR',   (0, 1), (-1, -1), colors.black),
    ]

    if hdr_bg_hex.lower() not in ['ffffff', 'fff', 'white']:
        style.append(('BACKGROUND', (0, 0), (-1, 0), hdr_bg))

    if num_cols == 3:
        f_rak = get_pdf_font_name(cfg.get('ext_font_rak_family', 'Helvetica'), cfg.get('ext_font_rak_bold', False))
        f_msku = get_pdf_font_name(cfg.get('ext_font_msku_family', 'Helvetica'), cfg.get('ext_font_msku_bold', False))
        f_qty = get_pdf_font_name(cfg.get('ext_font_qty_family', 'Helvetica'), cfg.get('ext_font_qty_bold', False))
        
        style.append(('FONTNAME', (0, 1), (0, -1), f_rak))
        style.append(('FONTNAME', (1, 1), (1, -1), f_msku))
        style.append(('FONTNAME', (2, 1), (2, -1), f_qty))
    else:
        f_msku = get_pdf_font_name(cfg.get('std_font_msku_family', 'Helvetica'), cfg.get('std_font_msku_bold', False))
        f_qty = get_pdf_font_name(cfg.get('std_font_qty_family', 'Helvetica'), cfg.get('std_font_qty_bold', False))
        
        style.append(('FONTNAME', (0, 1), (0, -1), f_msku))
        style.append(('FONTNAME', (1, 1), (1, -1), f_qty))
    
    if span_cmds:
        style.extend(span_cmds)

    t.setStyle(TableStyle(style))
    return t


def generate_table_data(chunk, is_extended, rak_map, label_cfg=None, picker_name=None):
    """
    Kembalikan dict dengan keys:
      - table_data   : list of row lists
      - row_heights  : list of float per row
      - span_cmds    : list of style commands
      - total_rows   : jumlah baris tabel (termasuk header)
    
    MSKU selalu menggunakan Paragraph agar word-wrap otomatis jika mentok border.
    """
    cfg = label_cfg or {}
    if not is_extended:
        std_f_msku = float(cfg.get('std_font_msku') or cfg.get('std_font_size', 8))
        std_f_qty = float(cfg.get('std_font_qty') or cfg.get('std_font_size', 8))
        max_font = max(std_f_msku, std_f_qty)
        row_height = max(float(cfg.get('std_row_height', 18)), max_font * 2.25)
        
        # Font & style untuk Paragraph word-wrap
        std_f_msku_fam = cfg.get('std_font_msku_family', 'Helvetica')
        std_f_msku_bold = cfg.get('std_font_msku_bold', False)
        std_f_msku_font = get_pdf_font_name(std_f_msku_fam, std_f_msku_bold)
        std_leading = max(14, round(std_f_msku * 1.75))
        col_msku_w = float(cfg.get('std_col_msku', 220))
        
        msku_para_style = ParagraphStyle(
            'msku_std_para',
            fontSize=std_f_msku,
            leading=std_leading,
            fontName=std_f_msku_font,
            leftIndent=0,
            rightIndent=0,
            spaceAfter=0,
            spaceBefore=0,
        )
        
        msku_header_text = 'MSKU'
        if picker_name:
            msku_header_text = f"MSKU   |   PIC : {str(picker_name).strip().upper()}"
            
        try:
            from reportlab.lib import colors
            hdr_txt_hex = str(cfg.get('header_color', '#ffffff')).lstrip('#')
            hdr_txt = colors.HexColor(f'#{hdr_txt_hex}')
        except Exception:
            from reportlab.lib import colors
            hdr_txt = colors.white
            
        hdr_style = ParagraphStyle(
            'hdr_para',
            fontSize=max(9.0, std_f_msku * 1.1),
            fontName='Helvetica-Bold',
            alignment=1, # center
            leading=max(14, round(std_f_msku * 1.75)),
            textColor=hdr_txt
        )
        msku_header_para = Paragraph(msku_header_text, hdr_style)
            
        table_data = [[msku_header_para, 'Qty']]
        
        # Hitung tinggi row header
        hdr_w, hdr_h = msku_header_para.wrap(col_msku_w - 8, 9999)
        row_heights = [max(row_height, hdr_h + 8)]
        
        for item in chunk:
            msku_text = item['msku']
            formatted_msku = format_msku_for_wrapping(msku_text, std_f_msku_font, std_f_msku, col_msku_w - 8)
            msku_para = Paragraph(formatted_msku, msku_para_style)
            # Hitung tinggi sebenarnya setelah wrap
            para_w, para_h = msku_para.wrap(col_msku_w - 8, 9999)  # 8 = leftPadding + rightPadding
            actual_rh = max(row_height, para_h + 4)  # +4 padding atas bawah
            table_data.append([msku_para, str(item['jumlah'])])
            row_heights.append(actual_rh)
        
        n = len(table_data)
        return {
            'table_data': table_data,
            'row_heights': row_heights,
            'span_cmds': [
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('FONTSIZE', (0, 1), (0, -1), std_f_msku),
                ('FONTSIZE', (1, 1), (1, -1), std_f_qty)
            ],
            'total_rows': n,
        }

    ext_f_rak  = float(cfg.get('ext_font_rak') or cfg.get('ext_font_size', 8))
    ext_f_msku = float(cfg.get('ext_font_msku') or cfg.get('ext_font_size', 8))
    ext_f_qty  = float(cfg.get('ext_font_qty') or cfg.get('ext_font_size', 8))
    
    ext_f_rak_fam  = cfg.get('ext_font_rak_family', 'Helvetica')
    ext_f_msku_fam = cfg.get('ext_font_msku_family', 'Helvetica')
    ext_f_qty_fam  = cfg.get('ext_font_qty_family', 'Helvetica')
    
    ext_f_msku_bold = cfg.get('ext_font_msku_bold', False)
    ext_f_msku_font = get_pdf_font_name(ext_f_msku_fam, ext_f_msku_bold)

    ext_max_f  = max(ext_f_rak, ext_f_msku, ext_f_qty)
    ext_row_h  = max(float(cfg.get('ext_row_height', 18)), ext_max_f * 2.25)
    ext_leading = max(14, round(ext_f_msku * 1.75))
    col_msku_w = float(cfg.get('ext_col_msku', 150))

    msku_para_style = ParagraphStyle(
        'msku_para',
        fontSize=ext_f_msku,
        leading=ext_leading,
        fontName=ext_f_msku_font,
        leftIndent=0,
        rightIndent=0,
        spaceAfter=0,
        spaceBefore=0,
    )

    msku_header_text = 'MSKU'
    if picker_name:
        msku_header_text = f"MSKU   |   PIC : {str(picker_name).strip().upper()}"
        
    try:
        from reportlab.lib import colors
        hdr_txt_hex = str(cfg.get('header_color', '#ffffff')).lstrip('#')
        hdr_txt = colors.HexColor(f'#{hdr_txt_hex}')
    except Exception:
        from reportlab.lib import colors
        hdr_txt = colors.white
        
    hdr_style_ext = ParagraphStyle(
        'hdr_para_ext',
        fontSize=max(9.0, ext_f_msku * 1.1),
        fontName='Helvetica-Bold',
        alignment=1,
        leading=max(14, round(ext_f_msku * 1.75)),
        textColor=hdr_txt
    )
    msku_header_para = Paragraph(msku_header_text, hdr_style_ext)
        
    table_data = [['Rak & ID', msku_header_para, 'Qty']]
    
    # Hitung tinggi row header
    hdr_w, hdr_h = msku_header_para.wrap(col_msku_w - 8, 9999)
    row_heights = [max(ext_row_h, hdr_h + 8)]
    span_cmds  = [
        ('FONTSIZE', (0, 1), (0, -1), ext_f_rak),
        ('FONTSIZE', (1, 1), (1, -1), ext_f_msku),
        ('FONTSIZE', (2, 1), (2, -1), ext_f_qty),
    ]
    
    current_row = 1
    for item in chunk:
        sku_upper  = item['msku'].strip().upper()
        rak_info   = rak_map.get(sku_upper, {"rak": "", "id": ""})
        rak_val    = rak_info.get('rak', '')
        id_val     = rak_info.get('id', '')
        combined = id_val if id_val else (rak_val or "-")

        # Selalu gunakan Paragraph agar word-wrap otomatis
        formatted_msku = format_msku_for_wrapping(item['msku'], ext_f_msku_font, ext_f_msku, col_msku_w - 8)
        msku_para = Paragraph(formatted_msku, msku_para_style)
        para_w, para_h = msku_para.wrap(col_msku_w - 8, 9999)  # 8 = leftPadding + rightPadding
        actual_rh = max(ext_row_h, para_h + 4)
        
        table_data.append([combined, msku_para, str(item['jumlah'])])
        row_heights.append(actual_rh)
        current_row += 1

    return {
        'table_data': table_data,
        'row_heights': row_heights,
        'span_cmds': span_cmds,
        'total_rows': current_row,
    }



def generate_recap_pdf(recap_dict, rak_map, label_cfg):
    """
    Generate a recap PDF page with aggregated items sorted by Rak & ID.
    """
    recap_items = []
    for msku, qty in recap_dict.items():
        sku_upper = msku.strip().upper()
        rak_info = rak_map.get(sku_upper, {"rak": "", "id": ""})
        rak_val = rak_info.get('rak', '')
        id_val = rak_info.get('id', '')
        combined = id_val if id_val else (rak_val or "-")
        recap_items.append({
            'rak_id': combined,
            'msku': msku,
            'jumlah': qty
        })
    
    def recap_sort_key(item):
        combined = item['rak_id']
        parts = combined.split('-') if combined else []
        zone  = parts[0] if parts else ''
        rest  = parts[1:] if len(parts) > 1 else []
        num_rest = []
        for p in rest:
            try: num_rest.append((0, int(p)))
            except: num_rest.append((1, p.upper()))
        try: zone_val = (0, int(zone))
        except: zone_val = (1, zone.upper())
        priority = 1 if combined == "-" else 0
        return (priority, zone_val, num_rest, item['msku'].upper())

    recap_items.sort(key=recap_sort_key)
    output_pdf = fitz.open()
    
    msku_para_style = ParagraphStyle(
        'msku_recap',
        fontSize=8,
        leading=10,
        fontName='Helvetica',
        leftIndent=0,
        rightIndent=0,
    )
    
    max_h = 330
    all_chunks = []
    current_chunk = []
    current_h = 18
    
    for item in recap_items:
        lines = split_msku_at_dash(item['msku'], max_chars=22)
        n = len(lines)
        row_h = max(18, n * 10 + 4)
        
        if current_h + row_h > max_h:
            all_chunks.append(current_chunk)
            current_chunk = [item]
            current_h = 18 + row_h
        else:
            current_chunk.append(item)
            current_h += row_h
    if current_chunk:
        all_chunks.append(current_chunk)

    for page_idx, chunk in enumerate(all_chunks):
        packet = io.BytesIO()
        W, H = 297, 420
        can = canvas.Canvas(packet, pagesize=(W, H))
        
        can.setFont("Helvetica-Bold", 10)
        can.drawCentredString(W/2, H - 25, "REKAP PENGAMBILAN BARANG (OVERLAP)")
        can.setFont("Helvetica", 7)
        can.drawCentredString(W/2, H - 35, f"Halaman {page_idx + 1} | Total SKUs: {len(recap_items)}")
        
        table_data = [['Rak & ID', 'MSKU', 'Qty']]
        row_heights = [18]
        
        for item in chunk:
            lines = split_msku_at_dash(item['msku'], max_chars=22)
            if len(lines) > 1:
                p = Paragraph("<br/>".join(lines), msku_para_style)
                table_data.append([item['rak_id'], p, str(item['jumlah'])])
                row_heights.append(max(18, len(lines) * 10 + 4))
            else:
                table_data.append([item['rak_id'], item['msku'], str(item['jumlah'])])
                row_heights.append(18)
        
        t = Table(table_data, colWidths=[75, 140, 50], rowHeights=row_heights)
        
        style = [
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ]
        t.setStyle(TableStyle(style))
        
        t.wrapOn(can, W, H)
        t.drawOn(can, 12, H - 50 - sum(row_heights))
        
        can.save()
        packet.seek(0)
        page_doc = fitz.open("pdf", packet.read())
        output_pdf.insert_pdf(page_doc)
        page_doc.close()
        
    return output_pdf


def generate_summary_page(title, final_ids):
    W_pts, H_pts = 283.46, 425.20 # 100x150mm
    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=(W_pts, H_pts))
    
    import qrcode
    from reportlab.lib.utils import ImageReader
    import math
    
    # 1. Top QR Code
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=0)
    qr.add_data(title)
    qr.make(fit=True)
    img_qr = qr.make_image(fill_color="black", back_color="white")
    qr_io = io.BytesIO()
    img_qr.save(qr_io, format='PNG')
    qr_io.seek(0)
    qr_reader = ImageReader(qr_io)
    
    qr_size = 90
    qr_x = (W_pts - qr_size) / 2
    qr_y = H_pts - qr_size - 15
    c.drawImage(qr_reader, qr_x, qr_y, width=qr_size, height=qr_size)
    
    # 2. Horizontal Line
    line_y = qr_y - 10
    c.setLineWidth(1)
    c.line(20, line_y, W_pts - 20, line_y)
    
    # 3. Title
    title_y = line_y - 15
    c.setFont("Helvetica-Bold", 12)
    title_font_size = 12
    while c.stringWidth(title, "Helvetica-Bold", title_font_size) > W_pts - 40 and title_font_size > 5:
        title_font_size -= 0.5
    c.setFont("Helvetica-Bold", title_font_size)
    c.drawCentredString(W_pts / 2, title_y, str(title))
    
    # 4. PACKING LIST
    pack_y = title_y - 15
    c.setFont("Helvetica", 9)
    c.drawCentredString(W_pts / 2, pack_y, "PACKING LIST / BATCH")
    
    # 5. Black Banner
    banner_y = pack_y - 20
    banner_height = 14
    banner_width = 120
    banner_x = (W_pts - banner_width) / 2
    c.setFillColorRGB(0, 0, 0)
    c.rect(banner_x, banner_y - banner_height/2, banner_width, banner_height, fill=1)
    
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(W_pts / 2, banner_y - 3, f"LIST ID PESANAN ({len(final_ids)})")
    
    # 6. Bottom Banner
    c.setFillColorRGB(0, 0, 0)
    bottom_banner_height = 25
    bottom_banner_y = 15
    box_x = 10
    box_width = W_pts - 20
    c.roundRect(box_x, bottom_banner_y, box_width, bottom_banner_height, 4, stroke=1, fill=0)
    
    c.setFont("Helvetica-Bold", 7)
    c.drawString(box_x + 30, bottom_banner_y + 13, "SCAN QR CODE INI UNTUK PENCATATAN")
    c.setFont("Helvetica", 6)
    c.drawString(box_x + 30, bottom_banner_y + 5, "Data akan otomatis masuk ke sistem.")
    
    # Smartphone Icon
    icon_x = box_x + 8
    icon_y = bottom_banner_y + 3
    c.roundRect(icon_x, icon_y, 14, 19, 2, stroke=1, fill=0)
    # mini screen
    screen_x = icon_x + 2
    screen_y = icon_y + 4
    screen_w = 10
    screen_h = 13
    c.rect(screen_x, screen_y, screen_w, screen_h, stroke=1, fill=0)
    c.circle(icon_x + 7, icon_y + 2, 1, stroke=1, fill=0)
    # Mini QR code inside screen
    c.drawImage(qr_reader, screen_x + 1, screen_y + 1, width=screen_w - 2, height=screen_w - 2)
    
    # 7. Main Box
    box_y_top = banner_y - banner_height/2 - 10
    box_y_bottom = bottom_banner_y + bottom_banner_height + 5
    box_height = box_y_top - box_y_bottom
    c.roundRect(box_x, box_y_bottom, box_width, box_height, 4, stroke=1, fill=0)
    
    # 8. Draw Multi-column list
    num_items = len(final_ids)
    if num_items == 0:
        c.save()
        packet.seek(0)
        return fitz.open(stream=packet.read(), filetype="pdf")
        
    num_cols = 5 if num_items >= 5 else num_items
    if num_cols == 0: num_cols = 1
    rows_per_col = math.ceil(num_items / num_cols)
    if rows_per_col == 0: rows_per_col = 1
    
    col_width = box_width / num_cols
    
    # draw column vertical dashed lines
    c.setDash(1, 2)
    for col in range(1, num_cols):
        line_x = box_x + col * col_width
        c.line(line_x, box_y_bottom + 2, line_x, box_y_top - 2)
    c.setDash()
    
    # Calculate row height
    margin_y = 5
    avail_height = box_height - (margin_y * 2)
    row_height = avail_height / rows_per_col
    row_height = min(12, row_height) # cap height
    
    # Base font size
    base_font_size = min(6.5, row_height * 0.8)
    
    start_y = box_y_top - margin_y - (row_height * 0.8)
    
    for i, pid in enumerate(final_ids):
        col_idx = i // rows_per_col
        row_idx = i % rows_per_col
        
        x = box_x + (col_idx * col_width) + 2
        y = start_y - (row_idx * row_height)
        
        # Adjust font size if text is too long for the column
        idx_str = f"{i+1:02d}. "
        pid_str = str(pid)
        
        c.setFont("Helvetica-Bold", base_font_size)
        idx_width = c.stringWidth(idx_str, "Helvetica-Bold", base_font_size)
        
        pid_font_size = base_font_size
        while c.stringWidth(pid_str, "Helvetica", pid_font_size) > (col_width - idx_width - 3) and pid_font_size > 3:
            pid_font_size -= 0.5
            
        c.setFont("Helvetica-Bold", base_font_size)
        c.drawString(x, y, idx_str)
        
        c.setFont("Helvetica", pid_font_size)
        c.drawString(x + idx_width, y, pid_str)
        
    c.save()
    packet.seek(0)
    return fitz.open(stream=packet.read(), filetype="pdf")


def generate_global_msku_summary_pdf(awb_to_items, matched_awbs, rak_map, label_cfg, barang_khusus_skus=None):
    """
    Generate a GLOBAL MSKU summary PDF that aggregates ONLY requested MSKU items
    (barang khusus) across ALL matched AWBs. Shows total quantities.
    """
    if barang_khusus_skus is None or not barang_khusus_skus:
        return fitz.open()  # Return empty doc if no data requested
    
    # Aggregate requested MSKU across all matched AWBs
    global_msku = {}  # msku -> total_qty
    for awb in matched_awbs:
        items = awb_to_items.get(awb, [])
        for item in items:
            msku = item.get('msku', '').strip()
            qty = item.get('jumlah', 1)
            if msku and msku.upper() in barang_khusus_skus:
                if msku in global_msku:
                    global_msku[msku] += qty
                else:
                    global_msku[msku] = qty
    
    if not global_msku:
        return fitz.open()  # Return empty doc
    
    # Build items list
    summary_items = []
    total_qty = 0
    
    for msku, qty in global_msku.items():
        summary_items.append({
            'msku': msku,
            'jumlah': qty
        })
        total_qty += qty
    
    # Sort alphabetically by MSKU
    summary_items.sort(key=lambda x: x['msku'].upper())
    
    output_pdf = fitz.open()
    
    msku_para_style = ParagraphStyle(
        'msku_global',
        fontSize=9.5,
        leading=12,
        fontName='Helvetica',
        leftIndent=0,
        rightIndent=0,
    )
    
    # Chunk into pages
    max_h = 320
    all_chunks = []
    current_chunk = []
    current_h = 18  # footer row
    
    for item in summary_items:
        lines = split_msku_at_dash(item['msku'], max_chars=30)
        n = len(lines)
        row_h = max(20, n * 12 + 6)
        
        if current_h + row_h > max_h:
            all_chunks.append(current_chunk)
            current_chunk = [item]
            current_h = 18 + row_h
        else:
            current_chunk.append(item)
            current_h += row_h
    if current_chunk:
        all_chunks.append(current_chunk)
    
    total_pages = len(all_chunks)
    
    for page_idx, chunk in enumerate(all_chunks):
        packet = io.BytesIO()
        W, H = 297, 420
        can = canvas.Canvas(packet, pagesize=(W, H))
        
        # Title
        can.setFont("Helvetica-Bold", 11)
        can.drawCentredString(W/2, H - 24, "TOTAL REKAP BARANG KESELURUHAN")
        can.setFont("Helvetica", 8)
        can.drawCentredString(W/2, H - 35, f"Halaman {page_idx + 1} dari {total_pages}")
        
        # Line separator
        can.setLineWidth(1)
        can.line(20, H - 40, W - 20, H - 40)
        
        # Table
        table_data = []
        row_heights = []
        
        for item in chunk:
            lines = split_msku_at_dash(item['msku'], max_chars=30)
            
            if len(lines) > 1:
                p = Paragraph("<br/>".join(lines), msku_para_style)
                table_data.append([p, str(item['jumlah'])])
                row_heights.append(max(20, len(lines) * 12 + 6))
            else:
                table_data.append([item['msku'], str(item['jumlah'])])
                row_heights.append(20)
                
        # Footer row on the last page
        if page_idx == total_pages - 1:
            table_data.append(['TOTAL KESELURUHAN SEMUA BARANG', str(total_qty)])
            row_heights.append(22)
        
        t = Table(table_data, colWidths=[220, 50], rowHeights=row_heights)
        
        style_cmds = [
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (0, -1), 9.5),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            
            # Qty Column
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 0), (1, -1), 13),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ]
        
        if page_idx == total_pages - 1:
            style_cmds.append(('FONTNAME', (0, -1), (0, -1), 'Helvetica-Bold'))
            style_cmds.append(('ALIGN', (0, -1), (0, -1), 'RIGHT'))
            style_cmds.append(('FONTSIZE', (0, -1), (0, -1), 10.5))
            
        t.setStyle(TableStyle(style_cmds))
        
        t.wrapOn(can, W, H)
        # Center the table: (297 - (220+50))/2 = 270/2 = 13.5
        t.drawOn(can, 13.5, H - 45 - sum(row_heights))
        
        # Bottom indicator
        if page_idx == total_pages - 1:
            can.setFont("Helvetica-Bold", 8)
            can.drawCentredString(W/2, 20, "--- AKHIR DARI REKAP BARANG ---")
        
        can.save()
        packet.seek(0)
        page_doc = fitz.open("pdf", packet.read())
        output_pdf.insert_pdf(page_doc)
        page_doc.close()
    
    return output_pdf


def calc_items_for_rows(items, is_extended, rak_map, max_rendered_rows: int):
    """
    Hitung item yang muat dengan slot baris (tidak termasuk header).
    """
    if not is_extended:
        fit = min(len(items), int(max_rendered_rows))
        return items[:fit], items[fit:]

    slots_used = 0
    for idx, item in enumerate(items):
        lines = split_msku_at_dash(item['msku'])
        needed = len(lines)
        if slots_used + needed > max_rendered_rows:
            if idx == 0: return [], items
            return items[:idx], items[idx:]
        slots_used += needed
    return items, []


def calc_items_for_rows_by_height(items, is_extended, rak_map, available_height: float, label_cfg, W_pts: float, force_first: bool = False):
    """
    Hitung item yang muat dalam `available_height` secara presisi menggunakan kalkulasi wrap teks Paragraph.
    """
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import ParagraphStyle
    
    scaled_cfg = scale_cfg_for_wide_label(label_cfg, W_pts, is_extended)
    if is_extended:
        f_msku = float(scaled_cfg.get('ext_font_msku') or scaled_cfg.get('ext_font_size', 8))
        f_rak = float(scaled_cfg.get('ext_font_rak') or scaled_cfg.get('ext_font_size', 8))
        f_qty = float(scaled_cfg.get('ext_font_qty') or scaled_cfg.get('ext_font_size', 8))
        ext_max_f = max(f_rak, f_msku, f_qty)
        row_h = max(float(scaled_cfg.get('ext_row_height', 18)), ext_max_f * 2.25)
        col_msku_w = float(scaled_cfg.get('ext_col_msku', 150))
        fam = scaled_cfg.get('ext_font_msku_family', 'Helvetica')
        is_bold = scaled_cfg.get('ext_font_msku_bold', False)
    else:
        f_msku = float(scaled_cfg.get('std_font_msku') or scaled_cfg.get('std_font_size', 8))
        f_qty = float(scaled_cfg.get('std_font_qty') or scaled_cfg.get('std_font_size', 8))
        ext_max_f = max(f_msku, f_qty)
        row_h = max(float(scaled_cfg.get('std_row_height', 18)), ext_max_f * 2.25)
        col_msku_w = float(scaled_cfg.get('std_col_msku', 220))
        fam = scaled_cfg.get('std_font_msku_family', 'Helvetica')
        is_bold = scaled_cfg.get('std_font_msku_bold', False)

    font_name = get_pdf_font_name(fam, is_bold)
    leading = max(14, round(f_msku * 1.75))
    
    style = ParagraphStyle(
        'msku_para_tmp',
        fontSize=f_msku, leading=leading, fontName=font_name,
        leftIndent=0, rightIndent=0, spaceAfter=0, spaceBefore=0
    )

    used_height = row_h # Header
    for idx, item in enumerate(items):
        formatted_msku = format_msku_for_wrapping(item['msku'], font_name, f_msku, col_msku_w - 8)
        msku_para = Paragraph(formatted_msku, style)
        para_w, para_h = msku_para.wrap(col_msku_w - 8, 9999)
        item_height = max(row_h, para_h + 4)
            
        if used_height + item_height > available_height:
            if idx == 0: 
                if force_first:
                    return [items[0]], items[1:]
                return [], items
            return items[:idx], items[idx:]
        used_height += item_height
    return items, []




def is_picking_list(content_check: bytes, filename: str) -> bool:
    try:
        import fitz
        doc = fitz.open(stream=content_check, filetype="pdf")
        # Cek halaman pertama saja untuk mempercepat
        if len(doc) > 0:
            text = doc[0].get_text().upper()
            import re
            # Cek keywords yang biasanya ada di Picking List, gunakan regex untuk handle enter/newline
            if re.search(r'PICKING\s+LIST', text) or \
               re.search(r'DAFTAR\s+PENGAMBILAN', text) or \
               re.search(r'DAFTAR\s+PESANAN', text) or \
               re.search(r'NO[:\s]*PL\d+', text):
                return True
        return False
    except Exception:
        return False

@app.post("/extract-matched-order-ids")
async def extract_matched_order_ids(
    excel_file: UploadFile = File(...),
    pdf_files: list[UploadFile] = File(...)
):
    try:
        excel_content = await excel_file.read()
        df_temp = pd.read_excel(io.BytesIO(excel_content), nrows=0)
        id_col = next((c for c in df_temp.columns if 'ID Pesanan' in c or 'Nomor Pesanan' in c or 'Order ID' in c), None)
        
        dtype_dict = {id_col: str} if id_col else {}
        df = pd.read_excel(io.BytesIO(excel_content), dtype=dtype_dict)
        
        col_mapping = {
            'AWB/No. Tracking': ['AWB/No. Tracking', 'AWB', 'No. Tracking', 'Tracking Number', 'Resi', 'No Resi'],
            'ID Pesanan': ['ID Pesanan', 'Order ID', 'No. Pesanan', 'Nomor Pesanan']
        }
        
        for target_col, alternatives in col_mapping.items():
            if target_col not in df.columns:
                for alt in alternatives:
                    if alt in df.columns:
                        df = df.rename(columns={alt: target_col})
                        break
                        
        has_id_pesanan = 'ID Pesanan' in df.columns
        has_awb = 'AWB/No. Tracking' in df.columns
        
        if has_id_pesanan: df['ID Pesanan'] = df['ID Pesanan'].replace('', float('nan')).ffill()
        if has_awb:
            df['AWB/No. Tracking'] = df['AWB/No. Tracking'].replace('', float('nan'))
            df['AWB/No. Tracking'] = df.groupby('ID Pesanan', sort=False)['AWB/No. Tracking'].transform(lambda x: x.ffill().bfill()) if has_id_pesanan else df['AWB/No. Tracking'].ffill()

        excel_awbs = set()
        id_to_awb_mapping = {}
        
        for _, row in df.iterrows():
            if has_id_pesanan:
                id_pesanan_val = row['ID Pesanan']
                if pd.notna(id_pesanan_val):
                    id_pesanan_norm = str(id_pesanan_val).strip()
                    if id_pesanan_norm:
                        excel_awbs.add(id_pesanan_norm)
                        if has_awb:
                            awb_val = row['AWB/No. Tracking']
                            if pd.notna(awb_val) and str(awb_val).strip():
                                awb_norm = str(awb_val).strip()
                                excel_awbs.add(awb_norm)
                                id_to_awb_mapping[id_pesanan_norm] = awb_norm
                                
        awb_to_id_mapping = {v: k for k, v in id_to_awb_mapping.items()}
        
        matched_ids = set()
        
        for pdf in pdf_files:
            doc = fitz.open(stream=await pdf.read(), filetype="pdf")
            for page_num in range(len(doc)):
                page = doc[page_num]
                text = page.get_text()
                candidates = extract_all_awb_candidates(text) + extract_order_ids(text)
                matched_awb = find_matching_awb(candidates, excel_awbs, page_num + 1)
                
                if matched_awb:
                    if matched_awb in id_to_awb_mapping:
                        matched_ids.add(matched_awb) # it was an ID Pesanan
                    elif matched_awb in awb_to_id_mapping:
                        matched_ids.add(awb_to_id_mapping[matched_awb])
                        
        return {"ids": list(matched_ids)}
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Error in extract-matched-order-ids: {e}")
        return {"ids": []}

@app.post("/process-labels")
async def process_labels(
    excel_file: UploadFile = File(...),
    pdf_files: list[UploadFile] = File(...),
    priority_kembar: bool = Form(False),
    sort_by_sku_count: bool = Form(False),
    include_global_msku: bool = Form(False),
    picker_name: Optional[str] = Form(None)
):
    """
    Process labels and return match statistics along with the PDF.
    Returns JSON with:
    - pdf_base64: Base64 encoded PDF
    - stats: Match statistics (matched, unmatched_excel, unmatched_pdf)
    """
    # Debug log to file
    with open("debug.log", "w") as f:
        f.write(f"=== PROCESS LABELS WITH STATS START ===\n")
        f.write(f"PDF files count: {len(pdf_files)}\n")
        for i, pdf in enumerate(pdf_files):
            f.write(f"  PDF {i+1}: {pdf.filename}\n")
        f.write(f"Excel file: {excel_file.filename}\n")
    
    try:
        # --- Fetch configurations (Custom Label Priority Top) ---
        is_priority_top_active = True
        special_skus = set()
        
        if sort_by_sku_count:
            try:
                features = await supabase_fetch("GET", "toolkit_feature_locks?select=feature_key,is_locked")
                if isinstance(features, list):
                    for f in features:
                        if f.get("feature_key") == "custom-label-priority-top" and f.get("is_locked", False):
                            is_priority_top_active = False
                            break
                            
                if is_priority_top_active:
                    priority_data = await supabase_fetch("GET", "sku_priority_bottom?select=sku")
                    if isinstance(priority_data, list):
                        special_skus = {str(r.get('sku')).strip().upper() for r in priority_data}
            except Exception as e:
                print(f"Error fetching custom label priority settings: {e}")
        # --------------------------------------------------------

        # PL Content Check (Early Validation)
        for pdf in pdf_files:
            content_check = await pdf.read()
            if is_picking_list(content_check, pdf.filename):
                 raise HTTPException(status_code=400, detail={"code": "PL_DETECTED", "message": f"File '{pdf.filename}' terdeteksi sebagai Picking List (PL). Mohon upload Resi Asli."})
            await pdf.seek(0)
            
        # 1. Baca File Excel Ginee
        excel_content = await excel_file.read()
        excel_filename = excel_file.filename
        df = pd.read_excel(io.BytesIO(excel_content), dtype=object)
        df = fix_excel_numeric_ids(df)
        
        # Normalize kolom
        col_mapping = {
            'AWB/No. Tracking': ['AWB/No. Tracking', 'AWB', 'No. Tracking', 'Tracking Number', 'Resi', 'No Resi'],
            'ID Pesanan': ['ID Pesanan', 'Order ID', 'No. Pesanan', 'Nomor Pesanan'],
            'MSKU': ['MSKU', 'SKU', 'Nama SKU', 'Product SKU', 'Master SKU'],
            'Jumlah': ['Jumlah', 'Qty', 'Quantity', 'QTY'],
            'Catatan Pembeli': ['Catatan Pembeli', 'Buyer Note', 'Buyer Notes', 'Customer Note', 'Note']
        }
        
        for target_col, alternatives in col_mapping.items():
            if target_col not in df.columns:
                for alt in alternatives:
                    if alt in df.columns:
                        df = df.rename(columns={alt: target_col})
                        break
        
        # Required: MSKU dan Jumlah
        required_cols = ['MSKU', 'Jumlah']
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            raise HTTPException(status_code=400, detail=f"Kolom tidak ditemukan: {missing}")
        
        # Pastikan minimal ada salah satu: ID Pesanan atau AWB/No. Tracking
        has_id_pesanan = 'ID Pesanan' in df.columns
        has_awb = 'AWB/No. Tracking' in df.columns
        has_catatan = 'Catatan Pembeli' in df.columns
        if not has_id_pesanan and not has_awb:
            raise HTTPException(status_code=400, detail="Kolom 'ID Pesanan' atau 'AWB/No. Tracking' tidak ditemukan")

        if has_id_pesanan:
            df['ID Pesanan'] = df['ID Pesanan'].replace('', float('nan')).ffill()
        if has_awb:
            # FIX: Only ffill AWB within rows that share the same ID Pesanan.
            # This prevents AWB from one order "leaking" into the next order whose AWB is empty.
            df['AWB/No. Tracking'] = df['AWB/No. Tracking'].replace('', float('nan'))
            if has_id_pesanan:
                df['AWB/No. Tracking'] = df.groupby('ID Pesanan', sort=False)['AWB/No. Tracking'].transform(lambda x: x.ffill())
            else:
                df['AWB/No. Tracking'] = df['AWB/No. Tracking'].ffill()

        # 2. Buat mapping - PRIORITAS: ID Pesanan dulu, jika kosong maka AWB
        awb_to_items = {}
        excel_awbs = set()  # Semua keys yang bisa digunakan untuk matching (ID + AWB)
        excel_primary_ids = set()  # HANYA identifier utama per order (untuk statistik)
        excel_awbs_raw_set = set() # Untuk menyimpan semua AWB raw yang ada di excel
        id_to_awb_mapping = {}  # Mapping ID Pesanan -> AWB/No. Tracking
        awb_to_catatan = {}  # Mapping ID/AWB -> Catatan Pembeli (untuk JX)
        
        for _, row in df.iterrows():
            identifier = None
            awb_value = None
            
            # Ambil AWB/No. Tracking value (untuk disimpan ke Supabase)
            if has_awb:
                awb_raw = row.get('AWB/No. Tracking', '')
                awb_norm = normalize_awb(awb_raw)
                if awb_norm and awb_norm not in ['NAN', 'NONE', 'NULL', '']:
                    awb_value = awb_norm
                    excel_awbs_raw_set.add(awb_value)
            
            # Ambil Catatan Pembeli (untuk JX labels)
            catatan_pembeli = ''
            if has_catatan:
                catatan_raw = row.get('Catatan Pembeli', '')
                if pd.notna(catatan_raw) and str(catatan_raw).strip() not in ['nan', 'NaN', 'NAN', '']:
                    catatan_pembeli = str(catatan_raw).strip()
            
            # Cek apakah ini Blibli
            is_blibli = awb_value and awb_value.startswith('BLI')
            
            # Prioritas 1: ID Pesanan (untuk validasi/matching) - KECUALI BLIBLI
            if has_id_pesanan and not is_blibli:
                id_pesanan = row.get('ID Pesanan', '')
                id_pesanan_norm = normalize_awb(id_pesanan)
                if id_pesanan_norm and id_pesanan_norm not in ['NAN', 'NONE', 'NULL', '']:
                    identifier = id_pesanan_norm
                    # Simpan mapping ID Pesanan -> AWB
                    if awb_value:
                        id_to_awb_mapping[identifier] = awb_value
                    # Simpan Catatan Pembeli untuk semua label
                    if catatan_pembeli:
                        awb_to_catatan[identifier] = catatan_pembeli
            
            # Prioritas 2: AWB/No. Tracking (jika ID Pesanan kosong)
            if not identifier and awb_value:
                identifier = awb_value
                # Simpan Catatan Pembeli untuk semua label
                if catatan_pembeli:
                    awb_to_catatan[identifier] = catatan_pembeli
            
            # === PENTING: Tambahkan KEDUA identifier ke excel_awbs ===
            # Ini memungkinkan matching dengan ID Pesanan ATAU AWB
            items_data = {
                'msku': str(row['MSKU']),
                'jumlah': int(row['Jumlah']) if pd.notna(row['Jumlah']) else 1
            }
            
            if identifier:
                excel_awbs.add(identifier)
                excel_primary_ids.add(identifier)  # Tambahkan ke primary IDs untuk statistik
                if identifier not in awb_to_items:
                    awb_to_items[identifier] = []
                awb_to_items[identifier].append(items_data.copy())
            
            # JUGA tambahkan AWB secara terpisah jika ada dan berbeda dari identifier
            # Ini memungkinkan matching via AWB meski identifier utama adalah ID Pesanan
            # CATATAN: AWB ekstra ini TIDAK dimasukkan ke excel_primary_ids (hanya untuk matching)
            if awb_value and awb_value != identifier:
                excel_awbs.add(awb_value)
                if awb_value not in awb_to_items:
                    awb_to_items[awb_value] = []
                awb_to_items[awb_value].append(items_data.copy())
                if catatan_pembeli:
                    awb_to_catatan[awb_value] = catatan_pembeli
        
        # 2c. Agregasi items per AWB: gabungkan baris dengan MSKU sama, jumlahkan qty
        #     Ini menangani format export Lazada/Ginee yang menghasilkan 1 baris per unit
        #     (misal: SHARPENER qty 24 → 24 baris qty=1 → digabung jadi 1 baris qty=24)
        print(f"\n[AGGREGATE] ===== STARTING MSKU AGGREGATION =====")
        print(f"[AGGREGATE] Total AWB keys to process: {len(awb_to_items)}")
        for _awb_key in list(awb_to_items.keys()):
            before_count = len(awb_to_items[_awb_key])
            merged = {}
            for _item in awb_to_items[_awb_key]:
                _key = _item['msku'].strip()
                if _key in merged:
                    merged[_key]['jumlah'] += _item['jumlah']
                else:
                    merged[_key] = {'msku': _item['msku'].strip(), 'jumlah': _item['jumlah']}
            after_count = len(merged)
            if before_count != after_count:
                print(f"[AGGREGATE] AWB {_awb_key}: {before_count} rows -> {after_count} unique MSKU (MERGED)")
                for mk, mv in merged.items():
                    print(f"  -> {mk}: qty={mv['jumlah']}")
            awb_to_items[_awb_key] = list(merged.values())
        print(f"[AGGREGATE] ===== AGGREGATION COMPLETE =====\n")
        
        # 2b. Fetch format mode SEBELUM sort agar bisa sort by Rak & ID
        try:
            feature_res = await supabase_fetch("GET", "toolkit_feature_locks")
            is_extended = any(f.get('feature_key') == 'label_extended_format' and f.get('is_locked') for f in feature_res)
            is_sort_rak_msku = any(f.get('feature_key') == 'label_sort_rak_msku' and f.get('is_locked') for f in feature_res)
        except:
            is_extended = False
            is_sort_rak_msku = False

        rak_map = {}
        if is_extended:
            try:
                mappings = await get_sku_mappings()
                rak_map = {m['sku'].strip().upper(): {"rak": m.get('rak', ''), "id": m.get('id', '')} for m in mappings}
            except:
                pass

        # Fetch label table config (ukuran kolom, font, border)
        try:
            label_cfg = await get_label_config()
        except:
            label_cfg = {}

        try:
            bottom_priorities_res = await supabase_fetch("GET", "label_bottom_priorities")
            bottom_priorities_std = [item['keyword'].strip().upper() for item in bottom_priorities_res if item['format_type'] == 'standar']
            bottom_priorities_rak = [item['keyword'].strip().upper() for item in bottom_priorities_res if item['format_type'] == 'rak_id']
        except:
            bottom_priorities_std = []
            bottom_priorities_rak = []

        def rak_id_sort_key(item):
            sku_upper = item['msku'].strip().upper()
            rak_info  = rak_map.get(sku_upper, {"rak": "", "id": ""})
            rak_val   = rak_info.get('rak', '')
            id_val    = rak_info.get('id', '')  # custom_id, sudah berisi full path: "12-EK-06-22"

            # Gunakan id_val (custom_id) langsung sebagai combined karena sudah berisi
            # lorong+rak+id secara lengkap, misal "12-EK-06-22" atau "2-O-01-03".
            # Jangan gabungkan dengan rak_val (akan duplikat: "EK-12-EK-06-22").
            combined = id_val if id_val else rak_val

            parts = combined.split('-') if combined else []
            zone  = parts[0] if parts else ''
            rest  = parts[1:] if len(parts) > 1 else []

            num_rest = []
            for p in rest:
                try:
                    num_rest.append((0, int(p)))
                except ValueError:
                    num_rest.append((1, p.upper()))

            # Lorong numerik di-sort numerik (2 < 13 < 14), huruf di-sort alfabet (B, Z)
            try:
                zone_val = (0, int(zone))
            except ValueError:
                zone_val = (1, zone.upper())

            priority = 1 if not combined else 0
            return (priority, zone_val, num_rest, sku_upper)

        for awb in awb_to_items:
            if is_extended:
                if is_sort_rak_msku:
                    rak_items = []
                    no_rak_items = []
                    for item in awb_to_items[awb]:
                        sku_upper = item['msku'].strip().upper()
                        rak_info = rak_map.get(sku_upper, {"rak": "", "id": ""})
                        rak_val = rak_info.get('rak', '')
                        id_val = rak_info.get('id', '')
                        combined = id_val if id_val else (rak_val or "")
                        
                        if combined:
                            rak_items.append(item)
                        else:
                            no_rak_items.append(item)
                    
                    rak_items.sort(key=rak_id_sort_key)
                    no_rak_items.sort(key=lambda x: x['msku'].strip().upper())
                    
                    result = []
                    rak_idx = 0
                    for no_rak in no_rak_items:
                        no_rak_msku = no_rak['msku'].strip().upper()
                        while rak_idx < len(rak_items):
                            if rak_items[rak_idx]['msku'].strip().upper() > no_rak_msku:
                                break
                            result.append(rak_items[rak_idx])
                            rak_idx += 1
                        result.append(no_rak)
                    
                    while rak_idx < len(rak_items):
                        result.append(rak_items[rak_idx])
                        rak_idx += 1
                        
                    awb_to_items[awb] = result
                else:
                    awb_to_items[awb].sort(key=rak_id_sort_key)
            else:
                awb_to_items[awb].sort(key=lambda x: x['msku'])

            # Extract bottom priority items
            if is_extended:
                normal_items = []
                bottom_items = []
                for item in awb_to_items[awb]:
                    sku_up = item['msku'].strip().upper()
                    r_info = rak_map.get(sku_up, {})
                    # 'id' = full custom_id (e.g. Z-IE-04-12), 'rak' = area prefix
                    c = str(r_info.get('id') or '').strip().upper()
                    c_rak = str(r_info.get('rak') or '').strip().upper()
                    is_bottom = any(
                        c == k or c.startswith(k) or c_rak == k or c_rak.startswith(k)
                        for k in bottom_priorities_rak
                    )
                    if is_bottom:
                        bottom_items.append(item)
                    else:
                        normal_items.append(item)
                # Sort bottom items A-Z by their location string (B before Z)
                def get_loc2(item):
                    r = rak_map.get(item['msku'].strip().upper(), {})
                    return str(r.get('id') or r.get('rak') or item['msku']).strip().upper()
                bottom_items.sort(key=get_loc2)
                awb_to_items[awb] = normal_items + bottom_items
            else:
                normal_items = []
                bottom_items = []
                for item in awb_to_items[awb]:
                    sku_up = item['msku'].strip().upper()
                    is_bottom = any(sku_up == k or sku_up.startswith(k) for k in bottom_priorities_std)
                    if is_bottom:
                        bottom_items.append(item)
                    else:
                        normal_items.append(item)
                bottom_items.sort(key=lambda x: x['msku'].strip().upper())
                awb_to_items[awb] = normal_items + bottom_items
        
        # Create reverse mapping for duplicate checking
        # awb_to_id_mapping: AWB -> ID Pesanan
        awb_to_id_mapping = {v: k for k, v in id_to_awb_mapping.items()}
        
        print(f"[DEBUG] Catatan Pembeli untuk JX: {len(awb_to_catatan)} entries")
        print(f"[DEBUG] Total AWB dari Excel: {len(excel_awbs)}")
        if excel_awbs:
            sample_awbs = list(excel_awbs)[:10]
            print(f"[DEBUG] Sample AWB dari Excel: {sample_awbs}")
            

        # 3. Baca PDF dan track AWBs
        all_pages = []
        pdf_awbs = set()  # AWB utama dari setiap page PDF
        pdf_filenames = []
        combined_pdf_content = io.BytesIO()  # Untuk backup: gabungan semua PDF asli
        combined_pdf_doc = fitz.open()  # Document untuk menggabungkan semua PDF asli
        
        for pdf in pdf_files:
            pdf_content = await pdf.read()
            pdf_filenames.append(pdf.filename)
            doc = fitz.open(stream=pdf_content, filetype="pdf")
            
            # Tambahkan ke combined PDF untuk backup
            combined_pdf_doc.insert_pdf(doc)
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                primary_identifier = None  # FIX: Reset setiap halaman baru agar tidak bocor dari halaman sebelumnya
                
                # FIX: Scan 70% bagian atas saja
                clip_rect = fitz.Rect(0, 0, page.rect.width, page.rect.height * 0.70)
                text = page.get_text("text", clip=clip_rect)
                full_text = page.get_text("text")  # Full text untuk junk/continuation check
                # Ekstrak kandidat dari FULL TEXT (bukan clip 70%) agar
                # barcode bawah Lazada (Order ID / Nomor Order) ikut ter-ekstrak.
                # Barcode Order ID Lazada ada di bawah separator --- di luar clip 70%.
                candidates = extract_all_awb_candidates(full_text)
                order_ids = extract_order_ids(full_text)  # Juga cari Order ID
                
                # Cek apakah halaman ini adalah junk page atau lanjutan kita
                page_is_junk = is_junk_page(page)
                page_is_continuation = 'LANJUTAN AWB:' in full_text.upper()
                
                # Gabungkan kandidat dengan PRIORITAS: alphanumerik (huruf+angka) lebih dulu, numerik murni terakhir
                all_raw = candidates + order_ids
                alpha_num_cands = list(dict.fromkeys([c for c in all_raw if any(ch.isalpha() for ch in c)]))
                pure_num_cands  = list(dict.fromkeys([c for c in all_raw if not any(ch.isalpha() for ch in c)]))
                priority_candidates = alpha_num_cands + pure_num_cands
                
                # Cek jika halaman ini adalah PACKING LIST / BATCH
                is_packing_list = 'PACKING LIST / BATCH' in full_text.upper() or 'LIST ID PESANAN' in full_text.upper()
                
                # Coba match dengan excel dulu (kecuali jika ini packing list)
                if is_packing_list:
                    matched_awb = None
                else:
                    matched_awb = find_matching_awb(priority_candidates, excel_awbs, page_num + 1)
                
                # DEBUG: Tulis info ekstraksi ke debug.log
                with open("debug_text.log", "a", encoding="utf-8") as dbg:
                    dbg.write(f"\n--- Page {page_num + 1} ---\n")
                    dbg.write(f"Is Junk: {page_is_junk}, Is Continuation: {page_is_continuation}\n")
                    dbg.write(f"AWB Candidates: {candidates[:10]}\n")
                    dbg.write(f"Order IDs extracted: {order_ids[:10]}\n")
                    dbg.write(f"Matched AWB: {matched_awb}\n")
                    dbg.write(f"Raw text (first 300 chars):\n{text[:300]}\n")
                
                if matched_awb:
                    primary_identifier = matched_awb
                else:
                    # Jika tidak match, ambil AWB/ID terbaik dari page
                    for c in candidates:
                        c_upper = c.upper()
                        known_prefixes = [
                            'SPXID', 'SPX',                              # Shopee
                            'JT', 'JX', 'JD',                            # J&T
                            'NVID', 'NV', 'NJVTT', 'NJV', 'NINJA',        # Ninja
                            'TKP', 'ANT', 'GTL',                          # Tokopedia/AnterAja/GoTo
                            'SAP', 'LEX', 'LAZ', 'ID',                    # SAP/Lazada
                            'JNE', 'LION', 'LP',                          # JNE/Lion
                            'KERRY', 'KE',                                # Kerry
                            'WHN', 'RPX', 'TIKI',                         # Wahana/RPX/Tiki
                            'BLIJC', 'BLI',                               # Blibli
                            'GB', 'GRAB',                                 # Grab
                            'POS',                                        # Pos
                        ]
                        if any(c_upper.startswith(p) for p in known_prefixes):
                            primary_identifier = c
                            break
                        if c_upper.startswith('00') and len(c) >= 12:
                            primary_identifier = c
                            break
                    
                    if not primary_identifier and candidates:
                        for c in candidates:
                            if len(c) >= 10:
                                primary_identifier = c
                                break
                    
                    # Fallback ke Order ID jika tidak ada AWB
                    if not primary_identifier and order_ids:
                        primary_identifier = order_ids[0]
                
                if primary_identifier:
                    pdf_awbs.add(primary_identifier)
                
                clear_start_y = find_clear_start_position(page)
                
                all_pages.append({
                    'page_num': page_num,
                    'page': page,
                    'doc': doc,
                    'width': page.rect.width,
                    'height': page.rect.height,
                    'awb': matched_awb,
                    'primary_pdf_awb': primary_identifier,
                    'is_junk': page_is_junk,
                    'is_continuation': page_is_continuation,
                    'clear_start_y': clear_start_y
                })

        # 4. Calculate match statistics
        matched_awbs = set()
        for page_info in all_pages:
            if page_info['awb']:
                matched_awbs.add(page_info['awb'])
        
        # Debug log to file
        with open("debug.log", "a") as f:
            f.write(f"\n=== MATCHING SUMMARY (with-stats) ===\n")
            f.write(f"Total PDF pages: {len(all_pages)}\n")
            f.write(f"Pages with matched AWB: {len([p for p in all_pages if p['awb']])}\n")
            f.write(f"Pages without matched AWB: {len([p for p in all_pages if not p['awb']])}\n")
            f.write(f"Excel unique identifiers: {len(excel_awbs)}\n")
            f.write(f"Sample Excel AWBs: {list(excel_awbs)[:10]}\n")
            f.write(f"Sample PDF AWBs: {list(pdf_awbs)[:10]}\n")
            f.write(f"Matched AWBs: {len(matched_awbs)}\n")
        
        # Order di Excel yang tidak ditemukan match di PDF
        # Gunakan canonical IDs: resolve semua matched_awbs ke canonical ID Pesanan
        matched_canonical_ids = set()
        for awb_key in matched_awbs:
            # Jika ini adalah AWB (bukan ID Pesanan), resolve ke ID Pesanan-nya
            if awb_key in awb_to_id_mapping:
                matched_canonical_ids.add(awb_to_id_mapping[awb_key])
            else:
                matched_canonical_ids.add(awb_key)
        
        # Bandingkan hanya primary IDs (bukan semua AWB + ID)
        unmatched_excel = excel_primary_ids - matched_canonical_ids
        
        # AWB di PDF yang tidak ada di Excel
        # Hanya ambil dari pages yang TIDAK matched, punya primary identifier,
        # DAN bukan junk page atau continuation page (lanjutan pretelan)
        unmatched_pdf = set()
        for page_info in all_pages:
            if not page_info['awb']:  # Page tidak matched dengan Excel
                # Skip junk pages dan halaman lanjutan (pretelan) kita sendiri
                if page_info.get('is_junk', False) or page_info.get('is_continuation', False):
                    print(f"[STATS] Page {page_info['page_num'] + 1} skipped from PDF Only: junk={page_info.get('is_junk')}, continuation={page_info.get('is_continuation')}")
                    continue
                primary = page_info.get('primary_pdf_awb')
                if primary:
                    unmatched_pdf.add(primary)
                    print(f"[STATS] Page {page_info['page_num'] + 1} counted as PDF Only: primary='{primary}'")
        
        master_doc = fitz.open()
        matched_count = 0
        label_number = 0
        processed_awbs = {}  # Track AWB -> last_page_num yang sudah diproses
        duplicate_count = 0
        duplicate_awbs = []  # Track detail AWB yang duplikat
        continuation_pages = []  # Track halaman lanjutan (pretelan)
        unique_matched_ids = set() # Track unique Order IDs for accurate matched count
        
        # RECAP DATA for priority_kembar summary
        recap_dict = {} # MSKU -> total_qty

        # 5. is_extended dan rak_map sudah di-fetch sebelum sort di atas.
        
        # --- CUSTOM SORT: Urutan label mengikuti urutan baris di Excel ---
        # Buat mapping: AWB/ID → urutan kemunculan pertama di Excel (index)
        excel_order = {}
        excel_order_idx = 0
        for _, row in df.iterrows():
            identifier = None
            if has_id_pesanan:
                id_val = normalize_awb(row.get('ID Pesanan', ''))
                if id_val and id_val not in ['NAN', 'NONE', 'NULL', '']:
                    identifier = id_val
            if not identifier and has_awb:
                awb_val = normalize_awb(row.get('AWB/No. Tracking', ''))
                if awb_val and awb_val not in ['NAN', 'NONE', 'NULL', '']:
                    identifier = awb_val
            if identifier and identifier not in excel_order:
                excel_order[identifier] = excel_order_idx
                excel_order_idx += 1

        import re
        def natural_sort_key(s):
            normalized_s = str(s).strip().upper()
            if normalized_s.startswith('BT-'):
                normalized_s = normalized_s.replace('BT-', 'BT1-', 1)
            elif normalized_s.startswith('T-'):
                normalized_s = normalized_s.replace('T-', 'T2-', 1)
            return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', normalized_s)]

        def get_excel_sort_key(page_info):
            awb = page_info.get('awb')
            
            sku_count_priority = 0
            rak_val = "ZZZZ"
            id_val = "ZZZZ"
            msku_val = "ZZZZ"
            
            excel_idx = 999999
            
            # 1. Cari urutan Excel
            if awb and awb in excel_order:
                excel_idx = excel_order[awb]
            elif awb and awb in awb_to_id_mapping and awb_to_id_mapping[awb] in excel_order:
                excel_idx = excel_order[awb_to_id_mapping[awb]]
            elif awb and awb in id_to_awb_mapping and id_to_awb_mapping[awb] in excel_order:
                excel_idx = excel_order[id_to_awb_mapping[awb]]
                
            # 2. Hitung jumlah SKU dan ambil data Rak jika = 1
            if sort_by_sku_count and awb:
                canonical_for_count = awb
                if awb in awb_to_id_mapping:
                    canonical_for_count = awb_to_id_mapping[awb]
                    
                items = []
                if canonical_for_count in awb_to_items:
                    items = awb_to_items[canonical_for_count]
                elif awb in awb_to_items:
                    items = awb_to_items[awb]
                    
                sku_count_priority = len(items)
                
                # JIKA SKU SATUAN, Ambil Rak, ID, MSKU
                if sku_count_priority == 1:
                    item = items[0]
                    msku_val = str(item.get('msku', '')).strip().upper()
                    
                    # rak_map sudah ada di memory
                    info = rak_map.get(msku_val, {})
                    r_v = str(info.get('rak', '')).strip().upper()
                    i_v = str(info.get('id', '')).strip().upper()
                    
                    if r_v and r_v != '-': rak_val = r_v
                    if i_v and i_v != '-': id_val = i_v

            if not sort_by_sku_count:
                # Jika bukan dari Upload 2, abaikan rak dan jumlah SKU
                return (0, 0, natural_sort_key("ZZZZ"), natural_sort_key("ZZZZ"), natural_sort_key("ZZZZ"), excel_idx)

            if not awb:
                return (999, 2, natural_sort_key("ZZZZ"), natural_sort_key("ZZZZ"), natural_sort_key("ZZZZ"), page_info.get('page_num', 9999))
                
            # Return tuple panjang agar tidak terjadi TypeError saat perbandingan
            # Menggunakan natural_sort_key agar "3" berada di atas "17"
            return (sku_count_priority, 0, natural_sort_key(rak_val), natural_sort_key(id_val), natural_sort_key(msku_val), excel_idx)

        all_pages.sort(key=get_excel_sort_key)
        
        if sort_by_sku_count:
            print(f"[SORT] SKU-Count sorting applied (Upload 2). {len(excel_order)} unique identifiers mapped.")
        else:
            print(f"[SORT] Excel-order sorting applied. {len(excel_order)} unique identifiers mapped from Excel.")

        # --- PRIORITY KEMBAR SORTING (Upload Kembar Only) ---
        # Catatan: Jika priority_kembar=True, sorting di bawah ini akan OVERRIDE urutan Excel di atas
        if priority_kembar:
            print("[SORT] Applying Priority Kembar sorting...")
            # Fingerprints based on MSKU sets (must have >= 2 unique MSKUs)
            page_fingerprints = {}  # Page Index -> tuple
            fp_counts = {}         # tuple -> count
            
            for idx, page_info in enumerate(all_pages):
                awb = page_info['awb']
                if awb and awb in awb_to_items:
                    items = awb_to_items[awb]
                    unique_mskus = sorted(list(set([str(i['msku']).strip().upper() for i in items])))
                    if len(unique_mskus) >= 1:
                        fp = tuple(unique_mskus)
                        page_fingerprints[idx] = fp
                        fp_counts[fp] = fp_counts.get(fp, 0) + 1
            
            # Add metadata for sorting
            for idx, page_info in enumerate(all_pages):
                fp = page_fingerprints.get(idx)
                is_twin = fp and fp_counts.get(fp, 0) >= 2
                page_info['is_kembar_priority'] = is_twin
                page_info['kembar_fingerprint'] = fp
                page_info['original_idx'] = idx
            
            # Sort: Twin Priority (0) -> Fingerprint -> Original Index
            all_pages.sort(key=lambda x: (
                0 if x.get('is_kembar_priority') else 1,
                x.get('kembar_fingerprint') or (),
                x.get('original_idx', 0)
            ))
            print(f"[SORT] Sorting complete. {len([p for p in all_pages if p.get('is_kembar_priority')])} pages prioritized.")

        # --- PRE-SCAN MSKU FREQUENCY (For Overlap Filtering) ---
        twin_mskus = set()
        if priority_kembar:
            msku_to_ids = {} # MSKU -> set of Canonical IDs
            for page_info in all_pages:
                awb = page_info['awb']
                # Filter ketat: Hanya hitung dari halaman label asli yang valid
                if not awb or awb not in awb_to_items:
                    continue
                if page_info.get('is_junk') or page_info.get('is_continuation'):
                    continue
                
                # Gunakan Canonical ID (Order ID) agar duplikat AWB tetap dihitung 1 order
                canonical_id = awb
                if awb in awb_to_id_mapping:
                    canonical_id = awb_to_id_mapping[awb]
                
                for item in awb_to_items[awb]:
                    m = str(item['msku']).strip().upper()
                    if m not in msku_to_ids: msku_to_ids[m] = set()
                    msku_to_ids[m].add(canonical_id)
            
            # Hanya rekap barang yang muncul di minimal 2 ORDER BERBEDA di PDF
            twin_mskus = {m for m, ids in msku_to_ids.items() if len(ids) >= 2}
            print(f"[RECAP] Found {len(twin_mskus)} shared MSKUs across different orders.")

        # 6. Proses setiap halaman (sama seperti endpoint biasa)
        for page_info in all_pages:
            page_awb = page_info['awb']
            page_num = page_info['page_num']
            W_pts = page_info['width']
            H_pts = page_info['height']
            clear_start_y = page_info['clear_start_y']
            table_start_y = clear_start_y
            
            if not page_awb or page_awb not in awb_to_items:
                if page_info.get('is_junk', False) or page_info.get('is_continuation', False):
                    print(f"[DEBUG] Page {page_num + 1} is junk/continuation. Skipping raw insert.")
                    continue
                # Page tanpa AWB match (misal cover PACKING LIST) - insert raw
                print(f"[DEBUG] Page {page_num + 1} not matched. Inserting raw page.")
                temp_doc = fitz.open()
                temp_doc.insert_pdf(page_info['doc'], from_page=page_num, to_page=page_num)
                master_doc.insert_pdf(temp_doc)
                temp_doc.close()
                continue
            
            # CEK DUPLIKAT vs PRETELAN (halaman lanjutan)
            status = "NEW"
            prev_page = -1
            
            # Debug Trace
            canonical_trace = page_awb
            if page_awb in awb_to_id_mapping:
                canonical_trace = awb_to_id_mapping[page_awb]
            
            print(f"[DEBUG] Processing Page {page_num + 1}: Raw='{page_awb}', Canonical='{canonical_trace}'")
            print(f"        processed_keys: {list(processed_awbs.keys())[:5]}... (Total: {len(processed_awbs)})")
            
            if page_awb in processed_awbs:
                last_page = processed_awbs[page_awb]
                prev_page = last_page
                awb_display = id_to_awb_mapping.get(page_awb, page_awb)
                
                print(f"        -> HIT processed_awbs via Raw ID")

                # Jika halaman berurutan (selisih <= 2), ini adalah PRETELAN (halaman lanjutan)
                if abs(page_num - last_page) <= 2:
                    status = "CONT"
                    print(f"  [CONT] PRETELAN (lanjutan): ID={page_awb}, AWB={awb_display}, Page={page_num + 1} (prev: {last_page + 1})")
                    continuation_pages.append({
                        'id_pesanan': page_awb,
                        'awb': awb_display,
                        'pdf_page': page_num + 1,
                        'first_page': last_page + 1
                    })
                    # Skip halaman lanjutan (pretelan) saja
                    continue
                # Gap > 2: JANGAN skip!
                # Bisa saja AWB yang sama muncul via text extraction di halaman lain (false positive)
                # Label tetap diproses, dedup via canonical ID di bawah
                print(f"  [INFO] AWB '{page_awb}' muncul lagi di page {page_num + 1} (gap > 2). Diproses, cek canonical dedup.")
            
            # Simpan page_awb sebagai sudah diproses
            processed_awbs[page_awb] = page_num
            
            # TRACK unique canonical Order ID
            # Jika canonical_id sudah ada = order ini sudah dirender sebelumnya (pretelan/duplikat)
            # → SKIP rendering label baru agar label_number tidak loncat
            canonical_id = page_awb
            if page_awb in awb_to_id_mapping:
                canonical_id = awb_to_id_mapping[page_awb]
            
            if canonical_id in unique_matched_ids:
                # Order sudah diproses sebelumnya, halaman ini adalah duplikat pretelan (gap > 2)
                print(f"  [SKIP-DUP] Canonical ID '{canonical_id}' sudah dirender. Halaman {page_num + 1} dilewati (duplikat pretelan gap>2).")
                continue
            
            unique_matched_ids.add(canonical_id)
            matched_count = len(unique_matched_ids)
            print(f"        -> Unique Track: Canonical='{canonical_id}', Count={matched_count}")
            
            items = awb_to_items[page_awb]
            
            # --- START PRIORITY TOP LOGIC ---
            if is_priority_top_active and len(items) > 1:
                def sort_item_key(item):
                    msku = str(item.get('msku', '')).strip().upper()
                    # 1. Is it a bulky SKU? (0 if yes, 1 if no) -> 0 will sort to TOP
                    bulky_score = 0 if msku in special_skus else 1
                    
                    # 2. Rak & ID
                    info = rak_map.get(msku, {})
                    rak_str = str(info.get('rak', '')).strip().upper()
                    if not rak_str or rak_str == '-':
                        rak_str = 'ZZZZ'
                        
                    id_str = str(info.get('id', '')).strip().upper()
                    if not id_str or id_str == '-':
                        id_str = 'ZZZZ'
                        
                    rak_val = natural_sort_key(rak_str)
                    id_val = natural_sort_key(id_str)
                    
                    # 3. MSKU
                    msku_val = natural_sort_key(msku)
                    return (bulky_score, rak_val, id_val, msku_val)
                    
                items = sorted(items, key=sort_item_key)
            # --- END PRIORITY TOP LOGIC ---
            
            # Add to recap if priority_kembar AND item is shared (twin)
            if priority_kembar:
                for item in items:
                    msku = str(item['msku']).strip().upper()
                    # Hanya rekap barang yang muncul di > 1 label
                    if msku in twin_mskus:
                        qty = int(item['jumlah'])
                        recap_dict[msku] = recap_dict.get(msku, 0) + qty
            
            # Tentukan row height sebenarnya dari config
            if is_extended:
                _ext_max_f = max(
                    float(label_cfg.get('ext_font_rak') or label_cfg.get('ext_font_size', 8)),
                    float(label_cfg.get('ext_font_msku') or label_cfg.get('ext_font_size', 8)),
                    float(label_cfg.get('ext_font_qty') or label_cfg.get('ext_font_size', 8))
                )
                actual_row_height = max(float(label_cfg.get('ext_row_height', 18)), _ext_max_f * 2.25)
            else:
                _std_max_f = max(
                    float(label_cfg.get('std_font_msku') or label_cfg.get('std_font_size', 8)),
                    float(label_cfg.get('std_font_qty') or label_cfg.get('std_font_size', 8))
                )
                actual_row_height = max(float(label_cfg.get('std_row_height', 18)), _std_max_f * 2.25)
                
            is_blibli_page = page_awb.startswith('BLI')
            margin_bottom = (65 if is_blibli_page else 50) if W_pts > 350 else (45 if is_blibli_page else 35)
            available_rows = calculate_available_rows(H_pts, table_start_y, row_height=actual_row_height, margin_bottom=margin_bottom)
            # Pagination dinamis — ITEM-aware untuk extended
            limit_extra = max(1, calculate_available_rows(H_pts, 50, row_height=actual_row_height, margin_bottom=margin_bottom))
            
            available_height_pts = max(0, H_pts - table_start_y - margin_bottom - 2)
            limit_extra_pts = max(0, H_pts - 50 - (margin_bottom - 10) - 2)
            
            scaled_cfg = scale_cfg_for_wide_label(label_cfg, W_pts, is_extended)
            
            # --- Uniform Page 1 Scaling ---
            worst_case_h = 70.0
            scale_f = 1.0
            if available_height_pts < worst_case_h and available_height_pts > 0:
                scale_f = max(0.65, available_height_pts / worst_case_h)
            page1_cfg = shrink_cfg_for_page1(scaled_cfg, scale_f)
            
            chunks = []
            rest = items
            first_page = True
            while rest:
                ah = available_height_pts if first_page else limit_extra_pts
                current_cfg = page1_cfg if first_page else scaled_cfg
                
                next_chunk, rest_after = calc_items_for_rows_by_height(rest, is_extended, rak_map, ah, current_cfg, W_pts, force_first=first_page)
                
                # Enforce max 2 items on the first page for multi-item (resi pretelan)
                if first_page and len(items) > 1:
                    if len(next_chunk) > 2:
                        next_chunk = rest[:2]
                        rest_after = rest[2:]
                
                if not next_chunk:
                    if first_page:
                        # Should not happen because of force_first, but fallback
                        chunks.append({'items': [], 'cfg': current_cfg})
                        first_page = False
                        continue
                    else:
                        next_chunk = rest[:1]
                        rest_after = rest[1:]
                        
                chunks.append({'items': next_chunk, 'cfg': current_cfg})
                rest = rest_after
                first_page = False

            total_pages = len(chunks)
            for i, chunk_data in enumerate(chunks):
                chunk = chunk_data['items']
                current_cfg = chunk_data['cfg']
                is_last_chunk = (i == total_pages - 1)
                
                if i == 0:
                    temp_doc = fitz.open()
                    temp_doc.insert_pdf(page_info['doc'], from_page=page_num, to_page=page_num)
                    page_copy = temp_doc[0]
                    
                    rect_clear = fitz.Rect(0, clear_start_y, max(W_pts, page_copy.rect.width), max(H_pts, page_copy.rect.height))
                    page_copy.add_redact_annot(rect_clear, fill=(1, 1, 1))
                    page_copy.apply_redactions()
                    page_copy.set_cropbox(fitz.Rect(0, 0, W_pts, H_pts))
                    
                    packet = io.BytesIO()
                    can = canvas.Canvas(packet, pagesize=(W_pts, H_pts))
                    
                    # FITUR BARU: Picker Name di bagian atas tabel (Solid, Normal)
                    if picker_name:
                        can.saveState()
                        can.setFillColorRGB(0, 0, 0)  # Hitam Solid agar jelas di thermal
                        
                        text_to_draw = f"PICKER: {str(picker_name).strip().upper()}"
                        f_size = min(18, max(11, 11 * (W_pts / 283.0))) # Skala font picker
                        can.setFont("Helvetica-Bold", f_size)
                        
                        # Letakkan di tengah atas dari area custom label
                        # y ditarik sedikit ke bawah dari clear_start_y
                        label_y = H_pts - clear_start_y - 20
                        
                        can.drawCentredString(W_pts / 2, label_y, text_to_draw)
                        can.restoreState()
                    
                    if chunk:
                        table_result = generate_table_data(chunk, is_extended, rak_map, current_cfg, picker_name=picker_name)
                        t = create_table(
                            table_result['table_data'],
                            row_heights=table_result['row_heights'],
                            span_cmds=table_result['span_cmds'],
                            label_cfg=current_cfg,
                        )
                        t.wrapOn(can, W_pts, H_pts)
                        total_table_h = sum(table_result['row_heights'])
                        # table_y dari bawah. Kurangi padding atas (2pt) agar sangat maksimal.
                        table_y = H_pts - table_start_y - total_table_h - 2
                        # Batas bawah aman: 20px dari bawah agar tidak timpa No:xxx (y=10)
                        safe_bottom = 35 if is_blibli_page else 20
                        t.drawOn(can, 7, max(safe_bottom, table_y))
                    else:
                        table_y = H_pts - table_start_y - 10
                        if not is_last_chunk:
                            can.saveState()
                            can.setFillColorRGB(0.5, 0.5, 0.5)
                            can.setFont("Helvetica-Oblique", 11)
                            # Gambar teks di tengah-tengah sisa ruang yang kosong
                            middle_y = (H_pts - clear_start_y) / 2
                            can.drawCentredString(W_pts / 2, middle_y, "Data pesanan berada di halaman berikutnya")
                            can.restoreState()
                    
                    # Customer Message HANYA di chunk terakhir
                    catatan_to_add = None
                    if is_last_chunk:
                        # Catatan Pembeli untuk semua label (tidak hanya JX)
                        catatan = awb_to_catatan.get(page_awb, '')
                        if catatan:
                            # Clean text: remove line breaks, keep unicode
                            catatan_to_add = clean_text_for_pdf(catatan)
                    
                    # Cek ruang untuk Customer Message
                    msg_space_needed = 0
                    msg_lines = []
                    if catatan_to_add:
                        max_chars = 55
                        remaining_text = catatan_to_add
                        while remaining_text:
                            if len(remaining_text) <= max_chars:
                                msg_lines.append(remaining_text)
                                break
                            else:
                                split_at = remaining_text[:max_chars].rfind(' ')
                                if split_at == -1:
                                    split_at = max_chars
                                msg_lines.append(remaining_text[:split_at])
                                remaining_text = remaining_text[split_at:].strip()
                        msg_space_needed = len(msg_lines) * 10 + 15  # 10pt per line + margin
                    
                    # Cek apakah cukup ruang (min 30pt dari bawah)
                    space_available = max(10, table_y) - 30
                    
                    if catatan_to_add and space_available >= msg_space_needed:
                        # Ada ruang - render di halaman ini
                        msg_y = max(10, table_y) - 12
                        can.setFont("Helvetica-Bold", 7)
                        msg_prefix = "Customer Message: "
                        can.drawString(7, msg_y, msg_prefix)
                        prefix_width = can.stringWidth(msg_prefix, "Helvetica-Bold", 7)
                        
                        # Use Unicode font for message content
                        can.setFont(UNICODE_FONT, 7)
                        if msg_lines:
                            can.drawString(7 + prefix_width, msg_y, msg_lines[0])
                            for j, line in enumerate(msg_lines[1:], 1):
                                can.drawString(7, msg_y - (j * 10), line)
                    
                    # Add label number (increment di sini agar berurutan berdasarkan output, bukan input PDF)
                    label_number += 1
                    label_text = f"No:{label_number:03d}"
                    no_font_size = 11 if W_pts > 350 else 7
                    can.setFont("Helvetica-Bold", no_font_size)
                    text_width = can.stringWidth(label_text, "Helvetica-Bold", no_font_size)
                    x_pos = W_pts - text_width - 10
                    y_pos = 10
                    can.drawString(x_pos, y_pos, label_text)
                    
                    # ===== INDIKATOR HALAMAN (BERSAMBUNG / AKHIR) =====
                    if total_pages > 1:
                        indicator_text = ">>> BERSAMBUNG KE HALAMAN BERIKUTNYA >>>" if not is_last_chunk else "--- AKHIR DARI PESANAN INI ---"
                        ind_font_size = 12 if W_pts > 350 else 8
                        can.setFont("Helvetica-Bold", ind_font_size)
                        ind_width = can.stringWidth(indicator_text, "Helvetica-Bold", ind_font_size)
                        can.drawString((W_pts - ind_width) / 2, 10, indicator_text)
                    
                    can.save()
                    
                    packet.seek(0)
                    overlay_pdf = fitz.open("pdf", packet.read())
                    cleared_area = fitz.Rect(0, clear_start_y, W_pts, H_pts)
                    page_copy.show_pdf_page(page_copy.rect, overlay_pdf, 0)
                    master_doc.insert_pdf(temp_doc, from_page=0, to_page=0)
                    overlay_pdf.close()
                    temp_doc.close()
                    
                    # Jika tidak cukup ruang, buat halaman baru untuk Customer Message
                    if catatan_to_add and space_available < msg_space_needed:
                        packet2 = io.BytesIO()
                        can2 = canvas.Canvas(packet2, pagesize=(W_pts, H_pts))
                        
                        can2.setFont("Helvetica-Bold", 10)
                        can2.drawString(10, H_pts - 25, f"Lanjutan AWB: {page_awb}")
                        
                        msg_y2 = H_pts - 50
                        can2.setFont("Helvetica-Bold", 8)
                        msg_prefix = "Customer Message: "
                        can2.drawString(10, msg_y2, msg_prefix)
                        prefix_width = can2.stringWidth(msg_prefix, "Helvetica-Bold", 8)
                        
                        can2.setFont(UNICODE_FONT, 8)
                        if msg_lines:
                            can2.drawString(10 + prefix_width, msg_y2, msg_lines[0])
                            for j, line in enumerate(msg_lines[1:], 1):
                                can2.drawString(10, msg_y2 - (j * 12), line)
                        
                        can2.save()
                        packet2.seek(0)
                        msg_page_doc = fitz.open("pdf", packet2.read())
                        master_doc.insert_pdf(msg_page_doc)
                        msg_page_doc.close()
                        
                else:
                    # Halaman lanjutan
                    packet = io.BytesIO()
                    can = canvas.Canvas(packet, pagesize=(W_pts, H_pts))
                    
                    can.setFont("Helvetica-Bold", 10)
                    hal_info = f" (Hal. {i+1} dari {total_pages})"
                    can.drawString(10, H_pts - 25, f"Lanjutan AWB: {page_awb}{hal_info}")
                    
                    table_result = generate_table_data(chunk, is_extended, rak_map, scaled_cfg, picker_name=picker_name)
                    t = create_table(
                        table_result['table_data'],
                        row_heights=table_result['row_heights'],
                        span_cmds=table_result['span_cmds'],
                        label_cfg=scaled_cfg,
                    )
                    t.wrapOn(can, W_pts, H_pts)
                    total_table_h = sum(table_result['row_heights'])
                    t_height = total_table_h
                    table_y_cont = H_pts - t_height - 45
                    t.drawOn(can, 7, table_y_cont)
                    
                    # Customer Message di halaman lanjutan TERAKHIR
                    if is_last_chunk:
                        # Catatan Pembeli untuk semua label (tidak hanya JX)
                        catatan = awb_to_catatan.get(page_awb, '')
                        if catatan:
                                # Clean text
                                catatan_clean = clean_text_for_pdf(catatan)
                                max_chars = 55
                                msg_lines_cont = []
                                remaining_text = catatan_clean
                                while remaining_text:
                                    if len(remaining_text) <= max_chars:
                                        msg_lines_cont.append(remaining_text)
                                        break
                                    else:
                                        split_at = remaining_text[:max_chars].rfind(' ')
                                        if split_at == -1:
                                            split_at = max_chars
                                        msg_lines_cont.append(remaining_text[:split_at])
                                        remaining_text = remaining_text[split_at:].strip()
                                
                                msg_y_cont = table_y_cont - 15
                                if msg_y_cont > 30:  # Cukup ruang
                                    can.setFont("Helvetica-Bold", 7)
                                    msg_prefix = "Customer Message: "
                                    can.drawString(7, msg_y_cont, msg_prefix)
                                    prefix_width = can.stringWidth(msg_prefix, "Helvetica-Bold", 7)
                                    
                                    can.setFont(UNICODE_FONT, 7)
                                    if msg_lines_cont:
                                        can.drawString(7 + prefix_width, msg_y_cont, msg_lines_cont[0])
                                        for j, line in enumerate(msg_lines_cont[1:], 1):
                                            can.drawString(7, msg_y_cont - (j * 10), line)
                    
                    # ===== INDIKATOR HALAMAN (BERSAMBUNG / AKHIR) =====
                    if total_pages > 1:
                        indicator_text = ">>> BERSAMBUNG KE HALAMAN BERIKUTNYA >>>" if not is_last_chunk else "--- AKHIR DARI PESANAN INI ---"
                        ind_font_size = 12 if W_pts > 350 else 8
                        can.setFont("Helvetica-Bold", ind_font_size)
                        ind_width = can.stringWidth(indicator_text, "Helvetica-Bold", ind_font_size)
                        can.drawString((W_pts - ind_width) / 2, 10, indicator_text)
                    
                    can.save()
                    
                    packet.seek(0)
                    new_page_doc = fitz.open("pdf", packet.read())
                    master_doc.insert_pdf(new_page_doc)
                    new_page_doc.close()

        if len(master_doc) == 0:
            # Revert to strict error as requested
            print("[PROCESS] No matching labels found. Aborting.")
            raise HTTPException(
                status_code=400, 
                detail="DATA MISMATCH: Tidak ada label di PDF yang cocok dengan data Excel. Pastikan Anda mengunggah file yang benar."
            )

        # Convert PDF to base64
        
        # PREPEND RECAP PAGE if priority_kembar
        if priority_kembar and recap_dict:
            try:
                recap_pdf = generate_recap_pdf(recap_dict, rak_map, label_cfg)
                if len(recap_pdf) > 0:
                    final_master = fitz.open()
                    final_master.insert_pdf(recap_pdf)
                    final_master.insert_pdf(master_doc)
                    master_doc = final_master
                recap_pdf.close()
            except Exception as e_recap:
                print(f"[RECAP] Error generating recap page: {e_recap}")

        # APPEND SUMMARY PAGE for Upload 2
        if sort_by_sku_count:
            try:
                # Cek apakah fitur packing list dimatikan di admin
                features = await supabase_fetch("GET", "toolkit_feature_locks?select=feature_key,is_locked")
                is_packing_list_locked = False
                if isinstance(features, list):
                    for f in features:
                        if f.get("feature_key") == "packing-list-upload-2":
                            is_packing_list_locked = f.get("is_locked", False)
                            break
                
                if not is_packing_list_locked:
                    final_ids = []
                    seen_ids = set()
                    for page_info in all_pages:
                        if page_info.get('is_continuation'):
                            continue
                        awb = page_info.get('awb')
                        if awb:
                            canonical_id = awb
                            if awb in awb_to_id_mapping:
                                canonical_id = awb_to_id_mapping[awb]
                            if canonical_id not in seen_ids:
                                seen_ids.add(canonical_id)
                                final_ids.append(canonical_id)
                    
                    if final_ids:
                        summary_title = pdf_files[0].filename if pdf_files else "SUMMARY_LABEL.pdf"
                        summary_pdf = generate_summary_page(summary_title, final_ids)
                        if len(summary_pdf) > 0:
                            master_doc.insert_pdf(summary_pdf)
                        summary_pdf.close()
            except Exception as e_summary:
                print(f"[SUMMARY] Error generating summary page: {e_summary}")
                
        # APPEND GLOBAL MSKU SUMMARY (if requested)
        if include_global_msku:
            try:
                # Fetch barang khusus SKUs
                try:
                    bk_data = await supabase_fetch("GET", "sku_barang_khusus?select=sku")
                    barang_khusus_skus = {str(item['sku']).strip().upper() for item in bk_data} if isinstance(bk_data, list) else set()
                except Exception as e_bk:
                    print(f"Failed to fetch barang_khusus: {e_bk}")
                    barang_khusus_skus = set()
                    
                global_msku_pdf = generate_global_msku_summary_pdf(awb_to_items, unique_matched_ids, rak_map, label_cfg, barang_khusus_skus)
                if len(global_msku_pdf) > 0:
                    master_doc.insert_pdf(global_msku_pdf)
                global_msku_pdf.close()
            except Exception as e_msku:
                print(f"[GLOBAL_MSKU] Error generating global msku page: {e_msku}")

        output_stream = io.BytesIO()
        master_doc.save(output_stream)
        master_doc.close()
        output_stream.seek(0)
        result_pdf_bytes = output_stream.read()
        pdf_base64 = base64.b64encode(result_pdf_bytes).decode('utf-8')
        
        # ===== BACKUP FILES =====
        try:
            # 1. GENERATE FILTERED EXCEL FOR PACKING LIST (Only matched rows)
            filtered_excel_bytes = None
            try:
                # Logic: Keep row if 'ID Pesanan' OR 'AWB' match match_awbs set
                print(f"[PROCESS] Filtering Excel for backup. Matched IDs: {len(matched_awbs)}")
                
                def is_matched_for_filter(row):
                    # Check ID Pesanan
                    if has_id_pesanan:
                        val = row.get('ID Pesanan', '')
                        norm = normalize_awb(val)
                        if norm and norm in matched_awbs: return True
                    # Check AWB as fallback
                    if has_awb:
                        val = row.get('AWB/No. Tracking', '')
                        norm = normalize_awb(val)
                        if norm and norm in matched_awbs: return True
                    return False

                # Filter the ORIGINAL DataFrame (df)
                filtered_df = df[df.apply(is_matched_for_filter, axis=1)].copy() # Copy to avoid SettingWithCopy
                
                if not filtered_df.empty:
                    # Force ID columns to string to prevent scientific notation in Excel
                    if has_id_pesanan:
                        filtered_df['ID Pesanan'] = filtered_df['ID Pesanan'].astype(str)
                    if has_awb:
                        filtered_df['AWB/No. Tracking'] = filtered_df['AWB/No. Tracking'].astype(str)

                    f_buffer = io.BytesIO()
                    # Use xlsxwriter engine
                    with pd.ExcelWriter(f_buffer, engine='xlsxwriter') as writer:
                        filtered_df.to_excel(writer, index=False)
                        # Auto-adjust columns width
                        worksheet = writer.sheets['Sheet1']
                        for i, col in enumerate(filtered_df.columns):
                            max_len = max(filtered_df[col].astype(str).map(len).max(), len(col)) + 2
                            worksheet.set_column(i, i, max_len)
                            
                    filtered_excel_bytes = f_buffer.getvalue()
                    print(f"[PROCESS] Filtered Excel created: {len(filtered_df)} rows (from {len(df)})")
                else:
                    print("[PROCESS] Filtered Excel is empty, skipping.")
            except Exception as fe:
                print(f"[PROCESS] Failed to filter Excel: {fe}")

            # 2. Simpan combined PDF asli ke bytes
            combined_pdf_stream = io.BytesIO()
            combined_pdf_doc.save(combined_pdf_stream)
            combined_pdf_doc.close()
            combined_pdf_stream.seek(0)
            combined_pdf_bytes = combined_pdf_stream.read()
            
            # Panggil save_backup dengan filtered data
            save_backup(
                excel_filename=excel_filename,
                pdf_filenames=pdf_filenames,
                excel_data=excel_content,
                pdf_data=combined_pdf_bytes,
                result_pdf=result_pdf_bytes,
                filtered_excel_data=filtered_excel_bytes
            )
        except Exception as backup_error:
            print(f"[BACKUP] Warning: Backup failed but process continues: {backup_error}")

        # Return JSON with stats
        # Create matched list with AWB info
        # FIX: Resolve ID Pesanan correctly. matched_awbs contains KEYS (can be ID or AWB).
        matched_with_awb = []
        processed_ids = set() # To avoid duplicates if both ID and AWB matched for same order
        
        for key in matched_awbs:
            real_id = None
            real_awb = None
            
            # Scenario 1: Key IS ID Pesanan
            if key in id_to_awb_mapping:
                real_id = key
                real_awb = id_to_awb_mapping[key]
            
            # Scenario 2: Key IS AWB (Reverse lookup)
            elif key in awb_to_id_mapping:
                real_id = awb_to_id_mapping[key]
                real_awb = key # The key itself is the AWB
            
            # Scenario 3: Key matches but isolated (maybe logic edge case)
            else:
                # If identifier was ID but had no AWB?
                # Check if it's in awb_to_items keys?
                if key in awb_to_items:
                    # It's a valid key pointing to items.
                    # Assume it's ID if not in reverse map?
                    real_id = key
                    real_awb = ''
            
            if real_id and real_id not in processed_ids:
                matched_with_awb.append({
                    'id_pesanan': real_id,
                    'awb': real_awb
                })
                processed_ids.add(real_id)
        
        # Create unmatched excel list with AWB info
        unmatched_excel_with_awb = []
        for id_pesanan in unmatched_excel:
            awb = id_to_awb_mapping.get(id_pesanan, '')
            unmatched_excel_with_awb.append({
                'id_pesanan': id_pesanan,
                'awb': awb
            })
        
        return JSONResponse({
            "success": True,
            "timestamp": datetime.now().isoformat(),
            "pdf_base64": pdf_base64,
            "stats": {
                "excel_filename": excel_filename,
                "pdf_filenames": pdf_filenames,
                "total_excel_awb": len(excel_awbs),
                "total_pdf_pages": len(all_pages),
                "matched_count": matched_count, # FIX: Use calculated unique count (deduplicated)
                "duplicate_count": duplicate_count,  # Duplikat asli (halaman tidak berurutan)
                "duplicate_awbs": duplicate_awbs,
                "continuation_count": len(continuation_pages),  # Pretelan (halaman lanjutan)
                "continuation_pages": continuation_pages,
                "unmatched_excel_count": len(unmatched_excel), # Jumlah order di Excel yang tidak ada di PDF
                "unmatched_pdf_count": len(unmatched_pdf),
                "matched_awbs": list(unique_matched_ids), # Use unique IDs list
                "matched_with_awb": matched_with_awb,
                "unmatched_excel_awbs": list(unmatched_excel),
                "unmatched_excel_with_awb": unmatched_excel_with_awb,
                "unmatched_excel_with_awb": unmatched_excel_with_awb,
                "unmatched_pdf_awbs": list(unmatched_pdf),
                "all_excel_awbs": list(excel_awbs_raw_set),
                "id_to_awb_mapping": id_to_awb_mapping
            }
        })

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/history/{id}")
async def delete_history(id: str, username: Optional[str] = None):
    """
    Delete history record and associated files.
    """
    try:
        # 1. Get record details first to find filenames
        data = await supabase_fetch("GET", f"label_process_history?id=eq.{id}&select=*")
        if not data:
            raise HTTPException(status_code=404, detail="History not found")
        
        record = data[0]
        
        # Security validation for User delete
        if username:
            if record.get('username') != username:
                raise HTTPException(status_code=403, detail="Akses ditolak: Anda bukan pemroses data ini.")
            
            created_at_str = record.get('created_at')
            if created_at_str:
                try:
                    if created_at_str.endswith('Z'):
                        created_at_str = created_at_str.replace('Z', '+00:00')
                    from datetime import datetime, timezone
                    record_time = datetime.fromisoformat(created_at_str)
                    now_time = datetime.now(timezone.utc)
                    diff_minutes = (now_time - record_time).total_seconds() / 60
                    
                    if diff_minutes > 60:
                        raise HTTPException(status_code=403, detail="Akses ditolak: Waktu penghapusan (60 menit) telah habis.")
                except Exception as parse_err:
                    print(f"Error parsing date {created_at_str}: {parse_err}")
                    raise HTTPException(status_code=500, detail="Gagal memverifikasi waktu data.")
        excel_filename = record.get('excel_filename')
        # pdf_filenames might be a list or string depending on how it was saved. In recent implementation it's a list.
        pdf_filenames = record.get('pdf_filenames', []) 
        created_at = record.get('created_at') # date string for folder structure

        # 2. Delete Physical Files (Backup Folder)
        # FIX: Gunakan find_backup_folder() untuk menemukan folder backup yang tepat
        # Logika lama salah karena mencari di subfolder tanggal (YYYY-MM-DD) yang tidak ada di struktur backup.
        if created_at and excel_filename:
            try:
                # Ambil nama PDF pertama sebagai identifier batch
                first_pdf_name = None
                if isinstance(pdf_filenames, list) and pdf_filenames:
                    first_pdf_name = pdf_filenames[0].replace('.pdf', '').replace('.PDF', '')
                elif isinstance(pdf_filenames, str) and pdf_filenames:
                    first_pdf_name = pdf_filenames.replace('.pdf', '').replace('.PDF', '')
                
                print(f"[DELETE] Searching backup folder: excel={excel_filename}, pdf={first_pdf_name}")
                
                # Gunakan find_backup_folder (mendukung 3 strategi pencocokan)
                backup_folder = find_backup_folder(
                    date_str=created_at,
                    excel_filename=excel_filename,
                    required_pdf_name=first_pdf_name
                )
                
                if backup_folder and backup_folder.exists():
                    shutil.rmtree(backup_folder)
                    print(f"[DELETE] ✅ Backup folder deleted: {backup_folder}")
                else:
                    print(f"[DELETE] ⚠️ Backup folder tidak ditemukan (mungkin sudah dihapus atau > 7 hari)")
                    
            except Exception as e:
                print(f"[DELETE] Error deleting backup folder: {e}")
                # Lanjutkan delete DB meski folder fisik gagal dihapus

        # 3. Delete from Supabase tables
        # A. Delete processed items via matched_awbs (lebih aman: tidak hapus batch lain)
        # FIX: Dulu delete semua processed_items by excel_filename, bisa hapus data batch lain
        matched_awbs_list = record.get('matched_awbs', [])
        if matched_awbs_list and isinstance(matched_awbs_list, list) and len(matched_awbs_list) > 0:
            print(f"[DELETE] Cascading delete for {len(matched_awbs_list)} matched AWBs")
            chunk_size = 30
            for i in range(0, len(matched_awbs_list), chunk_size):
                chunk = matched_awbs_list[i:i + chunk_size]
                awb_filter = ','.join([urllib.parse.quote(str(a)) for a in chunk])
                await supabase_fetch("DELETE", f"processed_items?order_id=in.({awb_filter})")
        elif excel_filename:
            # Fallback jika matched_awbs tidak tersimpan (data lama)
            print(f"[DELETE] Cascading delete fallback by excel_filename: {excel_filename}")
            await supabase_fetch("DELETE", f"processed_items?excel_filename=eq.{urllib.parse.quote(excel_filename)}")

        # B. Delete the history record itself
        await supabase_fetch("DELETE", f"label_process_history?id=eq.{id}")
        
        return {"success": True, "message": "History deleted"}

    except Exception as e:
        print(f"Delete History Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class SabotageManipulateRequest(BaseModel):
    target_folder: str
    first_pdf_name: str
    first_excel_name: str
    second_pdf_name: str
    second_excel_name: str
    backup_folder: str = ""

@app.post("/process-labels-with-stats")
async def process_labels_with_stats(
    excel_file: UploadFile = File(...),
    pdf_files: list[UploadFile] = File(...),
    priority_kembar: bool = Form(False),
    sort_by_sku_count: bool = Form(False),
    include_global_msku: bool = Form(False),
    picker_name: Optional[str] = Form(None)
):
    """
    Process labels and return match statistics along with the PDF.
    Returns JSON with:
    - pdf_base64: Base64 encoded PDF
    - stats: Match statistics (matched, unmatched_excel, unmatched_pdf)
    """
    # Debug log to file
    with open("debug.log", "w") as f:
        f.write(f"=== PROCESS LABELS WITH STATS START ===\n")
        f.write(f"PDF files count: {len(pdf_files)}\n")
        for i, pdf in enumerate(pdf_files):
            f.write(f"  PDF {i+1}: {pdf.filename}\n")
        f.write(f"Excel file: {excel_file.filename}\n")
    
    try:
        # --- Fetch configurations (Custom Label Priority Top) ---
        is_priority_top_active = True
        special_skus = set()
        
        if sort_by_sku_count:
            try:
                features = await supabase_fetch("GET", "toolkit_feature_locks?select=feature_key,is_locked")
                if isinstance(features, list):
                    for f in features:
                        if f.get("feature_key") == "custom-label-priority-top" and f.get("is_locked", False):
                            is_priority_top_active = False
                            break
                            
                if is_priority_top_active:
                    priority_data = await supabase_fetch("GET", "sku_priority_bottom?select=sku")
                    if isinstance(priority_data, list):
                        special_skus = {str(r.get('sku')).strip().upper() for r in priority_data}
            except Exception as e:
                print(f"Error fetching custom label priority settings: {e}")
        # --------------------------------------------------------

        # PL Content Check (Early Validation)
        for pdf in pdf_files:
            content_check = await pdf.read()
            if is_picking_list(content_check, pdf.filename):
                 raise HTTPException(status_code=400, detail={"code": "PL_DETECTED", "message": f"File '{pdf.filename}' terdeteksi sebagai Picking List (PL). Mohon upload Resi Asli."})
            await pdf.seek(0)
            
        # 1. Baca File Excel Ginee
        excel_content = await excel_file.read()
        excel_filename = excel_file.filename
        df = pd.read_excel(io.BytesIO(excel_content), dtype=object)
        df = fix_excel_numeric_ids(df)
        
        # Normalize kolom
        col_mapping = {
            'AWB/No. Tracking': ['AWB/No. Tracking', 'AWB', 'No. Tracking', 'Tracking Number', 'Resi', 'No Resi'],
            'ID Pesanan': ['ID Pesanan', 'Order ID', 'No. Pesanan', 'Nomor Pesanan'],
            'MSKU': ['MSKU', 'SKU', 'Nama SKU', 'Product SKU', 'Master SKU'],
            'Jumlah': ['Jumlah', 'Qty', 'Quantity', 'QTY'],
            'Catatan Pembeli': ['Catatan Pembeli', 'Buyer Note', 'Buyer Notes', 'Customer Note', 'Note']
        }
        
        for target_col, alternatives in col_mapping.items():
            if target_col not in df.columns:
                for alt in alternatives:
                    if alt in df.columns:
                        df = df.rename(columns={alt: target_col})
                        break
        
        # Required: MSKU dan Jumlah
        required_cols = ['MSKU', 'Jumlah']
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            raise HTTPException(status_code=400, detail=f"Kolom tidak ditemukan: {missing}")
        
        # Pastikan minimal ada salah satu: ID Pesanan atau AWB/No. Tracking
        has_id_pesanan = 'ID Pesanan' in df.columns
        has_awb = 'AWB/No. Tracking' in df.columns
        has_catatan = 'Catatan Pembeli' in df.columns
        if not has_id_pesanan and not has_awb:
            raise HTTPException(status_code=400, detail="Kolom 'ID Pesanan' atau 'AWB/No. Tracking' tidak ditemukan")

        if has_id_pesanan:
            df['ID Pesanan'] = df['ID Pesanan'].replace('', float('nan')).ffill()
        if has_awb:
            # FIX: Only ffill AWB within rows that share the same ID Pesanan.
            # This prevents AWB from one order "leaking" into the next order whose AWB is empty.
            df['AWB/No. Tracking'] = df['AWB/No. Tracking'].replace('', float('nan'))
            if has_id_pesanan:
                df['AWB/No. Tracking'] = df.groupby('ID Pesanan', sort=False)['AWB/No. Tracking'].transform(lambda x: x.ffill())
            else:
                df['AWB/No. Tracking'] = df['AWB/No. Tracking'].ffill()

        # 2. Buat mapping - PRIORITAS: ID Pesanan dulu, jika kosong maka AWB
        awb_to_items = {}
        excel_awbs = set()  # Semua keys yang bisa digunakan untuk matching (ID + AWB)
        excel_primary_ids = set()  # HANYA identifier utama per order (untuk statistik)
        excel_awbs_raw_set = set() # Untuk menyimpan semua AWB raw yang ada di excel
        id_to_awb_mapping = {}  # Mapping ID Pesanan -> AWB/No. Tracking
        awb_to_catatan = {}  # Mapping ID/AWB -> Catatan Pembeli (untuk JX)
        
        for _, row in df.iterrows():
            identifier = None
            awb_value = None
            
            # Ambil AWB/No. Tracking value (untuk disimpan ke Supabase)
            if has_awb:
                awb_raw = row.get('AWB/No. Tracking', '')
                awb_norm = normalize_awb(awb_raw)
                if awb_norm and awb_norm not in ['NAN', 'NONE', 'NULL', '']:
                    awb_value = awb_norm
                    excel_awbs_raw_set.add(awb_value)
            
            # Ambil Catatan Pembeli (untuk JX labels)
            catatan_pembeli = ''
            if has_catatan:
                catatan_raw = row.get('Catatan Pembeli', '')
                if pd.notna(catatan_raw) and str(catatan_raw).strip() not in ['nan', 'NaN', 'NAN', '']:
                    catatan_pembeli = str(catatan_raw).strip()
            
            # Cek apakah ini Blibli
            is_blibli = awb_value and awb_value.startswith('BLI')
            
            # Prioritas 1: ID Pesanan (untuk validasi/matching) - KECUALI BLIBLI
            if has_id_pesanan and not is_blibli:
                id_pesanan = row.get('ID Pesanan', '')
                id_pesanan_norm = normalize_awb(id_pesanan)
                if id_pesanan_norm and id_pesanan_norm not in ['NAN', 'NONE', 'NULL', '']:
                    identifier = id_pesanan_norm
                    # Simpan mapping ID Pesanan -> AWB
                    if awb_value:
                        id_to_awb_mapping[identifier] = awb_value
                    # Simpan Catatan Pembeli untuk semua label
                    if catatan_pembeli:
                        awb_to_catatan[identifier] = catatan_pembeli
            
            # Prioritas 2: AWB/No. Tracking (jika ID Pesanan kosong)
            if not identifier and awb_value:
                identifier = awb_value
                # Simpan Catatan Pembeli untuk semua label
                if catatan_pembeli:
                    awb_to_catatan[identifier] = catatan_pembeli
            
            # === PENTING: Tambahkan KEDUA identifier ke excel_awbs ===
            # Ini memungkinkan matching dengan ID Pesanan ATAU AWB
            items_data = {
                'msku': str(row['MSKU']),
                'jumlah': int(row['Jumlah']) if pd.notna(row['Jumlah']) else 1
            }
            
            if identifier:
                excel_awbs.add(identifier)
                excel_primary_ids.add(identifier)  # Tambahkan ke primary IDs untuk statistik
                if identifier not in awb_to_items:
                    awb_to_items[identifier] = []
                awb_to_items[identifier].append(items_data.copy())
            
            # JUGA tambahkan AWB secara terpisah jika ada dan berbeda dari identifier
            # Ini memungkinkan matching via AWB meski identifier utama adalah ID Pesanan
            # CATATAN: AWB ekstra ini TIDAK dimasukkan ke excel_primary_ids (hanya untuk matching)
            if awb_value and awb_value != identifier:
                excel_awbs.add(awb_value)
                if awb_value not in awb_to_items:
                    awb_to_items[awb_value] = []
                awb_to_items[awb_value].append(items_data.copy())
                if catatan_pembeli:
                    awb_to_catatan[awb_value] = catatan_pembeli
        
        # 2c. Agregasi items per AWB: gabungkan baris dengan MSKU sama, jumlahkan qty
        #     Ini menangani format export Lazada/Ginee yang menghasilkan 1 baris per unit
        #     (misal: SHARPENER qty 24 → 24 baris qty=1 → digabung jadi 1 baris qty=24)
        print(f"\n[AGGREGATE] ===== STARTING MSKU AGGREGATION =====")
        print(f"[AGGREGATE] Total AWB keys to process: {len(awb_to_items)}")
        for _awb_key in list(awb_to_items.keys()):
            before_count = len(awb_to_items[_awb_key])
            merged = {}
            for _item in awb_to_items[_awb_key]:
                _key = _item['msku'].strip()
                if _key in merged:
                    merged[_key]['jumlah'] += _item['jumlah']
                else:
                    merged[_key] = {'msku': _item['msku'].strip(), 'jumlah': _item['jumlah']}
            after_count = len(merged)
            if before_count != after_count:
                print(f"[AGGREGATE] AWB {_awb_key}: {before_count} rows -> {after_count} unique MSKU (MERGED)")
                for mk, mv in merged.items():
                    print(f"  -> {mk}: qty={mv['jumlah']}")
            awb_to_items[_awb_key] = list(merged.values())
        print(f"[AGGREGATE] ===== AGGREGATION COMPLETE =====\n")
        
        # 2b. Fetch format mode SEBELUM sort agar bisa sort by Rak & ID
        try:
            feature_res = await supabase_fetch("GET", "toolkit_feature_locks")
            is_extended = any(f.get('feature_key') == 'label_extended_format' and f.get('is_locked') for f in feature_res)
            is_sort_rak_msku = any(f.get('feature_key') == 'label_sort_rak_msku' and f.get('is_locked') for f in feature_res)
        except:
            is_extended = False
            is_sort_rak_msku = False

        rak_map = {}
        if is_extended:
            try:
                mappings = await get_sku_mappings()
                rak_map = {m['sku'].strip().upper(): {"rak": m.get('rak', ''), "id": m.get('id', '')} for m in mappings}
            except:
                pass

        # Fetch label table config (ukuran kolom, font, border)
        try:
            label_cfg = await get_label_config()
        except:
            label_cfg = {}

        try:
            bottom_priorities_res = await supabase_fetch("GET", "label_bottom_priorities")
            bottom_priorities_std = [item['keyword'].strip().upper() for item in bottom_priorities_res if item['format_type'] == 'standar']
            bottom_priorities_rak = [item['keyword'].strip().upper() for item in bottom_priorities_res if item['format_type'] == 'rak_id']
        except:
            bottom_priorities_std = []
            bottom_priorities_rak = []

        def rak_id_sort_key(item):
            sku_upper = item['msku'].strip().upper()
            rak_info  = rak_map.get(sku_upper, {"rak": "", "id": ""})
            rak_val   = rak_info.get('rak', '')
            id_val    = rak_info.get('id', '')  # custom_id, sudah berisi full path: "12-EK-06-22"

            # Gunakan id_val (custom_id) langsung sebagai combined karena sudah berisi
            # lorong+rak+id secara lengkap, misal "12-EK-06-22" atau "2-O-01-03".
            # Jangan gabungkan dengan rak_val (akan duplikat: "EK-12-EK-06-22").
            combined = id_val if id_val else rak_val

            parts = combined.split('-') if combined else []
            zone  = parts[0] if parts else ''
            rest  = parts[1:] if len(parts) > 1 else []

            num_rest = []
            for p in rest:
                try:
                    num_rest.append((0, int(p)))
                except ValueError:
                    num_rest.append((1, p.upper()))

            # Lorong numerik di-sort numerik (2 < 13 < 14), huruf di-sort alfabet (B, Z)
            try:
                zone_val = (0, int(zone))
            except ValueError:
                zone_val = (1, zone.upper())

            priority = 1 if not combined else 0
            return (priority, zone_val, num_rest, sku_upper)

        for awb in awb_to_items:
            if is_extended:
                if is_sort_rak_msku:
                    rak_items = []
                    no_rak_items = []
                    for item in awb_to_items[awb]:
                        sku_upper = item['msku'].strip().upper()
                        rak_info = rak_map.get(sku_upper, {"rak": "", "id": ""})
                        rak_val = rak_info.get('rak', '')
                        id_val = rak_info.get('id', '')
                        combined = id_val if id_val else (rak_val or "")
                        
                        if combined:
                            rak_items.append(item)
                        else:
                            no_rak_items.append(item)
                    
                    rak_items.sort(key=rak_id_sort_key)
                    no_rak_items.sort(key=lambda x: x['msku'].strip().upper())
                    
                    result = []
                    rak_idx = 0
                    for no_rak in no_rak_items:
                        no_rak_msku = no_rak['msku'].strip().upper()
                        while rak_idx < len(rak_items):
                            if rak_items[rak_idx]['msku'].strip().upper() > no_rak_msku:
                                break
                            result.append(rak_items[rak_idx])
                            rak_idx += 1
                        result.append(no_rak)
                    
                    while rak_idx < len(rak_items):
                        result.append(rak_items[rak_idx])
                        rak_idx += 1
                        
                    awb_to_items[awb] = result
                else:
                    awb_to_items[awb].sort(key=rak_id_sort_key)
            else:
                awb_to_items[awb].sort(key=lambda x: x['msku'])

            # Extract bottom priority items
            if is_extended:
                normal_items = []
                bottom_items = []
                for item in awb_to_items[awb]:
                    sku_up = item['msku'].strip().upper()
                    r_info = rak_map.get(sku_up, {})
                    # 'id' = full custom_id (e.g. Z-IE-04-12), 'rak' = area prefix
                    c = str(r_info.get('id') or '').strip().upper()
                    c_rak = str(r_info.get('rak') or '').strip().upper()
                    is_bottom = any(
                        c == k or c.startswith(k) or c_rak == k or c_rak.startswith(k)
                        for k in bottom_priorities_rak
                    )
                    if is_bottom:
                        bottom_items.append(item)
                    else:
                        normal_items.append(item)
                # Sort bottom items A-Z by their location string (B before Z)
                def get_loc2(item):
                    r = rak_map.get(item['msku'].strip().upper(), {})
                    return str(r.get('id') or r.get('rak') or item['msku']).strip().upper()
                bottom_items.sort(key=get_loc2)
                awb_to_items[awb] = normal_items + bottom_items
            else:
                normal_items = []
                bottom_items = []
                for item in awb_to_items[awb]:
                    sku_up = item['msku'].strip().upper()
                    is_bottom = any(sku_up == k or sku_up.startswith(k) for k in bottom_priorities_std)
                    if is_bottom:
                        bottom_items.append(item)
                    else:
                        normal_items.append(item)
                bottom_items.sort(key=lambda x: x['msku'].strip().upper())
                awb_to_items[awb] = normal_items + bottom_items
        
        # Create reverse mapping for duplicate checking
        # awb_to_id_mapping: AWB -> ID Pesanan
        awb_to_id_mapping = {v: k for k, v in id_to_awb_mapping.items()}
        
        print(f"[DEBUG] Catatan Pembeli untuk JX: {len(awb_to_catatan)} entries")
        print(f"[DEBUG] Total AWB dari Excel: {len(excel_awbs)}")
        if excel_awbs:
            sample_awbs = list(excel_awbs)[:10]
            print(f"[DEBUG] Sample AWB dari Excel: {sample_awbs}")
            

        # 3. Baca PDF dan track AWBs
        all_pages = []
        pdf_awbs = set()  # AWB utama dari setiap page PDF
        pdf_filenames = []
        combined_pdf_content = io.BytesIO()  # Untuk backup: gabungan semua PDF asli
        combined_pdf_doc = fitz.open()  # Document untuk menggabungkan semua PDF asli
        
        for pdf in pdf_files:
            pdf_content = await pdf.read()
            pdf_filenames.append(pdf.filename)
            doc = fitz.open(stream=pdf_content, filetype="pdf")
            
            # Tambahkan ke combined PDF untuk backup
            combined_pdf_doc.insert_pdf(doc)
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                primary_identifier = None  # FIX: Reset setiap halaman baru agar tidak bocor dari halaman sebelumnya
                
                # FIX: Scan 70% bagian atas saja
                clip_rect = fitz.Rect(0, 0, page.rect.width, page.rect.height * 0.70)
                text = page.get_text("text", clip=clip_rect)
                full_text = page.get_text("text")  # Full text untuk junk/continuation check
                # Ekstrak kandidat dari FULL TEXT (bukan clip 70%) agar
                # barcode bawah Lazada (Order ID / Nomor Order) ikut ter-ekstrak.
                # Barcode Order ID Lazada ada di bawah separator --- di luar clip 70%.
                candidates = extract_all_awb_candidates(full_text)
                order_ids = extract_order_ids(full_text)  # Juga cari Order ID
                
                # Cek apakah halaman ini adalah junk page atau lanjutan kita
                page_is_junk = is_junk_page(page)
                page_is_continuation = 'LANJUTAN AWB:' in full_text.upper()
                
                # Gabungkan kandidat dengan PRIORITAS: alphanumerik (huruf+angka) lebih dulu, numerik murni terakhir
                all_raw = candidates + order_ids
                alpha_num_cands = list(dict.fromkeys([c for c in all_raw if any(ch.isalpha() for ch in c)]))
                pure_num_cands  = list(dict.fromkeys([c for c in all_raw if not any(ch.isalpha() for ch in c)]))
                priority_candidates = alpha_num_cands + pure_num_cands
                
                # Cek jika halaman ini adalah PACKING LIST / BATCH
                is_packing_list = 'PACKING LIST / BATCH' in full_text.upper() or 'LIST ID PESANAN' in full_text.upper()
                
                # Coba match dengan excel dulu (kecuali jika ini packing list)
                if is_packing_list:
                    matched_awb = None
                else:
                    matched_awb = find_matching_awb(priority_candidates, excel_awbs, page_num + 1)
                
                # DEBUG: Tulis info ekstraksi ke debug.log
                with open("debug_text.log", "a", encoding="utf-8") as dbg:
                    dbg.write(f"\n--- Page {page_num + 1} ---\n")
                    dbg.write(f"Is Junk: {page_is_junk}, Is Continuation: {page_is_continuation}\n")
                    dbg.write(f"AWB Candidates: {candidates[:10]}\n")
                    dbg.write(f"Order IDs extracted: {order_ids[:10]}\n")
                    dbg.write(f"Matched AWB: {matched_awb}\n")
                    dbg.write(f"Raw text (first 300 chars):\n{text[:300]}\n")
                
                if matched_awb:
                    primary_identifier = matched_awb
                else:
                    # Jika tidak match, ambil AWB/ID terbaik dari page
                    for c in candidates:
                        c_upper = c.upper()
                        known_prefixes = [
                            'SPXID', 'SPX',                              # Shopee
                            'JT', 'JX', 'JD',                            # J&T
                            'NVID', 'NV', 'NJVTT', 'NJV', 'NINJA',        # Ninja
                            'TKP', 'ANT', 'GTL',                          # Tokopedia/AnterAja/GoTo
                            'SAP', 'LEX', 'LAZ', 'ID',                    # SAP/Lazada
                            'JNE', 'LION', 'LP',                          # JNE/Lion
                            'KERRY', 'KE',                                # Kerry
                            'WHN', 'RPX', 'TIKI',                         # Wahana/RPX/Tiki
                            'BLIJC', 'BLI',                               # Blibli
                            'GB', 'GRAB',                                 # Grab
                            'POS',                                        # Pos
                        ]
                        if any(c_upper.startswith(p) for p in known_prefixes):
                            primary_identifier = c
                            break
                        if c_upper.startswith('00') and len(c) >= 12:
                            primary_identifier = c
                            break
                    
                    if not primary_identifier and candidates:
                        for c in candidates:
                            if len(c) >= 10:
                                primary_identifier = c
                                break
                    
                    # Fallback ke Order ID jika tidak ada AWB
                    if not primary_identifier and order_ids:
                        primary_identifier = order_ids[0]
                
                if primary_identifier:
                    pdf_awbs.add(primary_identifier)
                
                clear_start_y = find_clear_start_position(page)
                
                all_pages.append({
                    'page_num': page_num,
                    'page': page,
                    'doc': doc,
                    'width': page.rect.width,
                    'height': page.rect.height,
                    'awb': matched_awb,
                    'primary_pdf_awb': primary_identifier,
                    'is_junk': page_is_junk,
                    'is_continuation': page_is_continuation,
                    'clear_start_y': clear_start_y
                })

        # 4. Calculate match statistics
        matched_awbs = set()
        for page_info in all_pages:
            if page_info['awb']:
                matched_awbs.add(page_info['awb'])
        
        # Debug log to file
        with open("debug.log", "a") as f:
            f.write(f"\n=== MATCHING SUMMARY (with-stats) ===\n")
            f.write(f"Total PDF pages: {len(all_pages)}\n")
            f.write(f"Pages with matched AWB: {len([p for p in all_pages if p['awb']])}\n")
            f.write(f"Pages without matched AWB: {len([p for p in all_pages if not p['awb']])}\n")
            f.write(f"Excel unique identifiers: {len(excel_awbs)}\n")
            f.write(f"Sample Excel AWBs: {list(excel_awbs)[:10]}\n")
            f.write(f"Sample PDF AWBs: {list(pdf_awbs)[:10]}\n")
            f.write(f"Matched AWBs: {len(matched_awbs)}\n")
        
        # Order di Excel yang tidak ditemukan match di PDF
        # Gunakan canonical IDs: resolve semua matched_awbs ke canonical ID Pesanan
        matched_canonical_ids = set()
        for awb_key in matched_awbs:
            # Jika ini adalah AWB (bukan ID Pesanan), resolve ke ID Pesanan-nya
            if awb_key in awb_to_id_mapping:
                matched_canonical_ids.add(awb_to_id_mapping[awb_key])
            else:
                matched_canonical_ids.add(awb_key)
        
        # Bandingkan hanya primary IDs (bukan semua AWB + ID)
        unmatched_excel = excel_primary_ids - matched_canonical_ids
        
        # AWB di PDF yang tidak ada di Excel
        # Hanya ambil dari pages yang TIDAK matched, punya primary identifier,
        # DAN bukan junk page atau continuation page (lanjutan pretelan)
        unmatched_pdf = set()
        for page_info in all_pages:
            if not page_info['awb']:  # Page tidak matched dengan Excel
                # Skip junk pages dan halaman lanjutan (pretelan) kita sendiri
                if page_info.get('is_junk', False) or page_info.get('is_continuation', False):
                    print(f"[STATS] Page {page_info['page_num'] + 1} skipped from PDF Only: junk={page_info.get('is_junk')}, continuation={page_info.get('is_continuation')}")
                    continue
                primary = page_info.get('primary_pdf_awb')
                if primary:
                    unmatched_pdf.add(primary)
                    print(f"[STATS] Page {page_info['page_num'] + 1} counted as PDF Only: primary='{primary}'")
        
        master_doc = fitz.open()
        matched_count = 0
        label_number = 0
        processed_awbs = {}  # Track AWB -> last_page_num yang sudah diproses
        duplicate_count = 0
        duplicate_awbs = []  # Track detail AWB yang duplikat
        continuation_pages = []  # Track halaman lanjutan (pretelan)
        unique_matched_ids = set() # Track unique Order IDs for accurate matched count
        
        # RECAP DATA for priority_kembar summary
        recap_dict = {} # MSKU -> total_qty

        # 5. is_extended dan rak_map sudah di-fetch sebelum sort di atas.
        
        # --- CUSTOM SORT: Urutan label mengikuti urutan baris di Excel ---
        # Buat mapping: AWB/ID → urutan kemunculan pertama di Excel (index)
        excel_order = {}
        excel_order_idx = 0
        for _, row in df.iterrows():
            identifier = None
            if has_id_pesanan:
                id_val = normalize_awb(row.get('ID Pesanan', ''))
                if id_val and id_val not in ['NAN', 'NONE', 'NULL', '']:
                    identifier = id_val
            if not identifier and has_awb:
                awb_val = normalize_awb(row.get('AWB/No. Tracking', ''))
                if awb_val and awb_val not in ['NAN', 'NONE', 'NULL', '']:
                    identifier = awb_val
            if identifier and identifier not in excel_order:
                excel_order[identifier] = excel_order_idx
                excel_order_idx += 1

        import re
        def natural_sort_key(s):
            normalized_s = str(s).strip().upper()
            if normalized_s.startswith('BT-'):
                normalized_s = normalized_s.replace('BT-', 'BT1-', 1)
            elif normalized_s.startswith('T-'):
                normalized_s = normalized_s.replace('T-', 'T2-', 1)
            return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', normalized_s)]

        def get_excel_sort_key(page_info):
            awb = page_info.get('awb')
            
            sku_count_priority = 0
            rak_val = "ZZZZ"
            id_val = "ZZZZ"
            msku_val = "ZZZZ"
            
            excel_idx = 999999
            
            # 1. Cari urutan Excel
            if awb and awb in excel_order:
                excel_idx = excel_order[awb]
            elif awb and awb in awb_to_id_mapping and awb_to_id_mapping[awb] in excel_order:
                excel_idx = excel_order[awb_to_id_mapping[awb]]
            elif awb and awb in id_to_awb_mapping and id_to_awb_mapping[awb] in excel_order:
                excel_idx = excel_order[id_to_awb_mapping[awb]]
                
            # 2. Hitung jumlah SKU dan ambil data Rak jika = 1
            if sort_by_sku_count and awb:
                canonical_for_count = awb
                if awb in awb_to_id_mapping:
                    canonical_for_count = awb_to_id_mapping[awb]
                    
                items = []
                if canonical_for_count in awb_to_items:
                    items = awb_to_items[canonical_for_count]
                elif awb in awb_to_items:
                    items = awb_to_items[awb]
                    
                sku_count_priority = len(items)
                
                # JIKA SKU SATUAN, Ambil Rak, ID, MSKU
                if sku_count_priority == 1:
                    item = items[0]
                    msku_val = str(item.get('msku', '')).strip().upper()
                    
                    # rak_map sudah ada di memory
                    info = rak_map.get(msku_val, {})
                    r_v = str(info.get('rak', '')).strip().upper()
                    i_v = str(info.get('id', '')).strip().upper()
                    
                    if r_v and r_v != '-': rak_val = r_v
                    if i_v and i_v != '-': id_val = i_v

            if not sort_by_sku_count:
                # Jika bukan dari Upload 2, abaikan rak dan jumlah SKU
                return (0, 0, natural_sort_key("ZZZZ"), natural_sort_key("ZZZZ"), natural_sort_key("ZZZZ"), excel_idx)

            if not awb:
                return (999, 2, natural_sort_key("ZZZZ"), natural_sort_key("ZZZZ"), natural_sort_key("ZZZZ"), page_info.get('page_num', 9999))
                
            # Return tuple panjang agar tidak terjadi TypeError saat perbandingan
            # Menggunakan natural_sort_key agar "3" berada di atas "17"
            return (sku_count_priority, 0, natural_sort_key(rak_val), natural_sort_key(id_val), natural_sort_key(msku_val), excel_idx)

        all_pages.sort(key=get_excel_sort_key)
        
        if sort_by_sku_count:
            print(f"[SORT] SKU-Count sorting applied (Upload 2). {len(excel_order)} unique identifiers mapped.")
        else:
            print(f"[SORT] Excel-order sorting applied. {len(excel_order)} unique identifiers mapped from Excel.")

        # --- PRIORITY KEMBAR SORTING (Upload Kembar Only) ---
        # Catatan: Jika priority_kembar=True, sorting di bawah ini akan OVERRIDE urutan Excel di atas
        if priority_kembar:
            print("[SORT] Applying Priority Kembar sorting...")
            # Fingerprints based on MSKU sets (must have >= 2 unique MSKUs)
            page_fingerprints = {}  # Page Index -> tuple
            fp_counts = {}         # tuple -> count
            
            for idx, page_info in enumerate(all_pages):
                awb = page_info['awb']
                if awb and awb in awb_to_items:
                    items = awb_to_items[awb]
                    unique_mskus = sorted(list(set([str(i['msku']).strip().upper() for i in items])))
                    if len(unique_mskus) >= 1:
                        fp = tuple(unique_mskus)
                        page_fingerprints[idx] = fp
                        fp_counts[fp] = fp_counts.get(fp, 0) + 1
            
            # Add metadata for sorting
            for idx, page_info in enumerate(all_pages):
                fp = page_fingerprints.get(idx)
                is_twin = fp and fp_counts.get(fp, 0) >= 2
                page_info['is_kembar_priority'] = is_twin
                page_info['kembar_fingerprint'] = fp
                page_info['original_idx'] = idx
            
            # Sort: Twin Priority (0) -> Fingerprint -> Original Index
            all_pages.sort(key=lambda x: (
                0 if x.get('is_kembar_priority') else 1,
                x.get('kembar_fingerprint') or (),
                x.get('original_idx', 0)
            ))
            print(f"[SORT] Sorting complete. {len([p for p in all_pages if p.get('is_kembar_priority')])} pages prioritized.")

        # --- PRE-SCAN MSKU FREQUENCY (For Overlap Filtering) ---
        twin_mskus = set()
        if priority_kembar:
            msku_to_ids = {} # MSKU -> set of Canonical IDs
            for page_info in all_pages:
                awb = page_info['awb']
                # Filter ketat: Hanya hitung dari halaman label asli yang valid
                if not awb or awb not in awb_to_items:
                    continue
                if page_info.get('is_junk') or page_info.get('is_continuation'):
                    continue
                
                # Gunakan Canonical ID (Order ID) agar duplikat AWB tetap dihitung 1 order
                canonical_id = awb
                if awb in awb_to_id_mapping:
                    canonical_id = awb_to_id_mapping[awb]
                
                for item in awb_to_items[awb]:
                    m = str(item['msku']).strip().upper()
                    if m not in msku_to_ids: msku_to_ids[m] = set()
                    msku_to_ids[m].add(canonical_id)
            
            # Hanya rekap barang yang muncul di minimal 2 ORDER BERBEDA di PDF
            twin_mskus = {m for m, ids in msku_to_ids.items() if len(ids) >= 2}
            print(f"[RECAP] Found {len(twin_mskus)} shared MSKUs across different orders.")

        # 6. Proses setiap halaman (sama seperti endpoint biasa)
        for page_info in all_pages:
            page_awb = page_info['awb']
            page_num = page_info['page_num']
            W_pts = page_info['width']
            H_pts = page_info['height']
            clear_start_y = page_info['clear_start_y']
            table_start_y = clear_start_y
            
            if not page_awb or page_awb not in awb_to_items:
                if page_info.get('is_junk', False) or page_info.get('is_continuation', False):
                    print(f"[DEBUG] Page {page_num + 1} is junk/continuation. Skipping raw insert.")
                    continue
                # Page tanpa AWB match (misal cover PACKING LIST) - insert raw
                print(f"[DEBUG] Page {page_num + 1} not matched. Inserting raw page.")
                temp_doc = fitz.open()
                temp_doc.insert_pdf(page_info['doc'], from_page=page_num, to_page=page_num)
                master_doc.insert_pdf(temp_doc)
                temp_doc.close()
                continue
            
            # CEK DUPLIKAT vs PRETELAN (halaman lanjutan)
            status = "NEW"
            prev_page = -1
            
            # Debug Trace
            canonical_trace = page_awb
            if page_awb in awb_to_id_mapping:
                canonical_trace = awb_to_id_mapping[page_awb]
            
            print(f"[DEBUG] Processing Page {page_num + 1}: Raw='{page_awb}', Canonical='{canonical_trace}'")
            print(f"        processed_keys: {list(processed_awbs.keys())[:5]}... (Total: {len(processed_awbs)})")
            
            if page_awb in processed_awbs:
                last_page = processed_awbs[page_awb]
                prev_page = last_page
                awb_display = id_to_awb_mapping.get(page_awb, page_awb)
                
                print(f"        -> HIT processed_awbs via Raw ID")

                # Jika halaman berurutan (selisih <= 2), ini adalah PRETELAN (halaman lanjutan)
                if abs(page_num - last_page) <= 2:
                    status = "CONT"
                    print(f"  [CONT] PRETELAN (lanjutan): ID={page_awb}, AWB={awb_display}, Page={page_num + 1} (prev: {last_page + 1})")
                    continuation_pages.append({
                        'id_pesanan': page_awb,
                        'awb': awb_display,
                        'pdf_page': page_num + 1,
                        'first_page': last_page + 1
                    })
                    # Skip halaman lanjutan (pretelan) saja
                    continue
                # Gap > 2: JANGAN skip!
                # Bisa saja AWB yang sama muncul via text extraction di halaman lain (false positive)
                # Label tetap diproses, dedup via canonical ID di bawah
                print(f"  [INFO] AWB '{page_awb}' muncul lagi di page {page_num + 1} (gap > 2). Diproses, cek canonical dedup.")
            
            # Simpan page_awb sebagai sudah diproses
            processed_awbs[page_awb] = page_num
            
            # TRACK unique canonical Order ID
            # Jika canonical_id sudah ada = order ini sudah dirender sebelumnya (pretelan/duplikat)
            # → SKIP rendering label baru agar label_number tidak loncat
            canonical_id = page_awb
            if page_awb in awb_to_id_mapping:
                canonical_id = awb_to_id_mapping[page_awb]
            
            if canonical_id in unique_matched_ids:
                # Order sudah diproses sebelumnya, halaman ini adalah duplikat pretelan (gap > 2)
                print(f"  [SKIP-DUP] Canonical ID '{canonical_id}' sudah dirender. Halaman {page_num + 1} dilewati (duplikat pretelan gap>2).")
                continue
            
            unique_matched_ids.add(canonical_id)
            matched_count = len(unique_matched_ids)
            print(f"        -> Unique Track: Canonical='{canonical_id}', Count={matched_count}")
            
            items = awb_to_items[page_awb]
            
            # --- START PRIORITY TOP LOGIC ---
            if is_priority_top_active and len(items) > 1:
                def sort_item_key(item):
                    msku = str(item.get('msku', '')).strip().upper()
                    # 1. Is it a bulky SKU? (0 if yes, 1 if no) -> 0 will sort to TOP
                    bulky_score = 0 if msku in special_skus else 1
                    
                    # 2. Rak & ID
                    info = rak_map.get(msku, {})
                    rak_str = str(info.get('rak', '')).strip().upper()
                    if not rak_str or rak_str == '-':
                        rak_str = 'ZZZZ'
                        
                    id_str = str(info.get('id', '')).strip().upper()
                    if not id_str or id_str == '-':
                        id_str = 'ZZZZ'
                        
                    rak_val = natural_sort_key(rak_str)
                    id_val = natural_sort_key(id_str)
                    
                    # 3. MSKU
                    msku_val = natural_sort_key(msku)
                    return (bulky_score, rak_val, id_val, msku_val)
                    
                items = sorted(items, key=sort_item_key)
            # --- END PRIORITY TOP LOGIC ---
            
            # Add to recap if priority_kembar AND item is shared (twin)
            if priority_kembar:
                for item in items:
                    msku = str(item['msku']).strip().upper()
                    # Hanya rekap barang yang muncul di > 1 label
                    if msku in twin_mskus:
                        qty = int(item['jumlah'])
                        recap_dict[msku] = recap_dict.get(msku, 0) + qty
            
            # Tentukan row height sebenarnya dari config
            if is_extended:
                _ext_max_f = max(
                    float(label_cfg.get('ext_font_rak') or label_cfg.get('ext_font_size', 8)),
                    float(label_cfg.get('ext_font_msku') or label_cfg.get('ext_font_size', 8)),
                    float(label_cfg.get('ext_font_qty') or label_cfg.get('ext_font_size', 8))
                )
                actual_row_height = max(float(label_cfg.get('ext_row_height', 18)), _ext_max_f * 2.25)
            else:
                _std_max_f = max(
                    float(label_cfg.get('std_font_msku') or label_cfg.get('std_font_size', 8)),
                    float(label_cfg.get('std_font_qty') or label_cfg.get('std_font_size', 8))
                )
                actual_row_height = max(float(label_cfg.get('std_row_height', 18)), _std_max_f * 2.25)
                
            is_blibli_page = page_awb.startswith('BLI')
            margin_bottom = (65 if is_blibli_page else 50) if W_pts > 350 else (45 if is_blibli_page else 35)
            available_rows = calculate_available_rows(H_pts, table_start_y, row_height=actual_row_height, margin_bottom=margin_bottom)
            # Pagination dinamis — ITEM-aware untuk extended
            limit_extra = max(1, calculate_available_rows(H_pts, 50, row_height=actual_row_height, margin_bottom=margin_bottom))
            
            available_height_pts = max(0, H_pts - table_start_y - margin_bottom - 2)
            limit_extra_pts = max(0, H_pts - 50 - (margin_bottom - 10) - 2)
            
            scaled_cfg = scale_cfg_for_wide_label(label_cfg, W_pts, is_extended)
            
            # --- Uniform Page 1 Scaling ---
            worst_case_h = 70.0
            scale_f = 1.0
            if available_height_pts < worst_case_h and available_height_pts > 0:
                scale_f = max(0.65, available_height_pts / worst_case_h)
            page1_cfg = shrink_cfg_for_page1(scaled_cfg, scale_f)
            
            chunks = []
            rest = items
            first_page = True
            while rest:
                ah = available_height_pts if first_page else limit_extra_pts
                current_cfg = page1_cfg if first_page else scaled_cfg
                
                next_chunk, rest_after = calc_items_for_rows_by_height(rest, is_extended, rak_map, ah, current_cfg, W_pts, force_first=first_page)
                
                # Enforce max 2 items on the first page for multi-item (resi pretelan)
                if first_page and len(items) > 1:
                    if len(next_chunk) > 2:
                        next_chunk = rest[:2]
                        rest_after = rest[2:]
                
                if not next_chunk:
                    if first_page:
                        # Should not happen because of force_first, but fallback
                        chunks.append({'items': [], 'cfg': current_cfg})
                        first_page = False
                        continue
                    else:
                        next_chunk = rest[:1]
                        rest_after = rest[1:]
                        
                chunks.append({'items': next_chunk, 'cfg': current_cfg})
                rest = rest_after
                first_page = False

            total_pages = len(chunks)
            for i, chunk_data in enumerate(chunks):
                chunk = chunk_data['items']
                current_cfg = chunk_data['cfg']
                is_last_chunk = (i == total_pages - 1)
                
                if i == 0:
                    temp_doc = fitz.open()
                    temp_doc.insert_pdf(page_info['doc'], from_page=page_num, to_page=page_num)
                    page_copy = temp_doc[0]
                    
                    rect_clear = fitz.Rect(0, clear_start_y, max(W_pts, page_copy.rect.width), max(H_pts, page_copy.rect.height))
                    page_copy.add_redact_annot(rect_clear, fill=(1, 1, 1))
                    page_copy.apply_redactions()
                    page_copy.set_cropbox(fitz.Rect(0, 0, W_pts, H_pts))
                    
                    packet = io.BytesIO()
                    can = canvas.Canvas(packet, pagesize=(W_pts, H_pts))
                    
                    if chunk:
                        table_result = generate_table_data(chunk, is_extended, rak_map, current_cfg, picker_name=picker_name)
                        t = create_table(
                            table_result['table_data'],
                            row_heights=table_result['row_heights'],
                            span_cmds=table_result['span_cmds'],
                            label_cfg=current_cfg,
                        )
                        t.wrapOn(can, W_pts, H_pts)
                        total_table_h = sum(table_result['row_heights'])
                        # table_y dari bawah. Kurangi padding atas (2pt) agar sangat maksimal.
                        table_y = H_pts - table_start_y - total_table_h - 2
                        # Batas bawah aman: 20px dari bawah agar tidak timpa No:xxx (y=10)
                        safe_bottom = 35 if is_blibli_page else 20
                        t.drawOn(can, 7, max(safe_bottom, table_y))
                    else:
                        table_y = H_pts - table_start_y - 10
                        if not is_last_chunk:
                            can.saveState()
                            can.setFillColorRGB(0.5, 0.5, 0.5)
                            can.setFont("Helvetica-Oblique", 11)
                            middle_y = (H_pts - clear_start_y) / 2
                            can.drawCentredString(W_pts / 2, middle_y, "Data pesanan berada di halaman berikutnya")
                            can.restoreState()
                    
                    # Customer Message HANYA di chunk terakhir
                    catatan_to_add = None
                    if is_last_chunk:
                        # Catatan Pembeli untuk semua label (tidak hanya JX)
                        catatan = awb_to_catatan.get(page_awb, '')
                        if catatan:
                            # Clean text: remove line breaks, keep unicode
                            catatan_to_add = clean_text_for_pdf(catatan)
                    
                    # Cek ruang untuk Customer Message
                    msg_space_needed = 0
                    msg_lines = []
                    if catatan_to_add:
                        max_chars = 55
                        remaining_text = catatan_to_add
                        while remaining_text:
                            if len(remaining_text) <= max_chars:
                                msg_lines.append(remaining_text)
                                break
                            else:
                                split_at = remaining_text[:max_chars].rfind(' ')
                                if split_at == -1:
                                    split_at = max_chars
                                msg_lines.append(remaining_text[:split_at])
                                remaining_text = remaining_text[split_at:].strip()
                        msg_space_needed = len(msg_lines) * 10 + 15  # 10pt per line + margin
                    
                    # Cek apakah cukup ruang (min 30pt dari bawah)
                    space_available = max(10, table_y) - 30
                    
                    if catatan_to_add and space_available >= msg_space_needed:
                        # Ada ruang - render di halaman ini
                        msg_y = max(10, table_y) - 12
                        can.setFont("Helvetica-Bold", 7)
                        msg_prefix = "Customer Message: "
                        can.drawString(7, msg_y, msg_prefix)
                        prefix_width = can.stringWidth(msg_prefix, "Helvetica-Bold", 7)
                        
                        # Use Unicode font for message content
                        can.setFont(UNICODE_FONT, 7)
                        if msg_lines:
                            can.drawString(7 + prefix_width, msg_y, msg_lines[0])
                            for j, line in enumerate(msg_lines[1:], 1):
                                can.drawString(7, msg_y - (j * 10), line)
                    
                    # Add label number (increment di sini agar berurutan berdasarkan output, bukan input PDF)
                    label_number += 1
                    label_text = f"No:{label_number:03d}"
                    no_font_size = 11 if W_pts > 350 else 7
                    can.setFont("Helvetica-Bold", no_font_size)
                    text_width = can.stringWidth(label_text, "Helvetica-Bold", no_font_size)
                    x_pos = W_pts - text_width - 10
                    y_pos = 10
                    can.drawString(x_pos, y_pos, label_text)
                    
                    # ===== INDIKATOR HALAMAN (BERSAMBUNG / AKHIR) =====
                    if total_pages > 1:
                        indicator_text = ">>> BERSAMBUNG KE HALAMAN BERIKUTNYA >>>" if not is_last_chunk else "--- AKHIR DARI PESANAN INI ---"
                        ind_font_size = 12 if W_pts > 350 else 8
                        can.setFont("Helvetica-Bold", ind_font_size)
                        ind_width = can.stringWidth(indicator_text, "Helvetica-Bold", ind_font_size)
                        can.drawString((W_pts - ind_width) / 2, 10, indicator_text)
                    
                    can.save()
                    
                    packet.seek(0)
                    overlay_pdf = fitz.open("pdf", packet.read())
                    cleared_area = fitz.Rect(0, clear_start_y, W_pts, H_pts)
                    page_copy.show_pdf_page(page_copy.rect, overlay_pdf, 0)
                    master_doc.insert_pdf(temp_doc, from_page=0, to_page=0)
                    overlay_pdf.close()
                    temp_doc.close()
                    
                    # Jika tidak cukup ruang, buat halaman baru untuk Customer Message
                    if catatan_to_add and space_available < msg_space_needed:
                        packet2 = io.BytesIO()
                        can2 = canvas.Canvas(packet2, pagesize=(W_pts, H_pts))
                        
                        can2.setFont("Helvetica-Bold", 10)
                        can2.drawString(10, H_pts - 25, f"Lanjutan AWB: {page_awb}")
                        
                        msg_y2 = H_pts - 50
                        can2.setFont("Helvetica-Bold", 8)
                        msg_prefix = "Customer Message: "
                        can2.drawString(10, msg_y2, msg_prefix)
                        prefix_width = can2.stringWidth(msg_prefix, "Helvetica-Bold", 8)
                        
                        can2.setFont(UNICODE_FONT, 8)
                        if msg_lines:
                            can2.drawString(10 + prefix_width, msg_y2, msg_lines[0])
                            for j, line in enumerate(msg_lines[1:], 1):
                                can2.drawString(10, msg_y2 - (j * 12), line)
                        
                        can2.save()
                        packet2.seek(0)
                        msg_page_doc = fitz.open("pdf", packet2.read())
                        master_doc.insert_pdf(msg_page_doc)
                        msg_page_doc.close()
                        
                else:
                    # Halaman lanjutan
                    packet = io.BytesIO()
                    can = canvas.Canvas(packet, pagesize=(W_pts, H_pts))
                    
                    can.setFont("Helvetica-Bold", 10)
                    hal_info = f" (Hal. {i+1} dari {total_pages})"
                    can.drawString(10, H_pts - 25, f"Lanjutan AWB: {page_awb}{hal_info}")
                    
                    table_result = generate_table_data(chunk, is_extended, rak_map, scaled_cfg, picker_name=picker_name)
                    t = create_table(
                        table_result['table_data'],
                        row_heights=table_result['row_heights'],
                        span_cmds=table_result['span_cmds'],
                        label_cfg=scaled_cfg,
                    )
                    t.wrapOn(can, W_pts, H_pts)
                    total_table_h = sum(table_result['row_heights'])
                    t_height = total_table_h
                    table_y_cont = H_pts - t_height - 45
                    t.drawOn(can, 7, table_y_cont)
                    
                    # Customer Message di halaman lanjutan TERAKHIR
                    if is_last_chunk:
                        # Catatan Pembeli untuk semua label (tidak hanya JX)
                        catatan = awb_to_catatan.get(page_awb, '')
                        if catatan:
                                # Clean text
                                catatan_clean = clean_text_for_pdf(catatan)
                                max_chars = 55
                                msg_lines_cont = []
                                remaining_text = catatan_clean
                                while remaining_text:
                                    if len(remaining_text) <= max_chars:
                                        msg_lines_cont.append(remaining_text)
                                        break
                                    else:
                                        split_at = remaining_text[:max_chars].rfind(' ')
                                        if split_at == -1:
                                            split_at = max_chars
                                        msg_lines_cont.append(remaining_text[:split_at])
                                        remaining_text = remaining_text[split_at:].strip()
                                
                                msg_y_cont = table_y_cont - 15
                                if msg_y_cont > 30:  # Cukup ruang
                                    can.setFont("Helvetica-Bold", 7)
                                    msg_prefix = "Customer Message: "
                                    can.drawString(7, msg_y_cont, msg_prefix)
                                    prefix_width = can.stringWidth(msg_prefix, "Helvetica-Bold", 7)
                                    
                                    can.setFont(UNICODE_FONT, 7)
                                    if msg_lines_cont:
                                        can.drawString(7 + prefix_width, msg_y_cont, msg_lines_cont[0])
                                        for j, line in enumerate(msg_lines_cont[1:], 1):
                                            can.drawString(7, msg_y_cont - (j * 10), line)
                    
                    # ===== INDIKATOR HALAMAN (BERSAMBUNG / AKHIR) =====
                    if total_pages > 1:
                        indicator_text = ">>> BERSAMBUNG KE HALAMAN BERIKUTNYA >>>" if not is_last_chunk else "--- AKHIR DARI PESANAN INI ---"
                        ind_font_size = 12 if W_pts > 350 else 8
                        can.setFont("Helvetica-Bold", ind_font_size)
                        ind_width = can.stringWidth(indicator_text, "Helvetica-Bold", ind_font_size)
                        can.drawString((W_pts - ind_width) / 2, 10, indicator_text)
                    
                    can.save()
                    
                    packet.seek(0)
                    new_page_doc = fitz.open("pdf", packet.read())
                    master_doc.insert_pdf(new_page_doc)
                    new_page_doc.close()

        if len(master_doc) == 0:
            # Revert to strict error as requested
            print("[PROCESS] No matching labels found. Aborting.")
            raise HTTPException(
                status_code=400, 
                detail="DATA MISMATCH: Tidak ada label di PDF yang cocok dengan data Excel. Pastikan Anda mengunggah file yang benar."
            )

        # Convert PDF to base64
        
        # PREPEND RECAP PAGE if priority_kembar
        if priority_kembar and recap_dict:
            try:
                recap_pdf = generate_recap_pdf(recap_dict, rak_map, label_cfg)
                if len(recap_pdf) > 0:
                    final_master = fitz.open()
                    final_master.insert_pdf(recap_pdf)
                    final_master.insert_pdf(master_doc)
                    master_doc = final_master
                recap_pdf.close()
            except Exception as e_recap:
                print(f"[RECAP] Error generating recap page: {e_recap}")

        # APPEND SUMMARY PAGE for Upload 2
        if sort_by_sku_count:
            try:
                # Cek apakah fitur packing list dimatikan di admin
                features = await supabase_fetch("GET", "toolkit_feature_locks?select=feature_key,is_locked")
                is_packing_list_locked = False
                if isinstance(features, list):
                    for f in features:
                        if f.get("feature_key") == "packing-list-upload-2":
                            is_packing_list_locked = f.get("is_locked", False)
                            break
                
                if not is_packing_list_locked:
                    final_ids = []
                    seen_ids = set()
                    for page_info in all_pages:
                        if page_info.get('is_continuation'):
                            continue
                        awb = page_info.get('awb')
                        if awb:
                            canonical_id = awb
                            if awb in awb_to_id_mapping:
                                canonical_id = awb_to_id_mapping[awb]
                            if canonical_id not in seen_ids:
                                seen_ids.add(canonical_id)
                                final_ids.append(canonical_id)
                    
                    if final_ids:
                        summary_title = pdf_files[0].filename if pdf_files else "SUMMARY_LABEL.pdf"
                        summary_pdf = generate_summary_page(summary_title, final_ids)
                        if len(summary_pdf) > 0:
                            master_doc.insert_pdf(summary_pdf)
                        summary_pdf.close()
            except Exception as e_summary:
                print(f"[SUMMARY] Error generating summary page: {e_summary}")
                
        # APPEND GLOBAL MSKU SUMMARY (if requested)
        if include_global_msku:
            try:
                # Fetch barang khusus SKUs
                try:
                    bk_data = await supabase_fetch("GET", "sku_barang_khusus?select=sku")
                    barang_khusus_skus = {str(item['sku']).strip().upper() for item in bk_data} if isinstance(bk_data, list) else set()
                except Exception as e_bk:
                    print(f"Failed to fetch barang_khusus: {e_bk}")
                    barang_khusus_skus = set()
                    
                global_msku_pdf = generate_global_msku_summary_pdf(awb_to_items, unique_matched_ids, rak_map, label_cfg, barang_khusus_skus)
                if len(global_msku_pdf) > 0:
                    master_doc.insert_pdf(global_msku_pdf)
                global_msku_pdf.close()
            except Exception as e_msku:
                print(f"[GLOBAL_MSKU] Error generating global msku page: {e_msku}")

        output_stream = io.BytesIO()
        master_doc.save(output_stream)
        master_doc.close()
        output_stream.seek(0)
        result_pdf_bytes = output_stream.read()
        pdf_base64 = base64.b64encode(result_pdf_bytes).decode('utf-8')
        
        # ===== BACKUP FILES =====
        try:
            # 1. GENERATE FILTERED EXCEL FOR PACKING LIST (Only matched rows)
            filtered_excel_bytes = None
            try:
                # Logic: Keep row if 'ID Pesanan' OR 'AWB' match match_awbs set
                print(f"[PROCESS] Filtering Excel for backup. Matched IDs: {len(matched_awbs)}")
                
                def is_matched_for_filter(row):
                    # Check ID Pesanan
                    if has_id_pesanan:
                        val = row.get('ID Pesanan', '')
                        norm = normalize_awb(val)
                        if norm and norm in matched_awbs: return True
                    # Check AWB as fallback
                    if has_awb:
                        val = row.get('AWB/No. Tracking', '')
                        norm = normalize_awb(val)
                        if norm and norm in matched_awbs: return True
                    return False

                # Filter the ORIGINAL DataFrame (df)
                filtered_df = df[df.apply(is_matched_for_filter, axis=1)].copy() # Copy to avoid SettingWithCopy
                
                if not filtered_df.empty:
                    # Force ID columns to string to prevent scientific notation in Excel
                    if has_id_pesanan:
                        filtered_df['ID Pesanan'] = filtered_df['ID Pesanan'].astype(str)
                    if has_awb:
                        filtered_df['AWB/No. Tracking'] = filtered_df['AWB/No. Tracking'].astype(str)

                    f_buffer = io.BytesIO()
                    # Use xlsxwriter engine
                    with pd.ExcelWriter(f_buffer, engine='xlsxwriter') as writer:
                        filtered_df.to_excel(writer, index=False)
                        # Auto-adjust columns width
                        worksheet = writer.sheets['Sheet1']
                        for i, col in enumerate(filtered_df.columns):
                            max_len = max(filtered_df[col].astype(str).map(len).max(), len(col)) + 2
                            worksheet.set_column(i, i, max_len)
                            
                    filtered_excel_bytes = f_buffer.getvalue()
                    print(f"[PROCESS] Filtered Excel created: {len(filtered_df)} rows (from {len(df)})")
                else:
                    print("[PROCESS] Filtered Excel is empty, skipping.")
            except Exception as fe:
                print(f"[PROCESS] Failed to filter Excel: {fe}")

            # 2. Simpan combined PDF asli ke bytes
            combined_pdf_stream = io.BytesIO()
            combined_pdf_doc.save(combined_pdf_stream)
            combined_pdf_doc.close()
            combined_pdf_stream.seek(0)
            combined_pdf_bytes = combined_pdf_stream.read()
            
            # Panggil save_backup dengan filtered data
            save_backup(
                excel_filename=excel_filename,
                pdf_filenames=pdf_filenames,
                excel_data=excel_content,
                pdf_data=combined_pdf_bytes,
                result_pdf=result_pdf_bytes,
                filtered_excel_data=filtered_excel_bytes
            )
        except Exception as backup_error:
            print(f"[BACKUP] Warning: Backup failed but process continues: {backup_error}")

        # Return JSON with stats
        # Create matched list with AWB info
        # FIX: Resolve ID Pesanan correctly. matched_awbs contains KEYS (can be ID or AWB).
        matched_with_awb = []
        processed_ids = set() # To avoid duplicates if both ID and AWB matched for same order
        
        for key in matched_awbs:
            real_id = None
            real_awb = None
            
            # Scenario 1: Key IS ID Pesanan
            if key in id_to_awb_mapping:
                real_id = key
                real_awb = id_to_awb_mapping[key]
            
            # Scenario 2: Key IS AWB (Reverse lookup)
            elif key in awb_to_id_mapping:
                real_id = awb_to_id_mapping[key]
                real_awb = key # The key itself is the AWB
            
            # Scenario 3: Key matches but isolated (maybe logic edge case)
            else:
                # If identifier was ID but had no AWB?
                # Check if it's in awb_to_items keys?
                if key in awb_to_items:
                    # It's a valid key pointing to items.
                    # Assume it's ID if not in reverse map?
                    real_id = key
                    real_awb = ''
            
            if real_id and real_id not in processed_ids:
                matched_with_awb.append({
                    'id_pesanan': real_id,
                    'awb': real_awb
                })
                processed_ids.add(real_id)
        
        # Create unmatched excel list with AWB info
        unmatched_excel_with_awb = []
        for id_pesanan in unmatched_excel:
            awb = id_to_awb_mapping.get(id_pesanan, '')
            unmatched_excel_with_awb.append({
                'id_pesanan': id_pesanan,
                'awb': awb
            })
        
        return JSONResponse({
            "success": True,
            "timestamp": datetime.now().isoformat(),
            "pdf_base64": pdf_base64,
            "stats": {
                "excel_filename": excel_filename,
                "pdf_filenames": pdf_filenames,
                "total_excel_awb": len(excel_awbs),
                "total_pdf_pages": len(all_pages),
                "matched_count": matched_count, # FIX: Use calculated unique count (deduplicated)
                "duplicate_count": duplicate_count,  # Duplikat asli (halaman tidak berurutan)
                "duplicate_awbs": duplicate_awbs,
                "continuation_count": len(continuation_pages),  # Pretelan (halaman lanjutan)
                "continuation_pages": continuation_pages,
                "unmatched_excel_count": len(unmatched_excel), # Jumlah order di Excel yang tidak ada di PDF
                "unmatched_pdf_count": len(unmatched_pdf),
                "matched_awbs": list(unique_matched_ids), # Use unique IDs list
                "matched_with_awb": matched_with_awb,
                "unmatched_excel_awbs": list(unmatched_excel),
                "unmatched_excel_with_awb": unmatched_excel_with_awb,
                "unmatched_excel_with_awb": unmatched_excel_with_awb,
                "unmatched_pdf_awbs": list(unmatched_pdf),
                "all_excel_awbs": list(excel_awbs_raw_set),
                "id_to_awb_mapping": id_to_awb_mapping
            }
        })

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/history/{id}")
async def delete_history(id: str, username: Optional[str] = None):
    """
    Delete history record and associated files.
    """
    try:
        # 1. Get record details first to find filenames
        data = await supabase_fetch("GET", f"label_process_history?id=eq.{id}&select=*")
        if not data:
            raise HTTPException(status_code=404, detail="History not found")
        
        record = data[0]
        
        # Security validation for User delete
        if username:
            if record.get('username') != username:
                raise HTTPException(status_code=403, detail="Akses ditolak: Anda bukan pemroses data ini.")
            
            created_at_str = record.get('created_at')
            if created_at_str:
                try:
                    if created_at_str.endswith('Z'):
                        created_at_str = created_at_str.replace('Z', '+00:00')
                    from datetime import datetime, timezone
                    record_time = datetime.fromisoformat(created_at_str)
                    now_time = datetime.now(timezone.utc)
                    diff_minutes = (now_time - record_time).total_seconds() / 60
                    
                    if diff_minutes > 60:
                        raise HTTPException(status_code=403, detail="Akses ditolak: Waktu penghapusan (60 menit) telah habis.")
                except Exception as parse_err:
                    print(f"Error parsing date {created_at_str}: {parse_err}")
                    raise HTTPException(status_code=500, detail="Gagal memverifikasi waktu data.")
        excel_filename = record.get('excel_filename')
        # pdf_filenames might be a list or string depending on how it was saved. In recent implementation it's a list.
        pdf_filenames = record.get('pdf_filenames', []) 
        created_at = record.get('created_at') # date string for folder structure

        # 2. Delete Physical Files (Backup Folder)
        # FIX: Gunakan find_backup_folder() untuk menemukan folder backup yang tepat
        # Logika lama salah karena mencari di subfolder tanggal (YYYY-MM-DD) yang tidak ada di struktur backup.
        if created_at and excel_filename:
            try:
                # Ambil nama PDF pertama sebagai identifier batch
                first_pdf_name = None
                if isinstance(pdf_filenames, list) and pdf_filenames:
                    first_pdf_name = pdf_filenames[0].replace('.pdf', '').replace('.PDF', '')
                elif isinstance(pdf_filenames, str) and pdf_filenames:
                    first_pdf_name = pdf_filenames.replace('.pdf', '').replace('.PDF', '')
                
                print(f"[DELETE] Searching backup folder: excel={excel_filename}, pdf={first_pdf_name}")
                
                # Gunakan find_backup_folder (mendukung 3 strategi pencocokan)
                backup_folder = find_backup_folder(
                    date_str=created_at,
                    excel_filename=excel_filename,
                    required_pdf_name=first_pdf_name
                )
                
                if backup_folder and backup_folder.exists():
                    shutil.rmtree(backup_folder)
                    print(f"[DELETE] ✅ Backup folder deleted: {backup_folder}")
                else:
                    print(f"[DELETE] ⚠️ Backup folder tidak ditemukan (mungkin sudah dihapus atau > 7 hari)")
                    
            except Exception as e:
                print(f"[DELETE] Error deleting backup folder: {e}")
                # Lanjutkan delete DB meski folder fisik gagal dihapus

        # 3. Delete from Supabase tables
        # A. Delete processed items via matched_awbs (lebih aman: tidak hapus batch lain)
        # FIX: Dulu delete semua processed_items by excel_filename, bisa hapus data batch lain
        matched_awbs_list = record.get('matched_awbs', [])
        if matched_awbs_list and isinstance(matched_awbs_list, list) and len(matched_awbs_list) > 0:
            print(f"[DELETE] Cascading delete for {len(matched_awbs_list)} matched AWBs")
            chunk_size = 30
            for i in range(0, len(matched_awbs_list), chunk_size):
                chunk = matched_awbs_list[i:i + chunk_size]
                awb_filter = ','.join([urllib.parse.quote(str(a)) for a in chunk])
                await supabase_fetch("DELETE", f"processed_items?order_id=in.({awb_filter})")
        elif excel_filename:
            # Fallback jika matched_awbs tidak tersimpan (data lama)
            print(f"[DELETE] Cascading delete fallback by excel_filename: {excel_filename}")
            await supabase_fetch("DELETE", f"processed_items?excel_filename=eq.{urllib.parse.quote(excel_filename)}")

        # B. Delete the history record itself
        await supabase_fetch("DELETE", f"label_process_history?id=eq.{id}")
        
        return {"success": True, "message": "History deleted"}

    except Exception as e:
        print(f"Delete History Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class SabotageManipulateRequest(BaseModel):
    target_folder: str
    first_pdf_name: str
    first_excel_name: str
    second_pdf_name: str
    second_excel_name: str
    backup_folder: str = ""

@app.post("/api/sabotage-manipulate")
async def sabotage_manipulate_local_files(data: SabotageManipulateRequest):
    try:
        folder = Path(data.target_folder)
        if not folder.exists() or not folder.is_dir():
            raise HTTPException(status_code=400, detail=f"Folder tidak ditemukan: {data.target_folder}")

        def resolve_file(fname: str) -> Path:
            if not fname: return Path("")
            p = folder / fname
            if p.exists(): return p
            dl = Path.home() / "Downloads" / fname
            if dl.exists(): return dl
            return p

        # Paths
        first_pdf = resolve_file(data.first_pdf_name)
        first_excel = resolve_file(data.first_excel_name)
        second_pdf = resolve_file(data.second_pdf_name)
        second_excel = resolve_file(data.second_excel_name)

        # Create Backup Folder
        if data.backup_folder and data.backup_folder.strip():
            backup_folder = Path(data.backup_folder.strip())
        else:
            backup_folder = Path(r"\\desktop-noq4lsr\Public\SCRIPT\testi\shipping-label-customizer 5\test")

            
        backup_folder.mkdir(parents=True, exist_ok=True)

        logs = []

        # 1. Backup Second Files
        if data.second_pdf_name and second_pdf.exists() and second_pdf.is_file():
            shutil.copy2(second_pdf, backup_folder / data.second_pdf_name)
            logs.append(f"Backed up {data.second_pdf_name}")
        
        if data.second_excel_name and second_excel.exists() and second_excel.is_file():
            shutil.copy2(second_excel, backup_folder / data.second_excel_name)
            logs.append(f"Backed up {data.second_excel_name}")

        # 2. Overwrite Second Files with First Files
        if data.first_pdf_name and first_pdf.exists() and first_pdf.is_file():
            shutil.copy2(first_pdf, second_pdf)
            logs.append(f"Overwrote {data.second_pdf_name} with {data.first_pdf_name}")
        elif data.first_pdf_name:
            logs.append(f"WARNING: File {data.first_pdf_name} tidak ditemukan di folder!")

        if data.first_excel_name and first_excel.exists() and first_excel.is_file():
            if data.second_excel_name:
                shutil.copy2(first_excel, second_excel)
                logs.append(f"Overwrote {data.second_excel_name} with {data.first_excel_name}")
        elif data.first_excel_name:
            logs.append(f"WARNING: File {data.first_excel_name} tidak ditemukan di folder!")

        return {"success": True, "logs": logs}

    except HTTPException:
        raise
    except Exception as e:
        print(f"[SABOTAGE MANIPULATE] Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/health")
async def health_check():
    return {"status": "ok", "message": "Backend is running"}




@app.post("/toolkit/generate-packing-list")
async def toolkit_generate_packing_list(
    excel_file: UploadFile = File(...),
    pdf_files: list[UploadFile] = File(...)
):
    try:
        content = await excel_file.read()
        df = pd.read_excel(io.BytesIO(content))
        
        # Normalize columns like process_labels
        col_mapping = {
            'AWB/No. Tracking': ['AWB/No. Tracking', 'AWB', 'No. Tracking', 'Tracking Number', 'Resi', 'No Resi', 'No. Resi'],
            'ID Pesanan': ['ID Pesanan', 'Order ID', 'No. Pesanan', 'Nomor Pesanan'],
            'MSKU': ['MSKU', 'SKU', 'Nama SKU', 'Product SKU', 'Master SKU', 'Nomor Referensi SKU', 'SKU Induk', 'Parent SKU'],
            'Jumlah': ['Jumlah', 'Qty', 'Quantity', 'QTY']
        }
        
        for target_col, alternatives in col_mapping.items():
            if target_col not in df.columns:
                for alt in alternatives:
                    if alt in df.columns:
                        df = df.rename(columns={alt: target_col})
                        break
                        
        # Pastikan kolom string
        if 'AWB/No. Tracking' in df.columns:
            df['AWB/No. Tracking'] = df['AWB/No. Tracking'].astype(str).apply(normalize_awb)
        if 'ID Pesanan' in df.columns:
            df['ID Pesanan'] = df['ID Pesanan'].astype(str).apply(normalize_awb)
            
        final_ids = []
        for pdf_file in pdf_files:
            pdf_content = await pdf_file.read()
            doc = fitz.open(stream=pdf_content, filetype="pdf")
            for page_num in range(len(doc)):
                page = doc[page_num]
                text_instances = page.get_text("dict")["blocks"]
                full_text = " ".join([span["text"] for block in text_instances if "lines" in block for line in block["lines"] for span in line["spans"]])
                
                is_packing_list = 'PACKING LIST / BATCH' in full_text.upper() or 'LIST ID PESANAN' in full_text.upper()
                if is_packing_list:
                    continue
                
                awbs = extract_all_awb_candidates(full_text)
                orders = extract_order_ids(full_text)
                
                if awbs:
                    final_ids.extend([normalize_awb(a) for a in awbs if normalize_awb(a)])
                if orders:
                    final_ids.extend([normalize_awb(o) for o in orders if normalize_awb(o)])
            doc.close()
            
        if not final_ids:
            raise HTTPException(status_code=400, detail="Tidak ada resi/pesanan ditemukan di PDF")
            
        df_matched = df.copy()
        mask = pd.Series(False, index=df.index)
        
        if 'AWB/No. Tracking' in df.columns:
            mask = mask | df['AWB/No. Tracking'].isin(final_ids)
            if 'ID Pesanan' in df.columns:
                mask = mask | (df['AWB/No. Tracking'].isna() & df['ID Pesanan'].isin(final_ids))
                mask = mask | (df['AWB/No. Tracking'] == 'NAN') & df['ID Pesanan'].isin(final_ids)
                mask = mask | (df['AWB/No. Tracking'] == '') & df['ID Pesanan'].isin(final_ids)
        elif 'ID Pesanan' in df.columns:
            mask = mask | df['ID Pesanan'].isin(final_ids)
            
        df_matched = df_matched[mask]
        
        if df_matched.empty:
            raise HTTPException(status_code=400, detail="Tidak ada data Excel yang cocok dengan PDF")
            
        col_msku = 'MSKU' if 'MSKU' in df_matched.columns else 'SKU'
        col_qty = 'Jumlah' if 'Jumlah' in df_matched.columns else 'QTY'
        col_id = 'ID Pesanan' if 'ID Pesanan' in df_matched.columns else 'NO. PESANAN'
        col_notes = 'Catatan Pembeli' if 'Catatan Pembeli' in df_matched.columns else None
        
        if col_msku not in df_matched.columns:
            df_matched[col_msku] = 'Unknown'
        if col_qty not in df_matched.columns:
            df_matched[col_qty] = 1
        if col_id not in df_matched.columns:
            df_matched[col_id] = ''
            
        # 1. Fetch Mapping ID & Rak
        rak_map = {}
        try:
            mappings = await get_sku_mappings()
            rak_map = {m['sku'].strip().upper(): {"rak": m.get('rak', ''), "id": m.get('id', '')} for m in mappings}
        except Exception as e:
            print(f"[PACKING LIST] Failed to fetch mappings: {e}")
            
        results = []
        for msku, group in df_matched.groupby(col_msku):
            total_qty = group[col_qty].sum()
            
            # Sub-group by Order ID to aggregate Qty per Order
            order_groups = group.groupby(col_id)[col_qty].sum()
            
            # Format: ID(Qty) sorted by ID
            detail_strs = []
            for oid, qty in sorted(order_groups.items()):
                 detail_strs.append(f"{oid}({int(qty)})")
            
            joined_details = '\n'.join(detail_strs)
            
            # Get Catatan Pembeli (Buyer Notes)
            notes_list = []
            if col_notes:
                for note in group[col_notes]:
                    if pd.notna(note) and str(note).strip():
                        notes_list.append(str(note).strip())
            
            # Format: SKU\nCatatan Pembeli:notes (or - if empty)
            if notes_list:
                unique_notes = list(set(notes_list))
                notes_str = '; '.join(unique_notes)
                sku_with_notes = f"{msku}\nCatatan Pembeli:{notes_str}"
            else:
                sku_with_notes = f"{msku}\nCatatan Pembeli:-"

            # Get ID Rak
            sku_upper = str(msku).strip().upper()
            rak_info = rak_map.get(sku_upper, {"rak": "", "id": ""})
            id_val = rak_info.get('id', '')
            rak_val = rak_info.get('rak', '')
            display_id = id_val if id_val else (rak_val if rak_val else '-')

            results.append({
                'ID': display_id,
                'SKU': sku_with_notes,
                'QTY': int(total_qty),
                'NO. PESANAN': joined_details,
                '_raw_sku': sku_upper
            })
            
        final_df = pd.DataFrame(results)
        
        # Sort by priority and QTY descending
        # Fetch Priority Bottom List
        priority_skus = set()
        try:
            p_data = await supabase_fetch("GET", "sku_priority_bottom?select=sku")
            priority_skus = {p['sku'].strip().upper() for p in p_data}
        except Exception:
            pass

        def get_priority(sku):
            return 1 if sku in priority_skus else 0

        if not final_df.empty:
            final_df['priority'] = final_df['_raw_sku'].apply(get_priority)
            final_df = final_df.sort_values(by=['priority', 'QTY'], ascending=[True, False])
            final_df = final_df.drop(columns=['priority', '_raw_sku'])
        
        output = io.BytesIO()
        # Fetch Formatting Rules
        format_colors = []
        format_styles = []
        try:
             format_colors = await supabase_fetch("GET", "sku_formatting_colors?select=keyword,color_code")
             format_styles = await supabase_fetch("GET", "sku_formatting_styles?select=keyword,font_size,is_bold")
             print(f"[PACKING LIST] Loaded {len(format_colors)} color rules, {len(format_styles)} style rules.")
        except Exception as e:
             print(f"[PACKING LIST] Failed to load formatting rules: {e}")

        # Fetch Column Settings
        column_settings = {}
        try:
            cs_data = await supabase_fetch("GET", "sku_column_settings?select=*")
            for cs in cs_data:
                column_settings[cs['column_name']] = cs
            print(f"[PACKING LIST] Loaded {len(column_settings)} column settings.")
        except Exception as e:
            print(f"[PACKING LIST] Failed to load column settings: {e}")

        # Default column settings if not in DB
        default_columns = {
            'ID': {'column_width': 9.64, 'font_size': 16, 'font_name': 'Rockwell', 'is_bold': False, 'text_align': 'center'},
            'SKU': {'column_width': 40.55, 'font_size': 16, 'font_name': 'Rockwell', 'is_bold': False, 'text_align': 'left'},
            'QTY': {'column_width': 15.82, 'font_size': 20, 'font_name': 'Rockwell', 'is_bold': True, 'text_align': 'center'},
            'NO. PESANAN': {'column_width': 29.91, 'font_size': 12, 'font_name': 'Rockwell', 'is_bold': False, 'text_align': 'center'}
        }

        # Merge with DB settings
        for col_name, defaults in default_columns.items():
            if col_name not in column_settings:
                column_settings[col_name] = defaults

        # Determine output stream
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            workbook = writer.book
            worksheet = workbook.add_worksheet('Sheet1')
            
            # --- DEFINE FORMATS ---
            # Header Format (Fixed)
            header_format = workbook.add_format({
                'font_name': 'Times New Roman', 'font_size': 20, 'bold': True,
                'align': 'center', 'valign': 'vcenter', 'bg_color': '#666666',
                'font_color': 'white', 'border': 1, 'text_wrap': True
            })
            
            # Dynamic Column Formats
            col_formats = {}
            col_order = ['ID', 'SKU', 'QTY', 'NO. PESANAN']
            for col_name in col_order:
                cs = column_settings.get(col_name, default_columns.get(col_name, {}))
                fmt = workbook.add_format({
                    'font_name': cs.get('font_name', 'Rockwell'),
                    'font_size': cs.get('font_size', 16),
                    'bold': cs.get('is_bold', False),
                    'align': cs.get('text_align', 'center'),
                    'valign': 'vcenter',
                    'border': 1,
                    'text_wrap': True
                })
                col_formats[col_name] = fmt

            # Assign specific format references for easier code use
            fmt_id = col_formats.get('ID')
            fmt_sku_base = col_formats.get('SKU')
            fmt_qty = col_formats.get('QTY')
            fmt_order = col_formats.get('NO. PESANAN')

            # Dynamic Color Formats Cache
            # Map: color_code -> Format Object (Base SKU + Color)
            color_formats = {} 
            for rule in format_colors:
                cc = rule['color_code']
                if cc not in color_formats:
                    cs = column_settings.get('SKU', default_columns.get('SKU', {}))
                    f = workbook.add_format({
                        'font_name': cs.get('font_name', 'Rockwell'),
                        'font_size': cs.get('font_size', 16),
                        'align': cs.get('text_align', 'left'),
                        'valign': 'vcenter',
                        'border': 1,
                        'text_wrap': True
                    })
                    f.set_font_color(cc)
                    color_formats[cc] = f

            # Dynamic Style Formats Cache
            # Map: (size, bold) -> Format Object (Just Font Metrics)
            style_formats = {}
            for rule in format_styles:
                key = (rule['font_size'], rule['is_bold'])
                if key not in style_formats:
                    # For rich string segments, we just define font props used in write_rich_string
                    # These segment formats do NOT need alignment/border usually, just font.
                    cs = column_settings.get('SKU', default_columns.get('SKU', {}))
                    f = workbook.add_format({'font_name': cs.get('font_name', 'Rockwell')})
                    if rule['font_size']: f.set_font_size(rule['font_size'])
                    if rule['is_bold']: f.set_bold()
                    style_formats[key] = f

            # --- WRITE HEADER ---
            headers = ['ID', 'SKU', 'QTY', 'NO. PESANAN']
            for col_num, header in enumerate(headers):
                worksheet.write(0, col_num, header, header_format)
            
            # Set Column Widths from Settings
            col_letters = ['A', 'B', 'C', 'D']
            for i, col_name in enumerate(col_order):
                cs = column_settings.get(col_name, default_columns.get(col_name, {}))
                width = cs.get('column_width', 20)
                worksheet.set_column(f'{col_letters[i]}:{col_letters[i]}', width)

            # --- WRITE DATA ---
            # Output DF: ID, SKU, QTY, NO. PESANAN
            # Iterating manually to apply rich text
            
            current_row = 1
            for index, row in final_df.iterrows():
                # 1. ID
                worksheet.write(current_row, 0, row['ID'], fmt_id)
                
                # 2. SKU (Rich Text Logic)
                sku_text = str(row['SKU'])
                
                # A. Determine Base Cell Color (from Color Rules)
                # Check formatting_colors
                # Find first matching rule? Or specific priority?
                # User: "jika ada sku ... mengandung kata BLUE maka ... teksnya warna merah"
                base_format = fmt_sku_base
                matched_color = None
                
                for rule in format_colors:
                    if rule['keyword'].upper() in sku_text.upper():
                        cc = rule['color_code']
                        if cc in color_formats:
                            base_format = color_formats[cc]
                            matched_color = cc
                        break # First match wins for cell color
                
                # B. Determine Rich Text Segments (from Style Rules)
                # This is complex if multiple keywords exist.
                # Simplification: Find ALL style keywords present.
                # Flatten the string into [ (text, format), (text, format), ... ]
                
                segments = []
                
                # Identify styled tokens
                # We need a list of (start_index, end_index, format_obj)
                style_matches = []
                for rule in format_styles:
                    kw = rule['keyword']
                    if not kw: continue
                    # Find all occurrences
                    start = 0
                    while True:
                        idx = sku_text.upper().find(kw.upper(), start)
                        if idx == -1: break
                        
                        # Store match
                        k_fmt = style_formats.get((rule['font_size'], rule['is_bold']))
                        style_matches.append({
                            'start': idx,
                            'end': idx + len(kw),
                            'fmt': k_fmt,
                            'priority': len(kw) # Longer keywords priority?
                        })
                        start = idx + 1
                
                # If no style matches, simple write
                if not style_matches:
                    worksheet.write(current_row, 1, sku_text, base_format)
                else:
                    # Sort matches by start position
                    style_matches.sort(key=lambda x: x['start'])
                    
                    # Handle overlaps? Simplest: First come first serve or verify no overlap.
                    # We will construct segments linearly.
                    
                    final_segments = []
                    last_pos = 0
                    
                    # Deduplicate/Resolve overlaps linearly
                    for match in style_matches:
                        if match['start'] < last_pos:
                            continue # Skip overlapping
                        
                        # Text before match - MUST include format for proper font!
                        if match['start'] > last_pos:
                            final_segments.append(base_format)  # Apply base format to non-styled text
                            final_segments.append(sku_text[last_pos:match['start']])
                        
                        # The Match - Add Format Object then Text
                        final_segments.append(match['fmt'])
                        final_segments.append(sku_text[match['start']:match['end']])
                        
                        last_pos = match['end']
                    
                    # Remaining text after last match - MUST include format!
                    if last_pos < len(sku_text):
                        final_segments.append(base_format)  # Apply base format to remaining text
                        final_segments.append(sku_text[last_pos:])
                    
                    # Write Rich String
                    # Args: row, col, *segments, cell_format (cell_format applies to cell properties like border)
                    # Each text segment should be preceded by its format
                    
                    try:
                        worksheet.write_rich_string(current_row, 1, *final_segments, base_format)
                    except Exception as e:
                        print(f"Rich Text Error: {e}")
                        worksheet.write(current_row, 1, sku_text, base_format)

                # 3. QTY
                worksheet.write(current_row, 2, row['QTY'], fmt_qty)
                
                # 4. NO PESANAN
                worksheet.write(current_row, 3, row['NO. PESANAN'], fmt_order)
                
                # Set Row Height (auto or fixed?)
                # user didn't ask, but multiline needs height. 
                # Excel usually auto-fits height if text wrap is on, but xlsxwriter sometimes needs explicit or separate step.
                # We'll leave default for now (xlsxwriter auto height is implied effectively if valid).
                # Actually, set row height to allow multi-line visibility if needed?
                # Using text_wrap=True usually handles it in Excel viewer.
                
                current_row += 1
                
            # Set Row Height for Header (Auto or fixed? Size 20 font needs space)
            worksheet.set_row(0, 30)
            
            # PAGE SETUP (from VBA specification + user top margin)
            # Margins: left/right 0.25", top 1.8" (pas untuk QR+teks header, jarak kecil ke tabel), bottom 0.75"
            worksheet.set_margins(left=0.25, right=0.25, top=1.8, bottom=0.75)
            
            # Print Title Rows: Row 1 repeats on each page
            worksheet.repeat_rows(0, 0)
            
            # Print Area: A1:D{lastRow}
            last_row = len(final_df) + 1  # +1 because header is row 1
            worksheet.print_area(f'A1:D{last_row}')

            # --- SHEET 2: UNIQUE ORDER IDs ---
            # Extract unique order IDs from NO. PESANAN column (column D)
            # Each cell may have multiple lines like: "582160320679544823(2)\n582160347278706463(1)..."
            # We split by newlines, extract ID before "(", and deduplicate
            
            unique_order_ids = set()
            for val in final_df['NO. PESANAN']:
                if pd.isna(val): continue
                lines = str(val).split('\n')
                for line in lines:
                    line = line.strip()
                    if not line: continue
                    # Remove quantity suffix like "(2)" or "(1)"
                    # Pattern: ID(qty) -> extract just ID
                    if '(' in line:
                        order_id = line.split('(')[0].strip()
                    else:
                        order_id = line.strip()
                    if order_id:
                        unique_order_ids.add(order_id)
            
            # Create Sheet2 and write unique IDs
            sheet2 = workbook.add_worksheet('Sheet2')
            sheet2.write(0, 0, 'ID Pesanan')
            # Write starting from A2 (row index 1)
            row_idx = 1
            for order_id in sorted(unique_order_ids):
                sheet2.write(row_idx, 0, order_id)
                row_idx += 1
            
            print(f"[PACKING LIST] Sheet2 created with {len(unique_order_ids)} unique order IDs.")

            # --- SHEET 3: UNIQUE AWBS ---
            # Cari kolom AWB di DataFrame asli
            col_awb = next((c for c in df.columns if 'AWB' in str(c).upper() or 'TRACKING' in str(c).upper() or 'RESI' in str(c).upper()), None)
            if col_awb:
                unique_awbs = set()
                for val in df[col_awb]:
                    if pd.notna(val):
                        awb = str(val).strip()
                        if awb:
                            unique_awbs.add(awb)
                
                sheet3 = workbook.add_worksheet('Sheet3')
                sheet3.write(0, 0, 'AWB/No. Tracking')
                row_idx = 1
                for awb in sorted(unique_awbs):
                    sheet3.write(row_idx, 0, awb)
                    row_idx += 1
                print(f"[PACKING LIST] Sheet3 created with {len(unique_awbs)} unique AWBs.")
            else:
                print(f"[PACKING LIST] Sheet3 skipped, AWB/No. Tracking column not found.")

            # 1. Generate QR Image locally (no internet needed)
            qr_file_path = None
            try:
                import qrcode
                # Use pdf_name for QR data (without extension)
                qr_data = (pdf_files[0].filename if pdf_files else excel_file.filename).replace('.pdf', '').replace('.xlsx', '').replace('.xls', '')
                print(f"[PACKING LIST] Generating QR locally for: {qr_data}")

                qr = qrcode.QRCode(
                    version=None,
                    error_correction=qrcode.constants.ERROR_CORRECT_M,
                    box_size=3,
                    border=1,
                )
                qr.add_data(qr_data)
                qr.make(fit=True)

                img = qr.make_image(fill_color="black", back_color="white")

                # Save to temp file
                temp_filename = f"temp_qr_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.png"
                img.save(temp_filename)
                qr_file_path = temp_filename
                print(f"[PACKING LIST] QR saved to: {temp_filename}")

            except Exception as e:
                print(f"[PACKING LIST] QR Generation Error: {e}")

            # 2. Set Header: &[Picture]        nama file PDF (tanpa extension)
            # Use pdf_name if provided, otherwise fallback to excel_file.filename
            header_name = pdf_files[0].filename if pdf_files else excel_file.filename
            header_display = header_name.replace('.pdf', '').replace('.xlsx', '').replace('.xls', '')
            if qr_file_path:
                # QR di baris pertama, teks nama di baris kedua (center, di bawah QR)
                header_str = f'&C&G\n&"Courier New,Bold"&16{header_display}'
                worksheet.set_header(header_str, {'image_center': qr_file_path, 'header_margin': 0.5})
            else:
                # Fallback text only if QR fails
                header_str = f'&C&"Courier New,Bold"&16[NO QR] {header_display}'
                worksheet.set_header(header_str, {'header_margin': 0.5})
            
            # 3. Set Footer: Page &[Page] of &[Pages] (Arial Black, Size 28)
            footer_str = '&C&"Arial Black,Regular"&28Page &P of &N'
            worksheet.set_footer(footer_str)
            
        output.seek(0)
        
        # Cleanup temp QR file
        if qr_file_path and os.path.exists(qr_file_path):
            try:
                os.remove(qr_file_path)
            except:
                pass
        
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="Packing List - {header_display}.xlsx"'
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"Toolkit Generate Packing List Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))




# --- SKU VIP (>50K) ROUTES ---

@app.get("/settings/sku-vip-50k")
async def get_sku_vip_50k():
    try:
        data = await supabase_fetch("GET", "sku_vip_50k?select=id,sku&order=created_at.desc")
        return data if data else []
    except Exception as e:
        print(f"Error getting SKU VIP 50K: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/settings/sku-vip-50k")
async def add_sku_vip_50k(request: Request):
    try:
        body = await request.json()
        sku = body.get("sku")
        if not sku:
            raise HTTPException(status_code=400, detail="SKU wajib diisi")
            
        data = [{"sku": sku}]
        await supabase_fetch("POST", "sku_vip_50k", data=data, headers={"Prefer": "resolution=ignore-duplicates"})
        return {"success": True}
    except Exception as e:
        print(f"Error adding SKU VIP 50K: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/settings/sku-vip-50k/{sku}")
async def delete_sku_vip_50k(sku: str):
    try:
        await supabase_fetch("DELETE", f"sku_vip_50k?sku=eq.{sku}")
        return {"success": True}
    except Exception as e:
        print(f"Error deleting SKU VIP 50K: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/settings/sku-vip-50k/bulk-delete")
async def bulk_delete_sku_vip_50k(request: Request):
    try:
        body = await request.json()
        ids = body.get("ids", [])
        if not ids:
             raise HTTPException(status_code=400, detail="Tidak ada SKU yang dipilih")
             
        for sku in ids:
             await supabase_fetch("DELETE", f"sku_vip_50k?sku=eq.{sku}")
             
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/settings/sku-vip-50k/export")
async def export_sku_vip_50k(request: Request):
    try:
        body = await request.json()
        ids = body.get("ids", [])
        
        if not ids:
            data = await supabase_fetch("GET", "sku_vip_50k?select=sku&order=created_at.desc")
        else:
            in_clause = ",".join(f"%22{sku}%22" for sku in ids)
            data = await supabase_fetch("GET", f"sku_vip_50k?sku=in.({in_clause})&select=sku")
            
        df = pd.DataFrame(data if data else [])
        if df.empty:
            df = pd.DataFrame(columns=["sku"])
            
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='SKU_VIP_50K')
        
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Export_SKU_VIP_50K.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/settings/import-sku-vip-50k")
async def import_sku_vip_50k(file: UploadFile = File(...)):
    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content), dtype=str)
        
        if 'MSKU' not in df.columns and 'SKU' not in df.columns:
            raise HTTPException(status_code=400, detail="Kolom 'MSKU' atau 'SKU' tidak ditemukan")
            
        col_sku = 'MSKU' if 'MSKU' in df.columns else 'SKU'
        
        skus = set()
        for _, row in df.iterrows():
            sku = str(row[col_sku]).strip()
            if sku and sku not in ('nan', 'None', '-'):
                skus.add(sku)
                
        if not skus:
             return {"success": True, "count": 0}
             
        to_import = [{"sku": s} for s in skus]
        
        # Batch insert
        chunk_size = 100
        for i in range(0, len(to_import), chunk_size):
            chunk = to_import[i:i + chunk_size]
            await supabase_fetch(
                "POST", "sku_vip_50k",
                data=chunk,
                headers={"Prefer": "resolution=ignore-duplicates"}
            )
            
        return {"success": True, "count": len(skus)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/toolkit/orderan-kilat-50k")
async def process_orderan_kilat_50k(file: UploadFile = File(...)):
    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False)
        
        if len(df) == 0:
            raise HTTPException(status_code=400, detail="File Excel kosong")
            
        col_id_pesanan = next((c for c in df.columns if 'ID PESANAN' in str(c).upper() or 'NO. PESANAN' in str(c).upper() or 'AWB' in str(c).upper()), None)
        col_msku = next((c for c in df.columns if 'MSKU' in str(c).upper()), None)
        
        if not col_id_pesanan or not col_msku:
            raise HTTPException(status_code=400, detail="Kolom 'ID Pesanan' atau 'MSKU' tidak ditemukan")
            
        # Ambil data SKU VIP 50K dari Supabase
        vip_data = await supabase_fetch("GET", "sku_vip_50k?select=sku")
        vip_skus = {e['sku'] for e in vip_data} if vip_data else set()
        
        if not vip_skus:
            raise HTTPException(status_code=404, detail="Data SKU VIP (>50K) masih kosong di database")
            
        import numpy as np
        df_ffill = df.copy()
        df_ffill[col_id_pesanan] = df_ffill[col_id_pesanan].replace('', np.nan)
        df_ffill[col_id_pesanan] = df_ffill[col_id_pesanan].ffill()
        
        vip_order_ids = set()
        for idx, row in df_ffill.iterrows():
            msku = str(row[col_msku]).strip()
            order_id = str(row[col_id_pesanan]).strip()
            if msku in vip_skus:
                vip_order_ids.add(order_id)
                
        if not vip_order_ids:
            raise HTTPException(status_code=404, detail="Tidak ditemukan resi dengan SKU VIP (>50K) di file ini")
            
        df_filtered = df_ffill[df_ffill[col_id_pesanan].isin(vip_order_ids)]
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df_filtered.to_excel(writer, index=False, sheet_name='Packing List')
            
            workbook  = writer.book
            worksheet = writer.sheets['Packing List']
            
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D9D9D9',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            cell_format_center = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'text_wrap': True
            })
            
            cell_format_left = workbook.add_format({
                'align': 'left',
                'valign': 'vcenter',
                'border': 1,
                'text_wrap': True
            })
            
            for col_num, value in enumerate(df_filtered.columns.values):
                worksheet.write(0, col_num, value, header_format)
            
            worksheet.set_column('A:A', 15)
            worksheet.set_column('B:B', 40)
            worksheet.set_column('C:C', 10)
            worksheet.set_column('D:D', 30)
            
            for row_num in range(len(df_filtered)):
                worksheet.write(row_num + 1, 0, df_filtered.iloc[row_num, 0], cell_format_center)
                worksheet.write(row_num + 1, 1, df_filtered.iloc[row_num, 1], cell_format_left)
                worksheet.write(row_num + 1, 2, df_filtered.iloc[row_num, 2], cell_format_center)
                worksheet.write(row_num + 1, 3, df_filtered.iloc[row_num, 3], cell_format_left)

        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Hasil_Orderan_Kilat_50K.xlsx"}
        )
        
    except HTTPException as he:
        raise he
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# --- ADMIN TABLE CLEANER ROUTES ---

@app.get("/admin/table-sizes")
async def get_table_sizes():
    """Get all public table sizes via Supabase RPC get_table_sizes()"""
    try:
        data = await supabase_fetch("POST", "rpc/get_table_sizes", data={})
        return data if data else []
    except Exception as e:
        print(f"Error getting table sizes: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/admin/table-cleanup")
async def table_cleanup(request: Request):
    """Delete rows from a table filtered by date range"""
    try:
        body = await request.json()
        table_name = body.get("table_name")
        date_column = body.get("date_column")
        date_from = body.get("date_from")  # YYYY-MM-DD
        date_to = body.get("date_to")      # YYYY-MM-DD (optional, same as date_from if single date)

        if not table_name or not date_column or not date_from:
            raise HTTPException(status_code=400, detail="table_name, date_column, dan date_from wajib diisi")

        # If date_to not provided, use date_from (single date mode)
        if not date_to:
            date_to = date_from

        # Build the filter query string for Supabase REST API
        # Delete rows where date_column >= date_from 00:00:00 AND date_column <= date_to 23:59:59
        start_dt = f"{date_from}T00:00:00"
        end_dt = f"{date_to}T23:59:59"

        # Use Supabase REST API DELETE with filters
        query = f"{table_name}?{date_column}=gte.{start_dt}&{date_column}=lte.{end_dt}"

        headers = {"Prefer": "return=representation"}
        result = await supabase_fetch("DELETE", query, headers=headers)

        deleted_count = len(result) if result else 0
        print(f"[TableCleaner] Deleted {deleted_count} rows from {table_name} ({date_from} to {date_to})")

        return {"success": True, "deleted_count": deleted_count, "table": table_name}

    except HTTPException as he:
        raise he
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/admin/table-count")
async def table_count_by_date(request: Request):
    """Count rows in a table filtered by date range (preview before delete)"""
    try:
        body = await request.json()
        table_name = body.get("table_name")
        date_column = body.get("date_column")
        date_from = body.get("date_from")
        date_to = body.get("date_to")

        if not table_name or not date_column or not date_from:
            raise HTTPException(status_code=400, detail="table_name, date_column, dan date_from wajib diisi")

        if not date_to:
            date_to = date_from

        start_dt = f"{date_from}T00:00:00"
        end_dt = f"{date_to}T23:59:59"

        # Use Supabase REST API with HEAD + Prefer: count=exact to get count
        query = f"{table_name}?{date_column}=gte.{start_dt}&{date_column}=lte.{end_dt}&select=id"
        headers = {"Prefer": "count=exact", "Range-Unit": "items", "Range": "0-0"}
        result = await supabase_fetch("GET", query, headers=headers)

        # The count comes from the content-range header, but since we're using
        # our helper, we'll just count the returned items with a limit
        # Alternative: just count returned array length with a high limit
        query_count = f"{table_name}?{date_column}=gte.{start_dt}&{date_column}=lte.{end_dt}&select=id&limit=100000"
        result = await supabase_fetch("GET", query_count)
        count = len(result) if result else 0

        return {"count": count, "table": table_name, "date_from": date_from, "date_to": date_to}

    except HTTPException as he:
        raise he
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/admin/table-columns")
async def get_table_columns(request: Request):
    """Get column names for a specific table by fetching 1 row"""
    try:
        body = await request.json()
        table_name = body.get("table_name")

        if not table_name:
            raise HTTPException(status_code=400, detail="table_name wajib diisi")

        # Fetch 1 row to get column names
        result = await supabase_fetch("GET", f"{table_name}?select=*&limit=1")
        if result and len(result) > 0:
            columns = list(result[0].keys())
            # Filter to only date/timestamp columns (heuristic: contains 'at', 'date', 'time')
            date_columns = [c for c in columns if any(kw in c.lower() for kw in ['_at', 'date', 'time', 'created', 'updated'])]
            return {"columns": columns, "date_columns": date_columns}
        else:
            return {"columns": [], "date_columns": []}

    except Exception as e:
        print(f"Error getting columns for {body.get('table_name', '?')}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)


